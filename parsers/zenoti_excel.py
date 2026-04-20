"""
Zenoti Employee KPI Excel Parser
Parses .xlsx files from Zenoti "Employee KPI" reports.

Schema (CORRECTED — header row is Row 3, not Row 4):
  Row 1: "Employee KPI Fantastic Sams"
  Row 2: "Employee KPI From : 01 Mar 2026 To : 31 Mar 2026 By : Karissa Grant..."
  Row 3: Column headers (Employee Name, Invoice Count, etc.)
  Row 4+: Stylist data rows
  Last data row: "Total" summary row (stop here)

Location extraction:
  Scan data rows (Row 4+) for any employee_name containing 'mgr'.
  Standard format  → "Andover mgr"        → location = "Andover"
  Non-standard     → "888-40098-F Sams Roseville, MN\\_mgr mgr"
                                           → location = "Roseville"

Karissa's authoritative formulas (applied here):
  guest_count = invoice_count   (override Zenoti column D — use column B)
  total_sales = service_sales + product_sales
  service_net = total_sales - product_sales   (== service_sales, explicit)
  ppg_net     = product_sales / guest_count   (0 when guest_count == 0)
"""

import re
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl

from config.locations import normalize_location


class ZenotiExcelParser:
    """Parse a single Zenoti Employee KPI .xlsx file."""

    # Zero-based column indices matching Row 3 headers
    COL_EMPLOYEE_NAME    = 0   # A — "Employee Name"
    COL_INVOICE_COUNT    = 1   # B — "Invoice Count"
    COL_AVG_INVOICE      = 2   # C — "Avg Invoice Value"
    COL_GUEST_COUNT      = 3   # D — "Guest Count"  (intentionally ignored per Karissa)
    COL_SERVICE_SALES    = 4   # E — "Service Sales"
    COL_AVG_SERVICE      = 5   # F — "Avg Service Value"
    COL_PRODUCT_SALES    = 6   # G — "Product Sales"
    COL_AVG_PRODUCT      = 7   # H — "Avg Product Value"

    # Row numbers (1-based, as openpyxl uses)
    HEADER_ROW  = 3   # CRITICAL CORRECTION: spec says Row 4, actual files use Row 3
    DATA_START  = 4   # First stylist data row

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet = self.wb.active

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def parse(self) -> Dict:
        """
        Parse the full file.

        Returns:
            {
              'location':      str,
              'location_id':   str,   # snake_case canonical name
              'location_name': str,   # display name (same as location)
              'period':        {'start_date': 'YYYY-MM-DD', 'end_date': 'YYYY-MM-DD'},
              'pos_system':    'zenoti',
              'stylists':      [ <stylist dict>, ... ]
            }
        """
        location    = self.extract_location()
        location_id = location.lower().replace(" ", "_")
        period      = self.extract_period()
        stylists    = self._parse_stylists(location, location_id, period)

        return {
            "location":      location,
            "location_id":   location_id,
            "location_name": location,
            "period":        period,
            "pos_system":    "zenoti",
            "stylists":      stylists,
        }

    # ------------------------------------------------------------------
    # Extraction helpers
    # ------------------------------------------------------------------

    def extract_period(self) -> Dict[str, Optional[str]]:
        """
        Extract reporting period from Row 2.

        Expected pattern:
            "Employee KPI From : 01 Mar 2026 To : 31 Mar 2026 By : ..."
        """
        row2_value = self.sheet.cell(row=2, column=1).value or ""
        row2_str   = str(row2_value)

        match = re.search(
            r"From\s*:\s*(\d{2}\s+\w+\s+\d{4})\s+To\s*:\s*(\d{2}\s+\w+\s+\d{4})",
            row2_str,
            re.IGNORECASE,
        )

        if match:
            try:
                start_date = datetime.strptime(match.group(1), "%d %b %Y").strftime("%Y-%m-%d")
                end_date   = datetime.strptime(match.group(2), "%d %b %Y").strftime("%Y-%m-%d")
                return {"start_date": start_date, "end_date": end_date}
            except ValueError:
                pass

        return {"start_date": None, "end_date": None}

    def extract_location(self) -> str:
        """
        Scan data rows for the 'mgr' marker row and extract location name.

        Standard:     "Andover mgr"  → "Andover"
        Non-standard: "888-40098-F Sams Roseville, MN\\_mgr mgr"
                      → looks for "Sams <City>" pattern → "Roseville"
        """
        for row_idx in range(self.DATA_START, self.sheet.max_row + 1):
            cell_val = self.sheet.cell(row=row_idx, column=1).value
            if cell_val is None:
                continue

            name = str(cell_val).strip()

            if "mgr" not in name.lower():
                continue

            # Check Sams/Roseville format FIRST — it also ends with " mgr" so
            # must be caught before the standard check or the ID string gets
            # returned as the location name.
            # "888-40098-F Sams Roseville, MN\_mgr mgr" → "Roseville"
            if " Sams " in name or name.lower().startswith("sams"):
                sams_match = re.search(r"Sams\s+([A-Za-z\s]+?)(?:,|\s*_|\s*mgr)", name, re.IGNORECASE)
                if sams_match:
                    raw = sams_match.group(1).strip()
                    return normalize_location(raw)

            # Standard: ends with " mgr" — e.g. "Andover mgr"
            if name.lower().endswith(" mgr"):
                raw = re.sub(r"\s+mgr$", "", name, flags=re.IGNORECASE).strip()
                return normalize_location(raw)

        return "Unknown"

    # ------------------------------------------------------------------
    # Stylist parsing
    # ------------------------------------------------------------------

    def _safe_float(self, value) -> float:
        """Convert a cell value to float, stripping commas/$ signs."""
        if value is None:
            return 0.0
        try:
            return float(str(value).replace(",", "").replace("$", "").strip())
        except (ValueError, AttributeError):
            return 0.0

    def _parse_stylists(
        self,
        location: str,
        location_id: str,
        period: Dict,
    ) -> List[Dict]:
        """
        Parse all stylist rows between DATA_START and the "Total" row.

        Skips:
          - Empty employee name
          - Rows containing 'mgr' (manager marker rows)
          - The "Total" summary row
        """
        stylists = []

        for row_idx in range(self.DATA_START, self.sheet.max_row + 1):
            employee_name = self.sheet.cell(row=row_idx, column=1).value

            # Stop at the totals row
            if employee_name is not None and str(employee_name).strip() == "Total":
                break

            # Skip empty rows
            if not employee_name:
                continue

            name_str = str(employee_name).strip()

            # Skip manager marker rows
            if "mgr" in name_str.lower():
                continue

            # Extract raw values
            invoice_count = self._safe_float(self.sheet.cell(row=row_idx, column=self.COL_INVOICE_COUNT + 1).value)
            service_sales = self._safe_float(self.sheet.cell(row=row_idx, column=self.COL_SERVICE_SALES  + 1).value)
            product_sales = self._safe_float(self.sheet.cell(row=row_idx, column=self.COL_PRODUCT_SALES  + 1).value)

            # Skip rows that look completely empty (all zeros and a name we can't validate)
            if invoice_count == 0 and service_sales == 0 and product_sales == 0:
                continue

            # ---- Karissa's authoritative formulas ----
            guest_count = invoice_count                                        # Override column D
            total_sales = service_sales + product_sales
            service_net = total_sales - product_sales                          # == service_sales
            product_net = product_sales                                        # explicit alias
            ppg_net     = (product_sales / guest_count) if guest_count > 0 else 0.0
            avg_ticket  = (total_sales   / guest_count) if guest_count > 0 else 0.0
            product_pct = (product_net   / total_sales * 100) if total_sales > 0 else 0.0

            stylists.append({
                # Identity
                "location":      location,
                "location_id":   location_id,
                "location_name": location,
                "stylist_name":  name_str,
                "period_start":  period["start_date"],
                "period_end":    period["end_date"],
                "pos_system":    "zenoti",

                # Raw values preserved from Excel
                "invoice_count": int(invoice_count),
                "service_sales": round(service_sales, 2),
                "product_sales": round(product_sales, 2),

                # Derived per Karissa's formulas
                "guest_count":   int(guest_count),
                "total_sales":   round(total_sales, 2),
                "service_net":   round(service_net, 2),
                "product_net":   round(product_net, 2),
                "ppg_net":       round(ppg_net, 2),
                "avg_ticket":    round(avg_ticket, 2),
                "product_pct":   round(product_pct, 2),

                # Zenoti does not report PPH or productive hours
                "pph_net":          None,
                "productive_hours": None,
            })

        return stylists
