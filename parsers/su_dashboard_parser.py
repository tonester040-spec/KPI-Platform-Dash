"""
KPI Salon Ultimate (SU) dashboard parser — salon-grain LOCATIONS_DATA source.

SU's "FS Salon Dashboard" export is a REAL binary OLE2 .xls — NOT the HTML-disguised
.xls Zenoti ships, and NOT xlrd-readable (it's written by a non-Excel tool, so xlrd
2.0 raises CompDocError). Read path that IS proven to work: LibreOffice headless
converts .xls -> .xlsx, then openpyxl reads the single flat 'Worksheet' sheet
(label-in-column-A, ~137 rows).

This parser is the SU sibling of `pdf_hours_parser.parse_salon_summary`. It returns the
SAME field dict so `locations_grouper._build_row` emits the IDENTICAL 39-col
LOCATIONS_DATA contract — there is NO schema fork. SU's one extra field, `tips_total`
(Zenoti does not report tips), rides along in the parse dict; the 39-col row ignores it.

Karissa's locked SU rules (KARISSA_GOLDEN_RULES.md / CLAUDE.md):
  guest_count = "TOTAL Guests"  (Serviced + Retail Only — already summed on the export)
  service_net = "Total Service"     product_net = "Total Retail"
  total       = service_net + product_net   (== the printed "Total Revenue" line)
  color_pct   = color_net / service_net  (engine);  wax_pct/treatment_pct = count / guests
Schema `pph_net` = total / hours (engine), which intentionally DIFFERS from SU's printed
PPH (= service / hours = 53.36 for Apple Valley). We keep the engine definition so PPH is
comparable across the Zenoti and SU halves of the network. (SU's printed PPH/PPG/Avg
Ticket are still used as parse-time cross-checks below.)

The Service Categories table carries NEGATIVE adjustment buckets ("None", "Other") that
ARE included in its TOTALS row, so reconciliation sums ALL categories (incl the negatives)
and they must equal service_net. None/Other are enumerated for reconciliation ONLY — never
mapped to a schema bucket. "None" is matched as the literal string (guard vs Python None).

The Employee Summary block (per-stylist Net Service/Retail, reconciles to the TOTALS row)
is NOT consumed here — salon-grain needs no stylist split. It exists, which means SU could
feed STYLISTS_DATA salon-side WITHOUT the Tableau path; that is deferred, not built now.

VALIDATED: Apple Valley 05/17-05/23/2026 -> service 22810.65 / product 2219.25 /
total 25029.90 / guests 428 / prod hours 427.45 / color 6501.50 / wax 1587.20 (84) /
treatment 2652.50 (136) / tips 4242.02 / avg_ticket 58.48 / reconciliation gap 0.00.
"""
from __future__ import annotations

import glob
import hashlib
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone

# ─────────────────────────────────────────────────────────────────────────────
# SU location crosswalk — store-name substring -> canonical location_id.
# IDs match config/customers/karissa_001.json (the project-wide SU identifiers).
# Apple Valley's store name is "FS - Apple Valley Pilot Knob"; Lakeville/Farmington
# store strings are UNSEEN (same POS, lower risk) — substring match + the
# reconciliation net is the backstop when their files arrive.
# ─────────────────────────────────────────────────────────────────────────────
SU_LOCATIONS: dict[str, str] = {
    "Apple Valley": "z010",
    "Lakeville":    "su001",
    "Farmington":   "su002",
}

# Service-category buckets we expose + accepted header-name aliases. Mirrors the
# Zenoti sibling's alias map (Wax/Waxing). Every OTHER category in the table
# (Haircut, Style, Perm, None, Other, ...) is enumerated for reconciliation only.
_BUCKET_ALIASES = {
    "color": ("color", "colour"),
    "wax": ("wax", "waxing"),
    "treatment": ("treatment", "treatments"),
}


# ─────────────────────────────────────────────────────────────────────────────
# LibreOffice .xls -> .xlsx conversion (the only way to read SU's binary OLE2).
# ─────────────────────────────────────────────────────────────────────────────
def _find_soffice() -> str:
    """Locate the LibreOffice headless binary. Order: SOFFICE_PATH env, PATH,
    then the usual per-OS install locations. Raises if none found."""
    env = os.environ.get("SOFFICE_PATH")
    if env and os.path.exists(env):
        return env
    for name in ("soffice", "soffice.exe"):
        found = shutil.which(name)
        if found:
            return found
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/opt/libreoffice/program/soffice",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    raise RuntimeError(
        "LibreOffice 'soffice' not found — required to read SU binary .xls files. "
        "Install LibreOffice or set the SOFFICE_PATH environment variable."
    )


def _xls_to_xlsx(xls_path: str, outdir: str) -> str:
    """Convert a binary SU .xls to .xlsx via LibreOffice headless into `outdir`.

    soffice can print warnings to stderr and even exit non-zero while still
    emitting a valid file, so success is judged by the PRESENCE of an .xlsx in
    outdir — never by the exit code. A per-call UserInstallation profile (inside
    outdir) avoids the profile-lock contention that makes concurrent soffice
    runs silently no-op.
    """
    soffice = _find_soffice()
    profile_dir = os.path.join(outdir, "lo_profile")
    profile_uri = "file:///" + profile_dir.replace("\\", "/")
    cmd = [
        soffice, "--headless", "--norestore", "--invisible",
        f"-env:UserInstallation={profile_uri}",
        "--convert-to", "xlsx", "--outdir", outdir, xls_path,
    ]
    try:
        subprocess.run(cmd, capture_output=True, text=True, timeout=180)
    except subprocess.TimeoutExpired as e:  # pragma: no cover - environment dependent
        raise RuntimeError(f"LibreOffice conversion timed out for {xls_path!r}") from e

    expected = os.path.join(outdir, os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx")
    if os.path.exists(expected):
        return expected
    produced = [p for p in glob.glob(os.path.join(outdir, "*.xlsx"))]
    if produced:
        return produced[0]
    raise RuntimeError(
        f"LibreOffice produced no .xlsx for {xls_path!r}. Check that the file is a "
        f"valid SU dashboard export and that LibreOffice can open it."
    )


def _read_su_grid(xls_path: str) -> list[list]:
    """Return the SU dashboard as an in-memory grid (list of row lists).

    Converts to .xlsx in a throwaway temp dir, reads every cell with openpyxl
    (data_only -> cached values for any formula cells), then lets the temp dir
    clean up. Row index in the returned list is 0-based; grid[r][c] mirrors the
    sheet's row r+1, column c+1.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read SU dashboards (pip install openpyxl).") from e
    import openpyxl as _ox

    if not os.path.exists(xls_path):
        raise FileNotFoundError(xls_path)

    tmp = tempfile.mkdtemp(prefix="su_xls_")
    try:
        xlsx = _xls_to_xlsx(xls_path, tmp)
        wb = _ox.load_workbook(xlsx, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# Pure parsing helpers over the grid (no I/O — unit-testable in isolation).
# ─────────────────────────────────────────────────────────────────────────────
def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _num(v):
    """Coerce a cell to float; None if blank/non-numeric. Strips $ and commas."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_hours(v):
    """'427h 27m' -> 427.45 (decimal hours). Already-numeric passes through.
    None if blank. Either of the h/m components may be absent."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v)
    hm = re.search(r"(\d+)\s*h", s, re.I)
    mm = re.search(r"(\d+)\s*m", s, re.I)
    if not hm and not mm:
        return None
    h = int(hm.group(1)) if hm else 0
    m = int(mm.group(1)) if mm else 0
    return round(h + m / 60.0, 2)


def _first_value(grid, label):
    """Value in column B for the first row whose column-A label == `label`
    (exact, case-insensitive after strip). None if the label is absent."""
    want = label.strip().lower()
    for row in grid:
        if row and _norm(row[0]).lower() == want and len(row) > 1:
            return row[1]
    return None


def _parse_period(grid):
    """'05/17/2026 - 05/23/2026' -> ('2026-05-17', '2026-05-23', raw). All None
    if the Report period row is missing/unparseable."""
    raw = _first_value(grid, "Report period:")
    if raw is None:
        return None, None, _norm(raw) or None
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4}).*?(\d{1,2})/(\d{1,2})/(\d{4})", str(raw))
    if not m:
        return None, None, str(raw).strip()
    a = f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    b = f"{int(m.group(6)):04d}-{int(m.group(4)):02d}-{int(m.group(5)):02d}"
    return a, b, str(raw).strip()


def _parse_service_categories(grid):
    """Service Categories table -> bucket nets/counts + full enumeration for
    reconciliation. Locates the 'Service Categories' header, resolves the Qty and
    Sales columns from the header text, then reads each row until 'TOTALS'.

    Returns: color_net, wax_net/wax_count, treatment_net/treatment_count,
    categories_net_sum (ALL rows incl negative None/Other), categories_found,
    and totals_row_sales (the printed TOTALS Sales cell, for a 2nd cross-check)."""
    header_idx = None
    qty_col = 1   # default col B
    sales_col = 4  # default col E
    for i, row in enumerate(grid):
        if row and _norm(row[0]).lower() == "service categories":
            header_idx = i
            for c, cell in enumerate(row):
                t = _norm(cell).lower()
                if t == "qty":
                    qty_col = c
                elif t == "sales":
                    sales_col = c
            break
    if header_idx is None:
        return {
            "color_net": None, "wax_net": None, "wax_count": None,
            "treatment_net": None, "treatment_count": None,
            "categories_net_sum": None, "categories_found": [], "totals_row_sales": None,
        }

    def _alias_bucket(name):
        nl = name.lower()
        for bucket, aliases in _BUCKET_ALIASES.items():
            if nl in aliases:
                return bucket
        return None

    found = {}                                   # label -> (qty, net) for every category row
    buckets = {"color": [], "wax": [], "treatment": []}
    totals_row_sales = None
    for row in grid[header_idx + 1:]:
        name = _norm(row[0]) if row else ""
        if not name:
            break  # blank row ends the table
        sales = _num(row[sales_col]) if len(row) > sales_col else None
        qty = _num(row[qty_col]) if len(row) > qty_col else None
        if name.lower() == "totals":
            totals_row_sales = sales
            break
        if sales is None:
            continue
        found[name] = (qty, sales)              # enumerate ALL (incl "None"/"Other" negatives)
        b = _alias_bucket(name)
        if b is not None:
            buckets[b].append((qty, sales))

    def _agg_net(pairs):
        return round(sum(n for _, n in pairs), 2) if pairs else None

    def _agg_qty(pairs):
        qs = [q for q, _ in pairs if q is not None]
        return int(round(sum(qs))) if qs else None

    return {
        "color_net": _agg_net(buckets["color"]),
        "wax_net": _agg_net(buckets["wax"]),
        "wax_count": _agg_qty(buckets["wax"]),
        "treatment_net": _agg_net(buckets["treatment"]),
        "treatment_count": _agg_qty(buckets["treatment"]),
        "categories_net_sum": round(sum(n for _, n in found.values()), 2) if found else None,
        "categories_found": sorted(found.keys()),
        "totals_row_sales": totals_row_sales,
    }


def resolve_su_location(store_name):
    """('FS - Apple Valley Pilot Knob') -> ('z010', 'Apple Valley'). Substring
    match against SU_LOCATIONS. Raises ValueError on no match (fail loud — an
    unknown store name means the crosswalk needs a new entry)."""
    s = (store_name or "").lower()
    for canonical, lid in SU_LOCATIONS.items():
        if canonical.lower() in s:
            return lid, canonical
    raise ValueError(
        f"Unrecognized SU store name {store_name!r}; known SU locations: {list(SU_LOCATIONS)}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────
def parse_su_dashboard(xls_path, *, strict=True):
    """Parse ONE SU dashboard .xls -> salon-grain field dict (sibling of
    pdf_hours_parser.parse_salon_summary, plus tips_total / store_name / period).

    I/O wrapper: convert + read the grid, then delegate to the pure `_parse_grid`.

    Reconciliation (the safety net): the sum of EVERY Service Categories row
    (including the negative None/Other adjustment buckets) must equal service_net
    within $0.01. When `strict` (default), a mismatch raises ValueError naming the
    gap and the categories found — that gap is the signature of a silently dropped
    or renamed category. With strict=False the dict is returned with
    service_mix_reconciled=False and a flag string instead (Zenoti-sibling parity).
    """
    return _parse_grid(_read_su_grid(xls_path), strict=strict)


def _parse_grid(grid, *, strict=True):
    """Pure SU-dashboard parse over an in-memory grid (no I/O — synthetic-grid
    testable). See parse_su_dashboard for the reconciliation contract."""
    store_name = _norm(_first_value(grid, "Store name:")) or None
    period_start, period_end, period_label = _parse_period(grid)

    service_net = _num(_first_value(grid, "Total Service"))
    product_net = _num(_first_value(grid, "Total Retail"))
    total_revenue_printed = _num(_first_value(grid, "Total Revenue"))
    total_sales_net = round(service_net + product_net, 2) if (service_net is not None and product_net is not None) else None

    guest_count = _num(_first_value(grid, "TOTAL Guests"))
    guest_count = int(guest_count) if guest_count is not None else None
    tips_total = _num(_first_value(grid, "Tips"))
    productive_hours = _parse_hours(_first_value(grid, "Production Hours"))

    cats = _parse_service_categories(grid)

    # ── Reconciliation: all categories (incl negatives) must equal service_net ──
    cats_sum = cats["categories_net_sum"]
    reconciled = (
        cats_sum is not None and service_net is not None and abs(cats_sum - service_net) <= 0.01
    )
    gap = round(service_net - cats_sum, 2) if (cats_sum is not None and service_net is not None) else None
    flags = []
    if not reconciled and cats_sum is not None and service_net is not None:
        msg = (
            f"su_service_mix_unreconciled [{store_name}]: categories_sum={cats_sum} "
            f"service_net={service_net} gap={gap} found={cats['categories_found']}"
        )
        flags.append(msg)
        if strict:
            raise ValueError(msg)

    # Secondary cross-checks (non-fatal — recorded as flags only):
    if (cats["totals_row_sales"] is not None and service_net is not None
            and abs(cats["totals_row_sales"] - service_net) > 0.01):
        flags.append(
            f"su_totals_row_mismatch [{store_name}]: TOTALS Sales={cats['totals_row_sales']} "
            f"service_net={service_net}"
        )
    if (total_revenue_printed is not None and total_sales_net is not None
            and abs(total_revenue_printed - total_sales_net) > 0.01):
        flags.append(
            f"su_total_revenue_mismatch [{store_name}]: printed Total Revenue="
            f"{total_revenue_printed} service+product={total_sales_net}"
        )

    location_id, canonical = (None, None)
    if store_name:
        location_id, canonical = resolve_su_location(store_name)

    return {
        "store_name": store_name,
        "location_id": location_id,
        "location_name_canonical": canonical,
        "period_start": period_start,
        "period_end": period_end,
        "period_label": period_label,
        "service_net": service_net,
        "product_net": product_net,
        "total_sales_net": total_sales_net,
        "color_net": cats["color_net"], "color_count": None,  # SU has no schema color_count column
        "wax_net": cats["wax_net"], "wax_count": cats["wax_count"],
        "treatment_net": cats["treatment_net"], "treatment_count": cats["treatment_count"],
        "guest_count": guest_count,
        "unique_guests": None,        # SU does not distinguish unique guests
        "productive_hours": productive_hours,
        "tips_total": tips_total,     # SU advantage — not in the 39-col row; rides along here
        "service_mix_reconciled": reconciled,
        "service_mix_gap": gap,
        "categories_found": cats["categories_found"],
        "flags": flags,
    }


def build_su_location_row(xls_path, *, location_id=None, period_start=None, period_end=None,
                          period_type="weekly", period_label=None, goals=None):
    """Parse ONE SU dashboard .xls -> one schema-ordered LOCATIONS_DATA row via the
    SHARED locations_grouper._build_row engine (the identical 39-col contract; no
    schema fork). The file carries its own date range, so period_* default to the
    parsed values when not supplied. location_id defaults to the store-name match."""
    # Imported here so the engine stays the single source of truth and a missing
    # pandas (only needed by the grouper's DataFrame path) never blocks this import.
    from parsers.locations_grouper import _build_row, CROSSWALK, _lookup_goal

    su = parse_su_dashboard(xls_path)
    if location_id is None:
        location_id = su["location_id"]
    metrics = {
        "service_net": su["service_net"],
        "product_net": su["product_net"],
        "total_sales_net": su["total_sales_net"],
        "color_net": su["color_net"],
        "wax_net": su["wax_net"],
        "wax_count": su["wax_count"],
        "treatment_net": su["treatment_net"],
        "treatment_count": su["treatment_count"],
    }
    canonical = CROSSWALK.get(location_id, {}).get("location_name_canonical")
    extract_hash = hashlib.sha256(open(xls_path, "rb").read()).hexdigest()  # hash the SOURCE .xls
    return _build_row(
        location_id=location_id,
        period_start=period_start or su["period_start"],
        period_end=period_end or su["period_end"],
        period_type=period_type,
        period_label=period_label or su["period_label"],
        metrics=metrics,
        guest_count=su["guest_count"],
        unique_guests=None,
        productive_hours=su["productive_hours"],
        monthly_goal=_lookup_goal(goals, location_id, canonical),
        extract_hash=extract_hash,
        generated_at=datetime.now(timezone.utc).isoformat(),
        extra_flags=None,
    )


if __name__ == "__main__":
    import argparse
    import json as _json

    ap = argparse.ArgumentParser(description="Parse one Salon Ultimate dashboard .xls.")
    ap.add_argument("xls", help="Path to the SU FS_Salon_Dashboard .xls")
    ap.add_argument("--row", action="store_true", help="Also print the assembled LOCATIONS_DATA row")
    args = ap.parse_args()

    parsed = parse_su_dashboard(args.xls)
    print(_json.dumps(parsed, indent=2))
    if args.row:
        from parsers import locations_grouper as L
        row = build_su_location_row(args.xls)
        print("\nLOCATIONS_DATA row:")
        for c in L.COLUMNS:
            print(f"  {c:<26} {row[c]}")
