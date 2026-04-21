"""
Salon Ultimate Stylist Tracking Excel Parser
Parses .xls files from Salon Ultimate "Stylist Tracking Report".

Schema:
  File extension : .xls  (legacy Excel format — requires LibreOffice conversion)
  Sheet name     : "Worksheet"
  Cell B1        : Store name  e.g. "FS - Apple Valley Pilot Knob"
  Cell B2        : Report period  e.g. "03/01/2026 - 03/31/2026"
  Row 5          : Column headers
  Row 6+         : Stylist data rows
  Last data row  : "Totals:" label in column A

Column mapping (Row 5 headers, 1-based):
  Col A (1)  : Stylist Name
  Col B (2)  : Service Sales  → service_net
  Col C (3)  : Retail Sales   → product_net
  Col D (4)  : Hours Worked   → productive_hours  (Totals row only)
  Col F (6)  : Serv $ Hour    → pph_net           (Totals row only)
  Col G (7)  : Retail Client  → retail_clients    (per-stylist guest component)
  Col H (8)  : Service Clients→ service_clients   (per-stylist guest component)

Excluded provider rows (skip, do not count as stylists):
  - "Totals:"        — summary row
  - "Booked Online"  — online booking phantom
  - "House _"        — house account prefix
  - "Salon Ultimate" — phantom phantom row

Karissa's authoritative formulas:
  guest_count  = service_clients + retail_clients
  total_sales  = service_net + product_net
  ppg_net      = product_net / guest_count       (0 when guest_count == 0)
  avg_ticket   = total_sales / guest_count       (0 when guest_count == 0)
  product_pct  = product_net / total_sales * 100 (0 when total_sales == 0)

pph_net and productive_hours are location-level metrics read from the Totals row only.
Per-stylist rows have pph_net: None and productive_hours: None.
"""

import logging
import os
import re
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl

from config.locations import normalize_location
from parsers.utils.xls_converter import convert_xls_to_xlsx, needs_conversion

logger = logging.getLogger(__name__)


class SalonUltimateExcelParser:
    """Parse a single Salon Ultimate Stylist Tracking .xls/.xlsx file."""

    SHEET_NAME  = "Worksheet"
    HEADER_ROW  = 5
    DATA_START  = 6

    # 1-based column indices (CORRECTED per spec)
    COL_STYLIST_NAME     = 1   # A — "Stylist Name"
    COL_SERVICE_NET      = 2   # B — "Service Sales"
    COL_PRODUCT_NET      = 3   # C — "Retail Sales"
    COL_HOURS_WORKED     = 4   # D — "Hours Worked"    (Totals row only)
    # Col E (5) not captured
    COL_PPH_NET          = 6   # F — "Serv $ Hour"     (Totals row only)
    COL_RETAIL_CLIENTS   = 7   # G — "Retail Client"
    COL_SERVICE_CLIENTS  = 8   # H — "Service Clients"

    # Provider name patterns to skip entirely
    SKIP_PATTERNS = ("Booked Online", "House _", "Salon Ultimate")

    def __init__(self, file_path: str):
        self.file_path       = file_path
        self._converted_path = None

        wb_path = self._resolve_workbook_path(file_path)
        self.wb = openpyxl.load_workbook(wb_path, data_only=True)

        # Resolve the active sheet — tolerate LibreOffice renames.
        # Salon Ultimate's original .xls is named "Worksheet", but after
        # LibreOffice conversion the sheet often takes the file's stem
        # (e.g. "stylist_tracking"). Fall back to the first sheet in that case.
        if self.SHEET_NAME in self.wb.sheetnames:
            self.sheet = self.wb[self.SHEET_NAME]
        else:
            self.sheet = self.wb.active
            logger.warning(
                "Expected sheet '%s' not found in %s — falling back to active sheet '%s'. "
                "Available sheets: %s",
                self.SHEET_NAME, wb_path, self.sheet.title, self.wb.sheetnames,
            )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def parse(self) -> Dict:
        """
        Parse the full file.

        Returns:
            {
              'location':       str,
              'location_id':    str,   # snake_case canonical name
              'location_name':  str,   # display name (same as location)
              'period':         {'start_date': 'YYYY-MM-DD', 'end_date': 'YYYY-MM-DD'},
              'pos_system':     'salon_ultimate',
              'location_totals':{'pph_net': float, 'productive_hours': float},
              'stylists':       [ <stylist dict>, ... ]
            }
        """
        location       = self.extract_location()
        location_id    = location.lower().replace(" ", "_")
        period         = self.extract_period()
        location_totals = self._extract_location_totals()
        stylists       = self._parse_stylists(location, location_id, period)

        return {
            "location":        location,
            "location_id":     location_id,
            "location_name":   location,
            "period":          period,
            "pos_system":      "salon_ultimate",
            "location_totals": location_totals,
            "stylists":        stylists,
        }

    # ------------------------------------------------------------------
    # .xls → .xlsx conversion
    # ------------------------------------------------------------------

    def _resolve_workbook_path(self, file_path: str) -> str:
        """
        Return the path to an .xlsx workbook, converting from .xls if necessary.

        Strategy:
          1. If the path already ends with .xlsx — use it directly.
          2. If a .xls, delegate to `parsers.utils.xls_converter.convert_xls_to_xlsx`.
             The utility auto-reuses an existing .xlsx sibling if present.

        The conversion path is recorded on `self._converted_path` so callers can
        (optionally) call `.cleanup()` to remove the generated file.
        """
        if not needs_conversion(file_path):
            return file_path

        xlsx_path = convert_xls_to_xlsx(file_path, reuse_existing=True)
        self._converted_path = str(xlsx_path)
        return str(xlsx_path)

    # ------------------------------------------------------------------
    # Extraction helpers
    # ------------------------------------------------------------------

    def extract_location(self) -> str:
        """
        Extract location from Cell B1.

        Examples:
          "FS - Apple Valley Pilot Knob"  → "Apple Valley"
          "FS - Prior Lake"               → "Prior Lake"
          "FS - Farmington"               → "Farmington"
          "FS - Lakeville"                → "Lakeville"

        Strategy: match "FS - <text>" then test against known location names.
        """
        b1 = self.sheet.cell(row=1, column=2).value or ""
        store_name = str(b1).strip()

        # Try known canonical names first (most reliable)
        for canonical in ["Apple Valley", "Prior Lake", "Farmington", "Lakeville"]:
            if canonical.lower() in store_name.lower():
                return canonical

        # Fallback: extract the segment after "FS - " and normalize
        match = re.search(r"FS\s*-\s*(.+)", store_name, re.IGNORECASE)
        if match:
            raw = match.group(1).strip()
            # Strip trailing descriptor words (e.g., "Pilot Knob", "MN")
            raw = re.split(r"\s+(Pilot|MN|Suite|#)", raw, maxsplit=1)[0].strip()
            return normalize_location(raw)

        return "Unknown"

    def extract_period(self) -> Dict[str, Optional[str]]:
        """
        Extract reporting period from Cell B2.

        Expected pattern: "03/01/2026 - 03/31/2026"
        """
        b2 = self.sheet.cell(row=2, column=2).value or ""
        period_str = str(b2).strip()

        match = re.search(
            r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})",
            period_str,
        )

        if match:
            try:
                start_date = datetime.strptime(match.group(1), "%m/%d/%Y").strftime("%Y-%m-%d")
                end_date   = datetime.strptime(match.group(2), "%m/%d/%Y").strftime("%Y-%m-%d")
                return {"start_date": start_date, "end_date": end_date}
            except ValueError:
                pass

        return {"start_date": None, "end_date": None}

    # ------------------------------------------------------------------
    # Stylist parsing
    # ------------------------------------------------------------------

    def _safe_float(self, value) -> float:
        """Convert a cell value to float, stripping commas and $ signs."""
        if value is None:
            return 0.0
        try:
            return float(str(value).replace(",", "").replace("$", "").strip())
        except (ValueError, AttributeError):
            return 0.0

    def _find_totals_row(self) -> Optional[int]:
        """Return the 1-based row number of the 'Totals:' row, or None."""
        for row_idx in range(self.DATA_START, self.sheet.max_row + 1):
            cell_val = self.sheet.cell(row=row_idx, column=self.COL_STYLIST_NAME).value
            if cell_val and "Totals:" in str(cell_val):
                return row_idx
        return None

    def _extract_location_totals(self) -> Dict:
        """
        Read location-level metrics from the 'Totals:' row.

        Returns dict with pph_net and productive_hours.
        Both default to 0.0 if Totals row is not found.
        """
        totals_row = self._find_totals_row()
        if totals_row is None:
            return {"pph_net": 0.0, "productive_hours": 0.0}

        productive_hours = self._safe_float(
            self.sheet.cell(row=totals_row, column=self.COL_HOURS_WORKED).value
        )
        pph_net = self._safe_float(
            self.sheet.cell(row=totals_row, column=self.COL_PPH_NET).value
        )

        return {
            "pph_net":          round(pph_net, 2),
            "productive_hours": round(productive_hours, 2),
        }

    def _should_skip(self, name: str) -> bool:
        """Return True if this provider row should be excluded."""
        for pattern in self.SKIP_PATTERNS:
            if pattern.lower() in name.lower():
                return True
        return False

    def _parse_stylists(
        self,
        location: str,
        location_id: str,
        period: Dict,
    ) -> List[Dict]:
        """
        Parse stylist rows from DATA_START up to (but not including) 'Totals:'.

        Skips:
          - Empty stylist name
          - "Totals:" row
          - Rows matching SKIP_PATTERNS (Booked Online, House _, Salon Ultimate)
          - Rows where all key metrics are zero (blank spacer rows)
        """
        totals_row = self._find_totals_row()
        end_row    = totals_row if totals_row else self.sheet.max_row + 1

        stylists = []

        for row_idx in range(self.DATA_START, end_row):
            stylist_name = self.sheet.cell(row=row_idx, column=self.COL_STYLIST_NAME).value

            # Skip empty
            if not stylist_name:
                continue

            name_str = str(stylist_name).strip()

            # Skip excluded provider patterns
            if self._should_skip(name_str):
                continue

            # Extract raw values (corrected column mapping)
            service_net      = self._safe_float(
                self.sheet.cell(row=row_idx, column=self.COL_SERVICE_NET).value
            )
            product_net      = self._safe_float(
                self.sheet.cell(row=row_idx, column=self.COL_PRODUCT_NET).value
            )
            retail_clients   = self._safe_float(
                self.sheet.cell(row=row_idx, column=self.COL_RETAIL_CLIENTS).value
            )
            service_clients  = self._safe_float(
                self.sheet.cell(row=row_idx, column=self.COL_SERVICE_CLIENTS).value
            )

            # Skip rows that are entirely zero (blank spacer rows)
            if (service_net == 0 and product_net == 0
                    and retail_clients == 0 and service_clients == 0):
                continue

            # ---- Karissa's authoritative formulas ----
            guest_count = service_clients + retail_clients
            total_sales = service_net + product_net
            ppg_net     = (product_net  / guest_count)  if guest_count  > 0 else 0.0
            avg_ticket  = (total_sales  / guest_count)  if guest_count  > 0 else 0.0
            product_pct = (product_net  / total_sales * 100) if total_sales > 0 else 0.0

            stylists.append({
                # Identity
                "location":      location,
                "location_id":   location_id,
                "location_name": location,
                "stylist_name":  name_str,
                "period_start":  period["start_date"],
                "period_end":    period["end_date"],
                "pos_system":    "salon_ultimate",

                # Raw values preserved from Excel
                "service_clients": int(service_clients),
                "retail_clients":  int(retail_clients),
                "service_net":     round(service_net, 2),
                "product_net":     round(product_net, 2),

                # Derived per Karissa's formulas
                "guest_count":  int(guest_count),
                "total_sales":  round(total_sales, 2),
                "ppg_net":      round(ppg_net, 2),
                "avg_ticket":   round(avg_ticket, 2),
                "product_pct":  round(product_pct, 2),

                # SU location-level metrics — None at stylist level (from location_totals)
                "pph_net":           None,
                "productive_hours":  None,
            })

        return stylists

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------

    def cleanup(self):
        """Remove temp .xlsx file created during .xls conversion (optional call)."""
        if self._converted_path and os.path.exists(self._converted_path):
            os.remove(self._converted_path)
            self._converted_path = None

    def __del__(self):
        """Best-effort cleanup on garbage collection."""
        try:
            self.cleanup()
        except Exception:
            pass
