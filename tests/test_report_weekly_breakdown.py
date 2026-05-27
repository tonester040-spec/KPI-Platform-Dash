"""
tests/test_report_weekly_breakdown.py
─────────────────────────────────────
Tests for the 'Weekly Breakdown' Excel sheet added in Phase 2
(Karissa Q7 view). Verifies sheet shape, headers, row counts, and the
network total footer.
"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl  # noqa: E402

from core.report_builder import _build_weekly_breakdown  # noqa: E402


def _mock_data() -> dict:
    """Mock data dict shape that build_report receives."""
    return {
        "network": {"week_ending": "2026-04-19"},
        "locations": [
            {
                "loc_name": "Andover FS",
                "hist": {
                    "weeks": ["2026-03-29", "2026-04-05", "2026-04-12", "2026-04-19"],
                    "total_sales": [7000, 6000, 6500, 5800],  # only April weeks count
                },
            },
            {
                "loc_name": "Blaine",
                "hist": {
                    "weeks": ["2026-04-05", "2026-04-12", "2026-04-19"],
                    "total_sales": [18000, 18500, 17000],
                },
            },
            {
                "loc_name": "Hudson",
                "hist": {
                    "weeks": ["2026-04-05"],   # only Week 1 so far
                    "total_sales": [13000],
                },
            },
        ],
    }


class TestWeeklyBreakdownSheet(unittest.TestCase):
    def test_sheet_created(self):
        wb = openpyxl.Workbook()
        # Remove default sheet so build creates Weekly Breakdown as first
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        self.assertIn("Weekly Breakdown", wb.sheetnames)

    def test_title_row_includes_current_month(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        ws = wb["Weekly Breakdown"]
        self.assertIn("2026-04", str(ws["A1"].value))

    def test_header_row_lists_each_week_plus_mtd(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        ws = wb["Weekly Breakdown"]
        # Header row is row 3 per _build_weekly_breakdown layout
        headers = [ws.cell(row=3, column=c).value for c in range(1, 10) if ws.cell(row=3, column=c).value]
        self.assertEqual(headers[0], "Location")
        self.assertEqual(headers[-1], "MTD Total")
        # Should have 3 week columns (max_weeks = 3 from Blaine/Andover)
        week_headers = [h for h in headers if h.startswith("Wk ")]
        self.assertEqual(len(week_headers), 3)
        self.assertTrue(week_headers[0].startswith("Wk 1"))
        self.assertIn("4/5", week_headers[0])  # 2026-04-05 -> 4/5

    def test_location_rows_in_alphabetical_order(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        ws = wb["Weekly Breakdown"]
        # Body rows start at row 4
        loc_col_values = [ws.cell(row=r, column=1).value for r in (4, 5, 6)]
        self.assertEqual(loc_col_values, ["Andover FS", "Blaine", "Hudson"])

    def test_missing_week_shows_dash(self):
        """Hudson only has Week 1 — Week 2 and Week 3 cells should be '—'."""
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        ws = wb["Weekly Breakdown"]
        # Hudson is row 6 (alphabetical: Andover=4, Blaine=5, Hudson=6)
        # Wk 1 = col 2, Wk 2 = col 3, Wk 3 = col 4, MTD = col 5
        self.assertEqual(ws.cell(row=6, column=2).value, 13000.0)
        self.assertEqual(ws.cell(row=6, column=3).value, "—")
        self.assertEqual(ws.cell(row=6, column=4).value, "—")
        self.assertEqual(ws.cell(row=6, column=5).value, 13000.0)  # MTD = Wk 1

    def test_mtd_total_column_for_each_location(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        ws = wb["Weekly Breakdown"]
        # MTD column = col 5 (Location + 3 weeks + MTD)
        # Andover April: 6000 + 6500 + 5800 = 18300 (only April weeks)
        self.assertAlmostEqual(ws.cell(row=4, column=5).value, 18300.0, places=2)
        # Blaine April: 18000 + 18500 + 17000 = 53500
        self.assertAlmostEqual(ws.cell(row=5, column=5).value, 53500.0, places=2)
        # Hudson April: only Week 1 = 13000
        self.assertAlmostEqual(ws.cell(row=6, column=5).value, 13000.0, places=2)

    def test_network_total_footer(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, _mock_data())
        ws = wb["Weekly Breakdown"]
        # Footer row is 1 row after last body row (which is row 6 = Hudson)
        footer_row = 8  # blank row 7, then footer at 8
        self.assertEqual(ws.cell(row=footer_row, column=1).value, "NETWORK TOTAL")
        # Wk 1 total: 6000 + 18000 + 13000 = 37000
        self.assertAlmostEqual(ws.cell(row=footer_row, column=2).value, 37000.0, places=2)
        # Wk 2 total: 6500 + 18500 + 0 (Hudson missing) = 25000
        self.assertAlmostEqual(ws.cell(row=footer_row, column=3).value, 25000.0, places=2)
        # Wk 3 total: 5800 + 17000 + 0 = 22800
        self.assertAlmostEqual(ws.cell(row=footer_row, column=4).value, 22800.0, places=2)
        # MTD total: 18300 + 53500 + 13000 = 84800
        self.assertAlmostEqual(ws.cell(row=footer_row, column=5).value, 84800.0, places=2)

    def test_empty_locations_writes_placeholder(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        _build_weekly_breakdown(wb, {"locations": []})
        ws = wb["Weekly Breakdown"]
        self.assertIn("No locations", str(ws["A1"].value))


if __name__ == "__main__":
    unittest.main()
