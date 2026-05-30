"""
Tests for core/report_generator — the Sheet-fed Karissa weekly Excel builder.

Style mirrors the Track A-D / history-store suites: plain script, `check()`, a
SKIP-gated real-file smoke. No network — the pure builder
``build_workbook_from_data`` takes plain dicts (the data_source CUMULATIVE_MTD /
DATA_MONTHLY / MONTHLY_GOALS shapes), so every layout assertion runs offline.

Covers:
  * single-week layout — 5 week tabs + YoY, 12 rows + Totals, header-name placement
  * HEADER DRIFT (the key test) — March-style and May-style header rows map to the
    SAME canonical fields, position-independently, with current vs 2025-reference
    'Guest Count'/'PPG'/'AT' disambiguated
  * cumulative weeks placed in their own tab, NOT summed
  * closed/zero week row PRESENT (not skipped)
  * provisional marker from data_complete_flag
  * YoY 2025 column blank when no prior-year rows; populated when present; goals fill
  * read-as-stored penny-exact placement
  * INTEGRATION (gated) — real Forest Lake parse -> snapshot -> generate -> penny-exact

Run:  python tests/test_report_generator.py
"""
from __future__ import annotations

import os
import shutil
import sys
import tempfile
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl  # noqa: E402
from core import report_generator as RG  # noqa: E402

FAILS: list[str] = []


def check(name, cond):
    print(("PASS" if cond else "FAIL"), "-", name)
    if not cond:
        FAILS.append(name)


def _fresh_dir():
    return tempfile.mkdtemp(prefix="kpi_report_")


ROSTER = ["Andover FS", "Blaine", "Crystal FS", "Elk River FS", "Forest Lake",
          "Prior Lake", "Hudson", "New Richmond", "Roseville", "Apple Valley",
          "Lakeville", "Farmington"]


def snap(loc, we, *, year_month="2026-05", total=10000.0, service=9000.0,
         product=1000.0, guests=200, prod_hours=250.0, ppg=5.0, pph=36.0,
         avg_ticket=50.0, product_pct=0.10, wax_count=20, wax=400.0, wax_pct=0.10,
         color=3000.0, color_pct=0.33, treat_count=30, treat=600.0, treat_pct=0.15,
         **over):
    """One CUMULATIVE_MTD snapshot dict (data_source shape). Values are arbitrary but
    distinct enough to catch mis-placement."""
    r = {
        "loc_name": loc, "year_month": year_month, "week_ending": we, "platform": "zenoti",
        "guests": guests, "total_sales": total, "service": service, "product": product,
        "product_pct": product_pct, "ppg": ppg, "pph": pph, "avg_ticket": avg_ticket,
        "prod_hours": prod_hours, "wax_count": wax_count, "wax": wax, "wax_pct": wax_pct,
        "color": color, "color_pct": color_pct, "treat_count": treat_count,
        "treat": treat, "treat_pct": treat_pct, "source": "test",
    }
    r.update(over)
    return r


def _field_cols(ws):
    """Header-name -> {field: 1-based col} for a week tab's top block (row 2)."""
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    return {field: col for col, field in RG.map_top_header_row(headers).items()}


# ─────────────────────────────────────────────────────────────────────────────
# Single-week layout + header-name placement.
# ─────────────────────────────────────────────────────────────────────────────
def t_single_week_layout():
    d = _fresh_dir()
    try:
        snaps = [snap(loc, "2026-05-03", guests=100 + i, total=10000.0 + i)
                 for i, loc in enumerate(ROSTER)]
        out = RG.build_workbook_from_data("2026-05", Path(d) / "w.xlsx",
                                          snapshots=snaps, roster=ROSTER)
        wb = openpyxl.load_workbook(out)
        check("5 week tabs + YoY in order",
              wb.sheetnames == RG.WEEK_TAB_NAMES + [RG.YOY_TAB_NAME])
        ws = wb[RG.WEEK_TAB_NAMES[0]]
        fc = _field_cols(ws)
        check("header 'Guest Count' resolves to guests field", "guests" in fc)
        check("header 'Total Sales Net' resolves to total_sales", "total_sales" in fc)
        # 12 location rows in roster order (rows 3..14).
        names = [ws.cell(row=3 + i, column=fc["loc_name"]).value for i in range(12)]
        check("12 location rows in roster order", names == ROSTER)
        # Forest Lake (idx 4) values land under the right headers.
        fl_row = 3 + ROSTER.index("Forest Lake")
        check("FL guests under 'Guest Count'", ws.cell(row=fl_row, column=fc["guests"]).value == 104)
        check("FL total under 'Total Sales Net'", ws.cell(row=fl_row, column=fc["total_sales"]).value == 10004.0)
        # Totals row (15).
        check("Totals label in col A row 15", ws.cell(row=15, column=fc["loc_name"]).value == "Totals")
        exp_guests = sum(100 + i for i in range(12))
        check("Totals guests = column sum", ws.cell(row=15, column=fc["guests"]).value == exp_guests)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# THE KEY TEST — header drift. March-style and May-style rows map identically.
# ─────────────────────────────────────────────────────────────────────────────
def t_header_drift_mapping():
    march = ["Name", "Guest Count", "Total Sales Net", "Service Net", "Product Net",
             "Product %", "PPG Net", "PPH Net", "Average Tkt", "Prod Hours",
             "Projection", "2025 Total Sales", " Guest Count", "2025PPG", "Diff",
             " AT", "Dif"]
    may = ["Name", "tickets ", "Guest Count", "Total Sales Net", "Service Net",
           "Product Net", "Product %", "PPG Net", "PPH Net", "Average Tkt",
           "Prod Hours", "Projection", "2025 Total Sales", " Guest Count", "PPG",
           "Diff", " AT", "Dif"]
    m = RG.map_top_header_row(march)
    y = RG.map_top_header_row(may)

    # March: positions are 1-based; current Guest Count at 2, ref Guest Count at 13.
    check("March guests @ col 2", m.get(2) == "guests")
    check("March total_sales @ col 3", m.get(3) == "total_sales")
    check("March projection @ col 11", m.get(11) == "projection")
    check("March py_total_sales @ col 12", m.get(12) == "py_total_sales")
    check("March py_guests @ col 13 (NOT guests)", m.get(13) == "py_guests")
    check("March py_ppg @ col 14", m.get(14) == "py_ppg")
    check("March diff_ppg @ col 15", m.get(15) == "diff_ppg")
    check("March py_avg_ticket @ col 16", m.get(16) == "py_avg_ticket")
    check("March diff_at @ col 17", m.get(17) == "diff_at")

    # May: 'tickets' at col 2 shifts everything +1; tickets maps to nothing.
    check("May tickets col 2 unmapped (noise)", 2 not in y)
    check("May guests @ col 3 (shifted)", y.get(3) == "guests")
    check("May total_sales @ col 4", y.get(4) == "total_sales")
    check("May py_total_sales @ col 13", y.get(13) == "py_total_sales")
    check("May py_guests @ col 14 (NOT guests)", y.get(14) == "py_guests")
    check("May py_ppg @ col 15", y.get(15) == "py_ppg")
    check("May diff_at @ col 18", y.get(18) == "diff_at")

    # Both styles resolve the SAME set of canonical fields.
    check("March & May map to identical field sets",
          set(m.values()) == set(y.values()))
    # The collision is genuinely disambiguated in both.
    check("guests != py_guests, both present (March)",
          "guests" in m.values() and "py_guests" in m.values())
    check("guests != py_guests, both present (May)",
          "guests" in y.values() and "py_guests" in y.values())


# ─────────────────────────────────────────────────────────────────────────────
# Cumulative weeks placed per tab, NOT summed.
# ─────────────────────────────────────────────────────────────────────────────
def t_cumulative_not_summed():
    d = _fresh_dir()
    try:
        # Forest Lake's MARCH-style cumulative progression (from the history docstring):
        # Wk1 8147 -> Wk2 15000 -> Wk3 22358. Each is already cumulative MTD.
        snaps = [
            snap("Forest Lake", "2026-05-03", total=8147.0),
            snap("Forest Lake", "2026-05-10", total=15000.0),
            snap("Forest Lake", "2026-05-17", total=22358.0),
        ]
        out = RG.build_workbook_from_data("2026-05", Path(d) / "w.xlsx",
                                          snapshots=snaps, roster=["Forest Lake"])
        wb = openpyxl.load_workbook(out)
        w1, w3 = wb[RG.WEEK_TAB_NAMES[0]], wb[RG.WEEK_TAB_NAMES[2]]
        c1, c3 = _field_cols(w1)["total_sales"], _field_cols(w3)["total_sales"]
        v1 = w1.cell(row=3, column=c1).value
        v3 = w3.cell(row=3, column=c3).value
        check("Week 1 shows its own cumulative (8147)", v1 == 8147.0)
        check("Week 3 shows its own cumulative (22358), NOT a sum", v3 == 22358.0)
        check("Week 3 is NOT 8147+15000+22358 (no double-count)", v3 != 45505.0)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# Closed/zero week row present (not skipped).
# ─────────────────────────────────────────────────────────────────────────────
def t_closed_week_zero_row():
    d = _fresh_dir()
    try:
        # Two weeks; in week 2 Hudson is a closed/holiday zero row.
        snaps = [snap(loc, "2026-05-03") for loc in ROSTER]
        snaps += [snap(loc, "2026-05-10") for loc in ROSTER if loc != "Hudson"]
        snaps.append(snap("Hudson", "2026-05-10", guests=0, total=0.0, service=0.0,
                          product=0.0, prod_hours=0.0, wax_count=0, wax=0.0,
                          color=0.0, treat_count=0, treat=0.0))
        out = RG.build_workbook_from_data("2026-05", Path(d) / "w.xlsx",
                                          snapshots=snaps, roster=ROSTER)
        wb = openpyxl.load_workbook(out)
        ws = wb[RG.WEEK_TAB_NAMES[1]]
        fc = _field_cols(ws)
        hud_row = 3 + ROSTER.index("Hudson")
        check("closed-week Hudson row PRESENT (name written)",
              ws.cell(row=hud_row, column=fc["loc_name"]).value == "Hudson")
        check("closed-week guests is 0 (zero row, not blank/skipped)",
              ws.cell(row=hud_row, column=fc["guests"]).value == 0)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# Provisional marker from data_complete_flag.
# ─────────────────────────────────────────────────────────────────────────────
def t_provisional_marker():
    d = _fresh_dir()
    try:
        snaps = [snap("Forest Lake", "2026-05-03", data_complete_flag=False)]
        out = RG.build_workbook_from_data("2026-05", Path(d) / "w.xlsx",
                                          snapshots=snaps, roster=["Forest Lake"])
        wb = openpyxl.load_workbook(out)
        ws = wb[RG.WEEK_TAB_NAMES[0]]
        name = ws.cell(row=3, column=1).value
        check("provisional row name gets '*' marker", name == "Forest Lake *")
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# YoY: blank without prior-year; populated with prior-year + goals.
# ─────────────────────────────────────────────────────────────────────────────
def t_yoy_blank_vs_populated():
    d = _fresh_dir()
    try:
        snaps = [snap("Forest Lake", "2026-05-31", total=33191.0)]
        # Case A — no prior-year rows: 2025 column blank, MTD still shown.
        out = RG.build_workbook_from_data("2026-05", Path(d) / "a.xlsx",
                                          snapshots=snaps, roster=["Forest Lake"])
        ws = openpyxl.load_workbook(out)[RG.YOY_TAB_NAME]
        check("YoY 2026 MTD populated", ws.cell(row=2, column=2).value == 33191.0)
        check("YoY 2025 column BLANK without prior-year (graceful)",
              ws.cell(row=2, column=3).value is None)

        # Case B — prior-year DATA_MONTHLY row + goals present.
        monthly = [{"year_month": "2025-05", "loc_name": "Forest Lake",
                    "total_sales": 32281.75, "guests": 659, "ppg": 3.58, "avg_ticket": 48.99}]
        goals = {"Forest Lake": {"monthly_goal": 38521.78, "daily_goal_divisor": 26,
                                 "day_goal_divisor": 26}}
        out2 = RG.build_workbook_from_data("2026-05", Path(d) / "b.xlsx",
                                           snapshots=snaps, monthly_rows=monthly,
                                           goals=goals, roster=["Forest Lake"])
        ws2 = openpyxl.load_workbook(out2)[RG.YOY_TAB_NAME]
        check("YoY 2025 column populated from prior-year row",
              ws2.cell(row=2, column=3).value == 32281.75)
        check("YoY Goal column populated from MONTHLY_GOALS",
              ws2.cell(row=2, column=5).value == 38521.78)
        check("YoY Daily Goal = goal/26",
              abs(ws2.cell(row=2, column=6).value - 38521.78 / 26) < 0.01)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# Read-as-stored, penny-exact placement.
# ─────────────────────────────────────────────────────────────────────────────
def t_penny_exact_read_as_stored():
    d = _fresh_dir()
    try:
        # A stored ratio that is NOT what naive recompute would give — proves we read
        # it verbatim. (avg_ticket here is intentionally not total/guests.)
        s = snap("Forest Lake", "2026-05-24", total=22720.0, service=20508.35,
                 guests=415, avg_ticket=999.99, ppg=1.23)
        out = RG.build_workbook_from_data("2026-05", Path(d) / "w.xlsx",
                                          snapshots=[s], roster=["Forest Lake"])
        ws = openpyxl.load_workbook(out)[RG.WEEK_TAB_NAMES[0]]
        fc = _field_cols(ws)
        check("service_net penny-exact (20508.35)",
              ws.cell(row=3, column=fc["service"]).value == 20508.35)
        check("avg_ticket read AS STORED (999.99, not recomputed)",
              ws.cell(row=3, column=fc["avg_ticket"]).value == 999.99)
        check("ppg read AS STORED (1.23)",
              ws.cell(row=3, column=fc["ppg"]).value == 1.23)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# Dry-run placeholder (the sandbox / main.py path).
# ─────────────────────────────────────────────────────────────────────────────
def t_dry_run_placeholder():
    d = _fresh_dir()
    try:
        out = RG.build_monthly_workbook("2026-05", Path(d) / "dry.xlsx",
                                        service=None, config={"locations": []}, dry_run=True)
        wb = openpyxl.load_workbook(out)
        check("dry-run writes a DRY_RUN placeholder tab", "DRY_RUN" in wb.sheetnames)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# INTEGRATION — real Forest Lake parse -> snapshot -> report (gated; Track D skip).
# ─────────────────────────────────────────────────────────────────────────────
_FL_PDF = str(_REPO_ROOT / "data" / "inbox" / "Forest Lake.pdf")


def integration_forest_lake():
    if not os.path.exists(_FL_PDF):
        print("SKIP - integration: data/inbox/Forest Lake.pdf not found")
        return
    try:
        from parsers.locations_grouper import build_location_row
        r = build_location_row(_FL_PDF, location_id="888-11812",
                               period_start="2026-05-01", period_end="2026-05-24",
                               period_type="mtd", period_label="MTD")
    except Exception as e:
        print(f"SKIP - integration: parse unavailable ({type(e).__name__}: {e})")
        return

    # Shape the penny-verified parser row as a CUMULATIVE_MTD snapshot.
    s = {
        "loc_name": "Forest Lake", "year_month": "2026-05", "week_ending": "2026-05-24",
        "platform": "zenoti", "guests": r["guest_count"], "total_sales": r["total_sales_net"],
        "service": r["service_net"], "product": r["product_net"],
        "product_pct": r["product_pct"], "ppg": r["ppg_net"], "pph": r["pph_net"],
        "avg_ticket": r["avg_ticket"], "prod_hours": r["productive_hours"],
        "wax_count": r["wax_count"], "wax": r["wax_net"], "wax_pct": r["wax_pct"],
        "color": r["color_net"], "color_pct": r["color_pct"],
        "treat_count": r["treatment_count"], "treat": r["treatment_net"],
        "treat_pct": r["treatment_pct"], "source": "integration",
    }
    d = _fresh_dir()
    try:
        out = RG.build_workbook_from_data("2026-05", Path(d) / "fl.xlsx",
                                          snapshots=[s], roster=["Forest Lake"])
        ws = openpyxl.load_workbook(out)[RG.WEEK_TAB_NAMES[0]]
        fc = _field_cols(ws)
        check("[integration] FL service_net penny-exact in 'Service Net' cell",
              ws.cell(row=3, column=fc["service"]).value == r["service_net"])
        check("[integration] FL total penny-exact in 'Total Sales Net' cell",
              ws.cell(row=3, column=fc["total_sales"]).value == r["total_sales_net"])
        print(f"-- integration: Forest Lake service_net={r['service_net']} placed penny-exact")
    finally:
        shutil.rmtree(d, ignore_errors=True)


if __name__ == "__main__":
    print("=== UNIT (synthetic snapshots; tmp workbook; no network) ===")
    t_single_week_layout()
    t_header_drift_mapping()
    t_cumulative_not_summed()
    t_closed_week_zero_row()
    t_provisional_marker()
    t_yoy_blank_vs_populated()
    t_penny_exact_read_as_stored()
    t_dry_run_placeholder()
    print("\n=== INTEGRATION (real Forest Lake -> report; gated) ===")
    integration_forest_lake()
    print("\n" + ("ALL PASS" if not FAILS else f"{len(FAILS)} FAILED: {FAILS}"))
    sys.exit(1 if FAILS else 0)
