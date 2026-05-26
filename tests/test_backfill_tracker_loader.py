"""
tests/test_backfill_tracker_loader.py
─────────────────────────────────────
Tests for scripts/backfill/tracker_loader.py.

Two layers:
  1. Unit tests against a synthetic in-memory workbook that mirrors the
     real tracker shape (Week 5 cumulative MTD layout). Always runs.
  2. Optional integration test against the real ``March 2026.xlsx``
     downloaded by Tony — skipped on machines that don't have the file.
"""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts.backfill.tracker_loader import (  # noqa: E402
    BOTTOM_BLOCK,
    TARGET_TAB_DEFAULT,
    TOP_BLOCK,
    _normalize,
    _to_number,
    get_tracker_week_ending,
    load_march_monthly_from_tracker,
)


# Stripped-down config mirroring the real config/customers/karissa_001.json
SAMPLE_CONFIG = {
    "customer_id": "test_001",
    "sheet_id": "test_sheet",
    "locations": [
        {"id": "z001", "name": "Andover FS",   "platform": "zenoti"},
        {"id": "z002", "name": "Blaine",       "platform": "zenoti"},
        {"id": "z003", "name": "Crystal FS",   "platform": "zenoti"},
        {"id": "z010", "name": "Apple Valley", "platform": "salon_ultimate"},
    ],
}


def _build_fixture_workbook(
    path: Path,
    *,
    locations_with_data: list[tuple[str, dict]] | None = None,
    include_bottom: bool = True,
) -> None:
    """Write a minimal tracker workbook with a Week 5 tab in the target layout.

    Each location dict supplies all 19 KPI values (top + bottom). Locations
    not in the list are simply omitted (simulating a missing row).
    """
    from datetime import datetime as _dt

    if locations_with_data is None:
        locations_with_data = [
            ("Andover FS", _full_kpi_dict(guests=578, total_sales=33291.80)),
            ("Blaine",     _full_kpi_dict(guests=1050, total_sales=63049.40)),
            ("Crystal FS", _full_kpi_dict(guests=1165, total_sales=65447.83)),
            # Trailing-whitespace bug from the real tracker — verify we strip it
            ("Apple Valley ", _full_kpi_dict(guests=1796, total_sales=105477.12)),
        ]

    wb = Workbook()
    # Remove default sheet so we control the layout
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet(title=TARGET_TAB_DEFAULT)

    # A1: date marker (Week 5 = 2026-03-31)
    ws["A1"] = _dt(2026, 3, 31)

    # Header rows
    ws.cell(row=2, column=1, value="Name")
    ws.cell(row=2, column=2, value="Guest Count")
    ws.cell(row=2, column=3, value="Total Sales Net")

    # Top block data starting row 3
    for i, (loc_name, kpis) in enumerate(locations_with_data):
        r = TOP_BLOCK["first_row"] + i
        ws.cell(row=r, column=TOP_BLOCK["cols"]["loc_name"], value=loc_name)
        ws.cell(row=r, column=TOP_BLOCK["cols"]["guests"], value=kpis["guests"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["total_sales"], value=kpis["total_sales"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["service"], value=kpis["service"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["product"], value=kpis["product"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["product_pct"], value=kpis["product_pct"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["ppg"], value=kpis["ppg"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["pph"], value=kpis["pph"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["avg_ticket"], value=kpis["avg_ticket"])
        ws.cell(row=r, column=TOP_BLOCK["cols"]["prod_hours"], value=kpis["prod_hours"])

    # Bottom block header
    ws.cell(row=18, column=2, value="Wax Count")

    if include_bottom:
        for i, (loc_name, kpis) in enumerate(locations_with_data):
            r = BOTTOM_BLOCK["first_row"] + i
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["loc_name"], value=loc_name)
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["wax_count"], value=kpis["wax_count"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["wax"], value=kpis["wax"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["wax_pct"], value=kpis["wax_pct"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["color"], value=kpis["color"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["color_pct"], value=kpis["color_pct"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["treat_count"], value=kpis["treat_count"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["treat"], value=kpis["treat"])
            ws.cell(row=r, column=BOTTOM_BLOCK["cols"]["treat_pct"], value=kpis["treat_pct"])

    wb.save(path)
    wb.close()


def _full_kpi_dict(*, guests, total_sales) -> dict:
    """Build a complete 19-field KPI dict with deterministic test values."""
    # Use guests as a seed so each location's numbers are distinct but predictable
    return {
        "guests": guests,
        "total_sales": total_sales,
        "service": total_sales * 0.95,
        "product": total_sales * 0.05,
        "product_pct": 0.05,
        "ppg": 3.5,
        "pph": 40.0,
        "avg_ticket": total_sales / guests if guests else 0,
        "prod_hours": 800.0,
        "wax_count": guests // 7,
        "wax": (guests // 7) * 18.0,
        "wax_pct": (guests // 7) / guests if guests else 0,
        "color": total_sales * 0.35,
        "color_pct": 0.35,
        "treat_count": guests // 10,
        "treat": (guests // 10) * 20.0,
        "treat_pct": (guests // 10) / guests if guests else 0,
    }


class TestNormalize(unittest.TestCase):
    def test_strips_trailing_space(self):
        self.assertEqual(_normalize("Apple Valley "), "Apple Valley")

    def test_collapses_internal_whitespace(self):
        self.assertEqual(_normalize("Apple  Valley"), "Apple Valley")

    def test_none_returns_empty(self):
        self.assertEqual(_normalize(None), "")

    def test_empty_returns_empty(self):
        self.assertEqual(_normalize(""), "")


class TestToNumber(unittest.TestCase):
    def test_none_returns_zero(self):
        self.assertEqual(_to_number(None), 0)

    def test_int_passes_through(self):
        self.assertEqual(_to_number(42), 42)

    def test_float_passes_through(self):
        self.assertAlmostEqual(_to_number(3.14), 3.14)

    def test_strips_currency_and_commas(self):
        self.assertEqual(_to_number("$1,092.45"), 1092.45)

    def test_strips_percent(self):
        self.assertEqual(_to_number("19.4%"), 19.4)

    def test_unparseable_string_returns_zero(self):
        self.assertEqual(_to_number("abc"), 0)


class TestLoaderSynthetic(unittest.TestCase):
    def test_happy_path_all_locations_present(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tracker.xlsx"
            _build_fixture_workbook(path)

            rows = load_march_monthly_from_tracker(path, SAMPLE_CONFIG)

            self.assertEqual(len(rows), 4)
            # Order matches config locations
            self.assertEqual([r["loc_name"] for r in rows],
                             ["Andover FS", "Blaine", "Crystal FS", "Apple Valley"])

            andover = rows[0]
            self.assertEqual(andover["year_month"], "2026-03")
            self.assertEqual(andover["platform"], "zenoti")
            self.assertEqual(andover["guests"], 578)
            self.assertAlmostEqual(andover["total_sales"], 33291.80, places=2)
            self.assertEqual(andover["source"], "tracker")
            self.assertEqual(andover["period_start"], "2026-03-01")
            self.assertEqual(andover["period_end"], "2026-03-31")

    def test_trailing_whitespace_in_tracker_name_is_matched(self):
        """Tracker stores 'Apple Valley ' but config has 'Apple Valley'."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tracker.xlsx"
            _build_fixture_workbook(path)

            rows = load_march_monthly_from_tracker(path, SAMPLE_CONFIG)
            apple = [r for r in rows if r["loc_name"] == "Apple Valley"]
            self.assertEqual(len(apple), 1)
            self.assertEqual(apple[0]["platform"], "salon_ultimate")
            # Use config's canonical name (no trailing space) in output
            self.assertEqual(apple[0]["loc_name"], "Apple Valley")

    def test_missing_location_raises_with_actionable_message(self):
        """If config has a location the tracker doesn't, we fail loudly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tracker.xlsx"
            # Tracker missing Crystal FS
            _build_fixture_workbook(
                path,
                locations_with_data=[
                    ("Andover FS", _full_kpi_dict(guests=578, total_sales=33291.80)),
                    ("Blaine",     _full_kpi_dict(guests=1050, total_sales=63049.40)),
                    ("Apple Valley", _full_kpi_dict(guests=1796, total_sales=105477.12)),
                ],
            )
            with self.assertRaises(ValueError) as cm:
                load_march_monthly_from_tracker(path, SAMPLE_CONFIG)
            self.assertIn("Crystal FS", str(cm.exception))

    def test_bottom_block_missing_does_not_crash(self):
        """If service-mix block absent, wax/color/treat fields default to 0."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tracker.xlsx"
            _build_fixture_workbook(path, include_bottom=False)

            rows = load_march_monthly_from_tracker(path, SAMPLE_CONFIG)
            andover = rows[0]
            self.assertEqual(andover["wax_count"], 0)
            self.assertEqual(andover["wax"], 0)
            self.assertEqual(andover["color"], 0)
            self.assertEqual(andover["treat_count"], 0)

    def test_get_tracker_week_ending(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tracker.xlsx"
            _build_fixture_workbook(path)
            self.assertEqual(get_tracker_week_ending(path), "2026-03-31")

    def test_file_not_found_raises(self):
        with self.assertRaises(FileNotFoundError):
            load_march_monthly_from_tracker("/no/such/file.xlsx", SAMPLE_CONFIG)

    def test_wrong_tab_name_raises_with_actionable_message(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tracker.xlsx"
            _build_fixture_workbook(path)
            with self.assertRaises(ValueError) as cm:
                load_march_monthly_from_tracker(path, SAMPLE_CONFIG, target_tab="No Such Tab")
            self.assertIn("Available tabs", str(cm.exception))


# ────────────────────────────────────────────────────────────────────────
# Optional integration test against the real March 2026.xlsx
# ────────────────────────────────────────────────────────────────────────

REAL_TRACKER_PATH = Path(r"C:\Users\TonyGrant\Downloads\KPI Historical Data\March 2026.xlsx")
REAL_CONFIG_PATH = _REPO_ROOT / "config" / "customers" / "karissa_001.json"


@unittest.skipUnless(
    REAL_TRACKER_PATH.exists() and REAL_CONFIG_PATH.exists(),
    f"Real tracker not present at {REAL_TRACKER_PATH} — skipping integration test",
)
class TestLoaderIntegration(unittest.TestCase):
    def test_real_tracker_loads_all_12_locations(self):
        import json
        config = json.loads(REAL_CONFIG_PATH.read_text())
        rows = load_march_monthly_from_tracker(REAL_TRACKER_PATH, config)
        self.assertEqual(len(rows), 12)

        # Spot-check Blaine — values come from the validated 2026-05-26 inspection
        blaine = [r for r in rows if r["loc_name"] == "Blaine"]
        self.assertEqual(len(blaine), 1)
        b = blaine[0]
        self.assertEqual(b["guests"], 1050)
        self.assertAlmostEqual(b["total_sales"], 63049.40, places=2)
        self.assertAlmostEqual(b["service"], 57848.31, places=2)
        self.assertEqual(b["platform"], "zenoti")
        self.assertEqual(b["source"], "tracker")

    def test_real_tracker_week_ending_is_2026_03_31(self):
        self.assertEqual(get_tracker_week_ending(REAL_TRACKER_PATH), "2026-03-31")


if __name__ == "__main__":
    unittest.main()
