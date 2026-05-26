"""
tests/test_backfill_zenoti_xlsx_loader.py
─────────────────────────────────────────
Tests for scripts/backfill/zenoti_xlsx_loader.py.

Synthetic-fixture coverage of _detect_location, _read_total_row,
load_zenoti_march_xlsxes, diff_against_tracker; plus an optional
integration test against the real March folder.
"""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts.backfill.zenoti_xlsx_loader import (  # noqa: E402
    _detect_location,
    _normalize,
    _read_stylist_rows,
    _read_total_row,
    diff_against_tracker,
    load_zenoti_march_stylists,
    load_zenoti_march_xlsxes,
)


SAMPLE_CONFIG = {
    "locations": [
        {"id": "z001", "name": "Andover FS",   "platform": "zenoti"},
        {"id": "z002", "name": "Blaine",       "platform": "zenoti"},
        {"id": "z003", "name": "Crystal FS",   "platform": "zenoti"},
        {"id": "z009", "name": "Roseville",    "platform": "zenoti"},
        {"id": "z010", "name": "Apple Valley", "platform": "salon_ultimate"},
    ],
}


def _build_xlsx(
    path: Path,
    mgr_name: str,
    *,
    invoices: int = 1000,
    guests: int = 950,
    service: float = 50000.0,
    product: float = 5000.0,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee KPI"
    ws.cell(row=1, column=1, value="Fantastic Sams")
    ws.cell(row=2, column=1, value="Employee KPI From: 01 Mar 2026 To: 31 Mar 2026")
    ws.cell(row=4, column=1, value="Employee Name")
    ws.cell(row=4, column=2, value="Invoice Count")
    ws.cell(row=5, column=1, value="Stylist A")
    ws.cell(row=6, column=1, value=mgr_name)
    ws.cell(row=7, column=1, value="Stylist B")
    # Total row
    ws.cell(row=8, column=1, value="Total")
    ws.cell(row=8, column=2, value=invoices)
    ws.cell(row=8, column=3, value=service / invoices if invoices else 0)
    ws.cell(row=8, column=4, value=guests)
    ws.cell(row=8, column=5, value=service)
    ws.cell(row=8, column=6, value=service / invoices if invoices else 0)
    ws.cell(row=8, column=7, value=product)
    ws.cell(row=8, column=8, value=product / invoices if invoices else 0)
    wb.save(path)
    wb.close()


class TestDetectLocation(unittest.TestCase):
    def test_exact_mgr_pattern(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            _build_xlsx(path, mgr_name="Blaine mgr")
            wb = __import__("openpyxl").load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            try:
                detected = _detect_location(ws, ["Andover FS", "Blaine", "Crystal FS"])
                self.assertEqual(detected, "Blaine")
            finally:
                wb.close()

    def test_andover_fs_matched_via_substring(self):
        """File has 'Andover mgr' but config name is 'Andover FS' — should still match."""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            _build_xlsx(path, mgr_name="Andover mgr")
            wb = __import__("openpyxl").load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            try:
                detected = _detect_location(ws, ["Andover FS", "Blaine"])
                # "Andover" appears as substring of "Andover fs" - substring match
                # (rank-1) returns Andover FS
                self.assertEqual(detected, "Andover FS")
            finally:
                wb.close()

    def test_roseville_weird_format(self):
        """Real Roseville mgr row: '888-40098-F Sams Roseville, MN_mgr mgr'."""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            _build_xlsx(path, mgr_name="888-40098-F Sams Roseville, MN_mgr mgr")
            wb = __import__("openpyxl").load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            try:
                detected = _detect_location(ws, ["Roseville", "Blaine"])
                self.assertEqual(detected, "Roseville")
            finally:
                wb.close()

    def test_no_match_returns_none(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            _build_xlsx(path, mgr_name="UnknownSalon mgr")
            wb = __import__("openpyxl").load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            try:
                detected = _detect_location(ws, ["Andover FS", "Blaine"])
                self.assertIsNone(detected)
            finally:
                wb.close()


class TestReadTotalRow(unittest.TestCase):
    def test_extracts_total_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            _build_xlsx(path, mgr_name="Blaine mgr",
                        invoices=1237, guests=1199, service=57848.31, product=5201.09)
            wb = __import__("openpyxl").load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            try:
                total = _read_total_row(ws)
                self.assertEqual(total["invoices"], 1237)
                self.assertEqual(total["guests"], 1199)
                self.assertAlmostEqual(total["service"], 57848.31, places=2)
                self.assertAlmostEqual(total["product"], 5201.09, places=2)
            finally:
                wb.close()


class TestLoadZenotiMarchXlsxes(unittest.TestCase):
    def test_loads_all_zenoti_locations(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            _build_xlsx(tmp_path / "a.xlsx", mgr_name="Andover mgr",
                        invoices=583, guests=566, service=31534.55, product=1757.25)
            _build_xlsx(tmp_path / "b.xlsx", mgr_name="Blaine mgr",
                        invoices=1237, guests=1199, service=57848.31, product=5201.09)
            _build_xlsx(tmp_path / "c.xlsx", mgr_name="Crystal mgr",
                        invoices=1177, guests=1136, service=61217.72, product=4230.10)
            _build_xlsx(tmp_path / "r.xlsx", mgr_name="888-40098-F Sams Roseville, MN_mgr mgr",
                        invoices=1077, guests=1031, service=39933.61, product=1028.18)
            # SU location should not be expected in xlsxes (we filter to zenoti)
            cfg = {
                "locations": [
                    {"id": "z001", "name": "Andover FS",   "platform": "zenoti"},
                    {"id": "z002", "name": "Blaine",       "platform": "zenoti"},
                    {"id": "z003", "name": "Crystal FS",   "platform": "zenoti"},
                    {"id": "z009", "name": "Roseville",    "platform": "zenoti"},
                    {"id": "z010", "name": "Apple Valley", "platform": "salon_ultimate"},
                ]
            }
            out = load_zenoti_march_xlsxes(tmp_path, cfg)
            self.assertEqual(set(out.keys()), {"Andover FS", "Blaine", "Crystal FS", "Roseville"})
            self.assertAlmostEqual(out["Blaine"]["service"], 57848.31, places=2)

    def test_missing_zenoti_location_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            _build_xlsx(tmp_path / "a.xlsx", mgr_name="Andover mgr")
            cfg = {
                "locations": [
                    {"id": "z001", "name": "Andover FS", "platform": "zenoti"},
                    {"id": "z002", "name": "Blaine",     "platform": "zenoti"},
                ]
            }
            with self.assertRaises(ValueError) as cm:
                load_zenoti_march_xlsxes(tmp_path, cfg)
            self.assertIn("Blaine", str(cm.exception))


class TestDiffAgainstTracker(unittest.TestCase):
    def test_matching_values_no_findings(self):
        tracker_rows = [
            {"loc_name": "Blaine", "service": 57848.31, "product": 5201.09, "guests": 1050},
        ]
        xlsx_totals = {
            "Blaine": {"service": 57848.31, "product": 5201.09, "guests": 1050},
        }
        findings = diff_against_tracker(tracker_rows, xlsx_totals)
        self.assertEqual(findings, [])

    def test_service_mismatch_above_tolerance_is_error(self):
        tracker_rows = [{"loc_name": "Blaine", "service": 57848.31, "product": 5201.09, "guests": 1050}]
        xlsx_totals = {"Blaine": {"service": 57850.00, "product": 5201.09, "guests": 1050}}
        findings = diff_against_tracker(tracker_rows, xlsx_totals)
        errors = [f for f in findings if f["severity"] == "error"]
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0]["code"], "SERVICE_MISMATCH")

    def test_product_mismatch_above_tolerance_is_error(self):
        tracker_rows = [{"loc_name": "Blaine", "service": 57848.31, "product": 5201.09, "guests": 1050}]
        xlsx_totals = {"Blaine": {"service": 57848.31, "product": 5210.00, "guests": 1050}}
        findings = diff_against_tracker(tracker_rows, xlsx_totals)
        errors = [f for f in findings if f["severity"] == "error"]
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0]["code"], "PRODUCT_MISMATCH")

    def test_guest_count_small_diff_is_info(self):
        tracker_rows = [{"loc_name": "Blaine", "service": 57848.31, "product": 5201.09, "guests": 1050}]
        xlsx_totals = {"Blaine": {"service": 57848.31, "product": 5201.09, "guests": 1100}}
        findings = diff_against_tracker(tracker_rows, xlsx_totals)
        # 50/1050 = 4.76% < 10% threshold → info
        infos = [f for f in findings if f["severity"] == "info"]
        warnings = [f for f in findings if f["severity"] == "warning"]
        errors = [f for f in findings if f["severity"] == "error"]
        self.assertEqual(len(infos), 1)
        self.assertEqual(infos[0]["code"], "GUEST_COUNT_DIFF")
        self.assertEqual(warnings, [])
        self.assertEqual(errors, [])

    def test_guest_count_large_diff_is_warning(self):
        tracker_rows = [{"loc_name": "Blaine", "service": 57848.31, "product": 5201.09, "guests": 1000}]
        xlsx_totals = {"Blaine": {"service": 57848.31, "product": 5201.09, "guests": 1500}}
        findings = diff_against_tracker(tracker_rows, xlsx_totals)
        # 500/1000 = 50% > 10% threshold → warning
        warnings = [f for f in findings if f["severity"] == "warning"]
        self.assertEqual(len(warnings), 1)
        self.assertEqual(warnings[0]["code"], "GUEST_COUNT_DIFF")


# ─── Optional integration test against real March folder ───────────────────────

REAL_MARCH_DIR = Path(r"C:\Users\TonyGrant\Downloads\KPI Historical Data\3-1-26 - 3-31-26")
REAL_CONFIG_PATH = _REPO_ROOT / "config" / "customers" / "karissa_001.json"


@unittest.skipUnless(
    REAL_MARCH_DIR.exists() and REAL_CONFIG_PATH.exists(),
    f"Real March folder not present at {REAL_MARCH_DIR} — skipping integration",
)
class TestRealMarchIntegration(unittest.TestCase):
    def test_all_9_zenoti_xlsxes_loaded(self):
        cfg = json.loads(REAL_CONFIG_PATH.read_text())
        out = load_zenoti_march_xlsxes(REAL_MARCH_DIR, cfg)
        expected = {loc["name"] for loc in cfg["locations"] if loc["platform"] == "zenoti"}
        self.assertEqual(set(out.keys()), expected)

        # Blaine spot-check (verified 2026-05-26)
        b = out["Blaine"]
        self.assertEqual(b["invoices"], 1237)
        self.assertEqual(b["guests"], 1199)
        self.assertAlmostEqual(b["service"], 57848.31, places=2)
        self.assertAlmostEqual(b["product"], 5201.09, places=2)

    def test_real_march_stylist_load(self):
        """Real March Zenoti xlsxes should produce stylist rows for the active
        roster across all 9 Zenoti locations."""
        cfg = json.loads(REAL_CONFIG_PATH.read_text())
        rows = load_zenoti_march_stylists(REAL_MARCH_DIR, cfg)
        # Sanity: at least 30 active stylists across the 9 Zenoti locations
        self.assertGreater(len(rows), 30, "expected >30 active March stylists")

        # All 9 Zenoti locations should be represented at least once
        zenoti_names = {loc["name"] for loc in cfg["locations"] if loc["platform"] == "zenoti"}
        seen_locs = {r["loc_name"] for r in rows}
        self.assertEqual(seen_locs, zenoti_names)

        # No inactive stylists (zero service AND zero product) made it through
        for r in rows:
            self.assertTrue(
                r["net_service"] > 0 or r["net_product"] > 0,
                f"inactive stylist leaked through: {r}",
            )

        # Provenance set
        self.assertTrue(all(r["source"] == "zenoti_xlsx" for r in rows))
        self.assertTrue(all(r["year_month"] == "2026-03" for r in rows))

        # Spot-check known Blaine stylist from validated 2026-05-26 inspection
        # Blaine xlsx Total: 1237 invoices, $57,848.31 service
        blaine_rows = [r for r in rows if r["loc_name"] == "Blaine"]
        blaine_service_sum = sum(r["net_service"] for r in blaine_rows)
        # Active stylist sum should be close to Total (within reasonable rounding,
        # since some inactive zero-rows are dropped from the active set)
        self.assertAlmostEqual(blaine_service_sum, 57848.31, places=2)

    def test_real_diff_against_tracker_only_guest_diffs(self):
        """Dollar amounts in tracker should match xlsx to within $0.01.
        Guest counts will differ (methodology) but only as info/warning."""
        from scripts.backfill.tracker_loader import load_march_monthly_from_tracker

        cfg = json.loads(REAL_CONFIG_PATH.read_text())
        tracker_rows = load_march_monthly_from_tracker(
            r"C:\Users\TonyGrant\Downloads\KPI Historical Data\March 2026.xlsx",
            cfg,
        )
        xlsx_totals = load_zenoti_march_xlsxes(REAL_MARCH_DIR, cfg)
        findings = diff_against_tracker(tracker_rows, xlsx_totals)

        errors = [f for f in findings if f["severity"] == "error"]
        self.assertEqual(errors, [], f"Unexpected $$ mismatches: {errors}")


if __name__ == "__main__":
    unittest.main()
