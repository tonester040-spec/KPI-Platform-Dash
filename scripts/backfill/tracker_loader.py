"""
scripts/backfill/tracker_loader.py
──────────────────────────────────
Loads Karissa's master tracker spreadsheet (``March 2026.xlsx``) into a
list of DATA_MONTHLY-ready row dicts.

Critical context: the tracker's weekly tabs are *cumulative month-to-date*,
not independent weekly snapshots. Week 5 == cumulative through 3/31 == the
full March monthly aggregate. This loader reads only Week 5 (or any tab
named in TARGET_TAB) and emits one row per location.

Tracker shape (Week 5 tab, validated 2026-05-26):
  Top KPI block — rows 3-14, 12 locations
    A loc_name  B guests  C total_sales  D service  E product  F product_pct
    G ppg       H pph     I avg_ticket   J prod_hours
  Bottom service-mix block — rows 19-30, 12 locations
    A loc_name  B wax_count  C wax  D wax_pct
    E color     F color_pct  G treat_count  H treat  I treat_pct

Tracker location names have trailing whitespace in some cells
(e.g. ``"Apple Valley "``) — normalize with .strip() before matching.

Karissa's tracker is the gold source — we read values verbatim and do NOT
recompute. Some ratios in the tracker (e.g. color_pct) use denominators
that don't exactly match the canonical formula in CLAUDE.md, but Karissa's
numbers are what she sees in her own reports, so we honor them.
"""

from __future__ import annotations

import datetime
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)


# The week-5 tab name as stored in March 2026.xlsx (note the leading space)
TARGET_TAB_DEFAULT = " week 5"

# Top block layout — col indexes are 1-based for openpyxl, value = column number
TOP_BLOCK = {
    "first_row": 3,
    "last_row": 14,
    "cols": {
        "loc_name": 1,
        "guests": 2,
        "total_sales": 3,
        "service": 4,
        "product": 5,
        "product_pct": 6,
        "ppg": 7,
        "pph": 8,
        "avg_ticket": 9,
        "prod_hours": 10,
    },
}

# Bottom block layout
BOTTOM_BLOCK = {
    "first_row": 19,
    "last_row": 30,
    "cols": {
        "loc_name": 1,
        "wax_count": 2,
        "wax": 3,
        "wax_pct": 4,
        "color": 5,
        "color_pct": 6,
        "treat_count": 7,
        "treat": 8,
        "treat_pct": 9,
    },
}


def _normalize(name: str | None) -> str:
    """Strip whitespace and collapse internal multi-spaces — tracker cells are messy."""
    if not name:
        return ""
    return " ".join(str(name).split()).strip()


def _to_number(v: Any) -> float | int:
    """Convert cell value to a number; None or non-numeric → 0."""
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        cleaned = v.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0
    return 0


def _read_block(ws, block: dict) -> dict[str, dict[str, Any]]:
    """Read a block (top or bottom) and return {normalized_name: {col_key: value}}."""
    out: dict[str, dict[str, Any]] = {}
    for r in range(block["first_row"], block["last_row"] + 1):
        loc_raw = ws.cell(row=r, column=block["cols"]["loc_name"]).value
        loc_norm = _normalize(loc_raw)
        if not loc_norm or loc_norm.lower() == "totals":
            continue
        row_data: dict[str, Any] = {}
        for key, col in block["cols"].items():
            if key == "loc_name":
                row_data[key] = loc_norm
            else:
                row_data[key] = _to_number(ws.cell(row=r, column=col).value)
        out[loc_norm] = row_data
    return out


def load_march_monthly_from_tracker(
    tracker_path: str | Path,
    customer_config: dict,
    *,
    target_tab: str = TARGET_TAB_DEFAULT,
    year_month: str = "2026-03",
    period_start: str = "2026-03-01",
    period_end: str = "2026-03-31",
) -> list[dict]:
    """Parse Karissa's tracker into DATA_MONTHLY-ready row dicts.

    Returns one dict per location in ``customer_config["locations"]`` whose
    name appears in the tracker. Raises ValueError if any config location
    is missing from the tracker — the backfill is gated on having all 12
    locations present.
    """
    tracker_path = Path(tracker_path)
    if not tracker_path.exists():
        raise FileNotFoundError(f"Tracker file not found: {tracker_path}")

    wb = load_workbook(tracker_path, data_only=True, read_only=True)
    try:
        if target_tab not in wb.sheetnames:
            raise ValueError(
                f"Tab {target_tab!r} not found in {tracker_path.name}. "
                f"Available tabs: {wb.sheetnames}"
            )
        ws = wb[target_tab]

        top = _read_block(ws, TOP_BLOCK)
        bottom = _read_block(ws, BOTTOM_BLOCK)

        rows: list[dict] = []
        missing: list[str] = []

        for loc in customer_config["locations"]:
            cfg_name = _normalize(loc["name"])
            if cfg_name not in top:
                missing.append(loc["name"])
                continue

            t = top[cfg_name]
            b = bottom.get(cfg_name, {})

            row = {
                "loc_name": loc["name"],  # use canonical name from config, not tracker
                "year_month": year_month,
                "platform": loc["platform"],
                "guests": _to_number(t.get("guests")),
                "total_sales": _to_number(t.get("total_sales")),
                "service": _to_number(t.get("service")),
                "product": _to_number(t.get("product")),
                "product_pct": _to_number(t.get("product_pct")),
                "ppg": _to_number(t.get("ppg")),
                "pph": _to_number(t.get("pph")),
                "avg_ticket": _to_number(t.get("avg_ticket")),
                "prod_hours": _to_number(t.get("prod_hours")),
                "wax_count": _to_number(b.get("wax_count")),
                "wax": _to_number(b.get("wax")),
                "wax_pct": _to_number(b.get("wax_pct")),
                "color": _to_number(b.get("color")),
                "color_pct": _to_number(b.get("color_pct")),
                "treat_count": _to_number(b.get("treat_count")),
                "treat": _to_number(b.get("treat")),
                "treat_pct": _to_number(b.get("treat_pct")),
                "source": "tracker",
                "period_start": period_start,
                "period_end": period_end,
            }
            rows.append(row)

        if missing:
            raise ValueError(
                f"Tracker tab {target_tab!r} is missing rows for: {missing}. "
                f"Found in tracker: {sorted(top.keys())}"
            )

        log.info(
            "Loaded %d locations from tracker %s (tab=%r, year_month=%s)",
            len(rows), tracker_path.name, target_tab, year_month,
        )
        return rows
    finally:
        wb.close()


def get_tracker_week_ending(tracker_path: str | Path, tab: str = TARGET_TAB_DEFAULT) -> str | None:
    """Return the date stored in A1 of the given tab as YYYY-MM-DD, or None.

    Used to confirm Week 5 == 2026-03-31 before we trust its values
    as the March monthly aggregate.
    """
    tracker_path = Path(tracker_path)
    wb = load_workbook(tracker_path, data_only=True, read_only=True)
    try:
        if tab not in wb.sheetnames:
            return None
        a1 = wb[tab]["A1"].value
        if isinstance(a1, datetime.datetime):
            return a1.date().isoformat()
        if isinstance(a1, datetime.date):
            return a1.isoformat()
        return None
    finally:
        wb.close()
