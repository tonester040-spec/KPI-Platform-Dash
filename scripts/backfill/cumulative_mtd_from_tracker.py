"""
scripts/backfill/cumulative_mtd_from_tracker.py
───────────────────────────────────────────────
One-off loader that seeds two tabs from Karissa's master May 2026 tracker:

  CUMULATIVE_MTD     <- May 2026 Week 1/Week 2/Week 4 cumulative-MTD snapshots
                        (Week 3 left blank in her tracker — skipped;
                         Week 5 will be written by the live pipeline on 6/1)

  DATA_MONTHLY       <- 2025-05, 2019-05, 2025-Q2, 2019-Q2 totals per location
                        (sparse rows — only ``total_sales`` is populated;
                         year_month "2025-Q2" / "2019-Q2" are non-standard but
                         documented here; the YoY report builder filters by
                         exact year_month value, so the schema bend is contained.)

Z-mode contract (per the 2026-05-27 architecture discussion):

  This loader extracts RAW values verbatim from her tracker. It does NOT load
  her tracker's pre-computed derived ratios (PPG, PPH, etc.) because those use
  POS-reported denominators that conflict with Karissa's canonical formula
  contract in CLAUDE.md (Zenoti PPG uses unique-guest count in the POS report;
  canonical PPG uses invoice count).

  The downstream report builder recomputes derived ratios canonically from
  these raw values. CUMULATIVE_MTD rows store only raw fields populated;
  percentage / ratio columns are written as 0 and will be recomputed at read
  time by the report builder.

Usage:
  # Dry-run (default tracker path on Tony's Desktop)
  python -m scripts.backfill.cumulative_mtd_from_tracker --dry-run

  # Write to Sheets after dry-run review
  python -m scripts.backfill.cumulative_mtd_from_tracker --write

  # Override the tracker location
  python -m scripts.backfill.cumulative_mtd_from_tracker --dry-run \\
      --tracker "C:/path/to/May 2026.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(_REPO_ROOT / ".env")
except ImportError:
    pass

log = logging.getLogger(__name__)


DEFAULT_TRACKER = Path(r"C:\Users\TonyGrant\Desktop\May 2026.xlsx")
DEFAULT_CONFIG = _REPO_ROOT / "config" / "customers" / "karissa_001.json"


# ─── Per-week tab layouts ─────────────────────────────────────────────────────
#
# Karissa's May 2026.xlsx has 5 weekly tabs. Week 1 has an extra "tickets"
# column in the top KPI block AND a blank column in the service-mix block,
# shifting everything one column to the right. Weeks 2/3/4/5 share a layout.
# Week 4 adds a "Goals" column we don't read (live pipeline already has goals
# in the GOALS tab).

@dataclass(frozen=True)
class WeekLayout:
    tab_name: str
    week_ending: str          # YYYY-MM-DD (Sunday of that week)
    top_first_row: int
    top_last_row: int
    bottom_first_row: int
    bottom_last_row: int
    # Top block columns (1-based for openpyxl)
    col_name: int
    col_guests: int
    col_total_sales: int
    col_service: int
    col_product: int
    col_prod_hours: int
    # Bottom block columns
    col_bot_name: int
    col_wax_count: int
    col_wax_net: int
    col_color_net: int
    col_treat_count: int
    col_treat_net: int


# Week 1: tickets col at B shifts top block; blank col at B shifts bottom block.
WEEK_1 = WeekLayout(
    tab_name=" Week 1 ",
    week_ending="2026-05-03",
    top_first_row=3, top_last_row=14,
    bottom_first_row=19, bottom_last_row=30,
    col_name=1, col_guests=3, col_total_sales=4, col_service=5,
    col_product=6, col_prod_hours=11,
    col_bot_name=1, col_wax_count=3, col_wax_net=4,
    col_color_net=6, col_treat_count=8, col_treat_net=9,
)

# Week 2/3/4/5 share the same raw-field column positions.
WEEK_2 = WeekLayout(
    tab_name=" week 2",
    week_ending="2026-05-10",
    top_first_row=3, top_last_row=14,
    bottom_first_row=19, bottom_last_row=30,
    col_name=1, col_guests=2, col_total_sales=3, col_service=4,
    col_product=5, col_prod_hours=10,
    col_bot_name=1, col_wax_count=2, col_wax_net=3,
    col_color_net=5, col_treat_count=7, col_treat_net=8,
)

WEEK_4 = WeekLayout(
    tab_name=" week 4",
    week_ending="2026-05-24",
    top_first_row=3, top_last_row=14,
    bottom_first_row=19, bottom_last_row=30,
    col_name=1, col_guests=2, col_total_sales=3, col_service=4,
    col_product=5, col_prod_hours=10,
    col_bot_name=1, col_wax_count=2, col_wax_net=3,
    col_color_net=5, col_treat_count=7, col_treat_net=8,
)

# Week 3 is left blank in her May tracker (she skipped it).
# Week 5 will be written by the live pipeline on Monday 6/1.
WEEKS_TO_LOAD = [WEEK_1, WEEK_2, WEEK_4]


# ─── Year Over Year tab layout ────────────────────────────────────────────────
#
# Three sub-tables:
#   Table 1 (rows 2-13): current-month YoY — col C = 2025 May, col J = 2019 May
#   Table 2 (rows 18-29): Q2 YoY — col C = Q2 2025, col E = Q2 2019
#   Table 3 (rows 19-30 right side): Q2 monthly breakdown — col K = April, col L = May, ...
#                                    (used here only for cross-checking the April 2026 row
#                                    that already exists in DATA_MONTHLY from the existing backfill)

YOY_TAB = "Year Over Year"
YOY_TABLE1_FIRST_ROW = 2
YOY_TABLE1_LAST_ROW = 13
YOY_TABLE1_NAME_COL = 1
YOY_TABLE1_2025_COL = 3
YOY_TABLE1_2019_COL = 10

YOY_TABLE2_FIRST_ROW = 18
YOY_TABLE2_LAST_ROW = 29
YOY_TABLE2_NAME_COL = 1
YOY_TABLE2_Q2_2025_COL = 3
YOY_TABLE2_Q2_2019_COL = 5

YOY_TABLE3_FIRST_ROW = 19
YOY_TABLE3_LAST_ROW = 30
YOY_TABLE3_NAME_COL = 11        # col K (was col J in display layout; openpyxl 1-based)
YOY_TABLE3_APRIL_COL = 12       # col L


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _normalize(name: str | None) -> str:
    if not name:
        return ""
    return " ".join(str(name).split()).strip()


def _to_number(v: Any) -> float | int:
    """Convert cell value to a number; None / non-numeric / DIV-by-zero → 0."""
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        # Tracker may have '#DIV/0!' from her formulas where she hasn't filled inputs
        cleaned = v.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0
    return 0


def _read_week_top(ws, layout: WeekLayout) -> dict[str, dict[str, Any]]:
    """Read the top KPI block (raw fields only). Returns {normalized_name: row_dict}."""
    out: dict[str, dict[str, Any]] = {}
    for r in range(layout.top_first_row, layout.top_last_row + 1):
        loc_norm = _normalize(ws.cell(row=r, column=layout.col_name).value)
        if not loc_norm or loc_norm.lower() == "totals":
            continue
        out[loc_norm] = {
            "guests": _to_number(ws.cell(row=r, column=layout.col_guests).value),
            "total_sales": _to_number(ws.cell(row=r, column=layout.col_total_sales).value),
            "service": _to_number(ws.cell(row=r, column=layout.col_service).value),
            "product": _to_number(ws.cell(row=r, column=layout.col_product).value),
            "prod_hours": _to_number(ws.cell(row=r, column=layout.col_prod_hours).value),
        }
    return out


def _read_week_bottom(ws, layout: WeekLayout) -> dict[str, dict[str, Any]]:
    """Read the service-mix block (raw counts + net dollars only). Returns {name: row_dict}."""
    out: dict[str, dict[str, Any]] = {}
    for r in range(layout.bottom_first_row, layout.bottom_last_row + 1):
        loc_norm = _normalize(ws.cell(row=r, column=layout.col_bot_name).value)
        if not loc_norm or loc_norm.lower() == "totals":
            continue
        out[loc_norm] = {
            "wax_count": _to_number(ws.cell(row=r, column=layout.col_wax_count).value),
            "wax_net": _to_number(ws.cell(row=r, column=layout.col_wax_net).value),
            "color_net": _to_number(ws.cell(row=r, column=layout.col_color_net).value),
            "treat_count": _to_number(ws.cell(row=r, column=layout.col_treat_count).value),
            "treat_net": _to_number(ws.cell(row=r, column=layout.col_treat_net).value),
        }
    return out


# ─── Public loader ────────────────────────────────────────────────────────────

def load_cumulative_mtd_rows(
    tracker_path: str | Path,
    customer_config: dict,
) -> list[dict]:
    """Build CUMULATIVE_MTD-ready rows for May Wk1/Wk2/Wk4 from her tracker.

    Returns one row per (location, week_ending). Raw fields are populated;
    derived ratios are written as 0 (the report builder recomputes canonically).
    """
    tracker_path = Path(tracker_path)
    if not tracker_path.exists():
        raise FileNotFoundError(f"Tracker file not found: {tracker_path}")

    wb = load_workbook(tracker_path, data_only=True, read_only=True)
    try:
        rows: list[dict] = []
        missing_global: list[tuple[str, str]] = []  # (tab, loc_name)

        for layout in WEEKS_TO_LOAD:
            if layout.tab_name not in wb.sheetnames:
                raise ValueError(
                    f"Tab {layout.tab_name!r} not found in {tracker_path.name}. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[layout.tab_name]
            top = _read_week_top(ws, layout)
            bottom = _read_week_bottom(ws, layout)

            for loc in customer_config["locations"]:
                cfg_name = _normalize(loc["name"])
                if cfg_name not in top:
                    missing_global.append((layout.tab_name, loc["name"]))
                    continue

                t = top[cfg_name]
                b = bottom.get(cfg_name, {})

                rows.append({
                    "loc_name": loc["name"],
                    "year_month": "2026-05",
                    "week_ending": layout.week_ending,
                    "platform": loc["platform"],
                    "guests": t["guests"],
                    "total_sales": t["total_sales"],
                    "service": t["service"],
                    "product": t["product"],
                    # Derived ratios — written as 0; report builder recomputes from raw.
                    "product_pct": 0,
                    "ppg": 0,
                    "pph": 0,
                    "avg_ticket": 0,
                    "prod_hours": t["prod_hours"],
                    "wax_count": b.get("wax_count", 0),
                    "wax": b.get("wax_net", 0),
                    "wax_pct": 0,
                    "color": b.get("color_net", 0),
                    "color_pct": 0,
                    "treat_count": b.get("treat_count", 0),
                    "treat": b.get("treat_net", 0),
                    "treat_pct": 0,
                    "source": "karissa_tracker_may_2026",
                })

        if missing_global:
            raise ValueError(
                "Tracker is missing rows for these (tab, location) pairs: "
                f"{missing_global}"
            )

        log.info("Loaded %d CUMULATIVE_MTD rows from %s", len(rows), tracker_path.name)
        return rows
    finally:
        wb.close()


def load_historical_data_monthly_rows(
    tracker_path: str | Path,
    customer_config: dict,
) -> list[dict]:
    """Build DATA_MONTHLY-ready rows for 2025-05, 2019-05, 2025-Q2, 2019-Q2.

    Sparse rows — only ``total_sales`` is populated (the source data is
    Karissa's tracker which only has the total). Other numeric columns
    default to 0; ``platform`` is copied from the customer config.

    The YoY report builder reads these by exact ``year_month`` value, so the
    non-monthly "Q2" sentinel year_month strings are intentional.
    """
    tracker_path = Path(tracker_path)
    if not tracker_path.exists():
        raise FileNotFoundError(f"Tracker file not found: {tracker_path}")

    wb = load_workbook(tracker_path, data_only=True, read_only=True)
    try:
        if YOY_TAB not in wb.sheetnames:
            raise ValueError(f"{YOY_TAB!r} tab not found in {tracker_path.name}")
        ws = wb[YOY_TAB]

        # Build a name->row index for Tables 1 and 2 (both use col A for the location name).
        table1_by_name: dict[str, int] = {}
        for r in range(YOY_TABLE1_FIRST_ROW, YOY_TABLE1_LAST_ROW + 1):
            n = _normalize(ws.cell(row=r, column=YOY_TABLE1_NAME_COL).value)
            if n and n.lower() != "totals":
                table1_by_name[n] = r

        table2_by_name: dict[str, int] = {}
        for r in range(YOY_TABLE2_FIRST_ROW, YOY_TABLE2_LAST_ROW + 1):
            n = _normalize(ws.cell(row=r, column=YOY_TABLE2_NAME_COL).value)
            if n and n.lower() != "totals":
                table2_by_name[n] = r

        rows: list[dict] = []
        missing: list[tuple[str, str]] = []  # (year_month_key, loc_name)

        for loc in customer_config["locations"]:
            cfg_name = _normalize(loc["name"])

            # 2025-05
            if cfg_name in table1_by_name:
                r = table1_by_name[cfg_name]
                total_2025 = _to_number(ws.cell(row=r, column=YOY_TABLE1_2025_COL).value)
                total_2019 = _to_number(ws.cell(row=r, column=YOY_TABLE1_2019_COL).value)
                rows.append({
                    "loc_name": loc["name"], "year_month": "2025-05",
                    "platform": loc["platform"], "total_sales": total_2025,
                    "source": "karissa_tracker_may_2026_yoy",
                    "period_start": "2025-05-01", "period_end": "2025-05-31",
                })
                rows.append({
                    "loc_name": loc["name"], "year_month": "2019-05",
                    "platform": loc["platform"], "total_sales": total_2019,
                    "source": "karissa_tracker_may_2026_yoy",
                    "period_start": "2019-05-01", "period_end": "2019-05-31",
                })
            else:
                missing.append(("table1", loc["name"]))

            # Q2 totals
            if cfg_name in table2_by_name:
                r = table2_by_name[cfg_name]
                q2_2025 = _to_number(ws.cell(row=r, column=YOY_TABLE2_Q2_2025_COL).value)
                q2_2019 = _to_number(ws.cell(row=r, column=YOY_TABLE2_Q2_2019_COL).value)
                rows.append({
                    "loc_name": loc["name"], "year_month": "2025-Q2",
                    "platform": loc["platform"], "total_sales": q2_2025,
                    "source": "karissa_tracker_may_2026_yoy",
                    "period_start": "2025-04-01", "period_end": "2025-06-30",
                })
                rows.append({
                    "loc_name": loc["name"], "year_month": "2019-Q2",
                    "platform": loc["platform"], "total_sales": q2_2019,
                    "source": "karissa_tracker_may_2026_yoy",
                    "period_start": "2019-04-01", "period_end": "2019-06-30",
                })
            else:
                missing.append(("table2", loc["name"]))

        if missing:
            raise ValueError(
                f"YoY tab is missing location rows: {missing}. "
                f"Table1 names: {sorted(table1_by_name)}; "
                f"Table2 names: {sorted(table2_by_name)}"
            )

        log.info("Loaded %d DATA_MONTHLY historical rows from %s", len(rows), tracker_path.name)
        return rows
    finally:
        wb.close()


# ─── Pretty-print review tables ───────────────────────────────────────────────

def render_cumulative_review(rows: list[dict]) -> str:
    """Render a per-week table grouped by week_ending for human review."""
    if not rows:
        return "(no CUMULATIVE_MTD rows)"
    by_week: dict[str, list[dict]] = {}
    for r in rows:
        by_week.setdefault(r["week_ending"], []).append(r)
    lines = []
    for we in sorted(by_week.keys()):
        lines.append(f"\n  Week ending {we} ({len(by_week[we])} locations)")
        lines.append(
            f"    {'Location':<16}  {'guests':>7}  {'total_sales':>12}  "
            f"{'service':>10}  {'product':>9}  {'prod_hrs':>8}  "
            f"{'wax_ct':>6} {'wax_$':>8} {'color_$':>9} {'trt_ct':>6} {'trt_$':>8}"
        )
        for r in sorted(by_week[we], key=lambda x: x["loc_name"]):
            lines.append(
                f"    {r['loc_name']:<16}  {r['guests']:>7.0f}  "
                f"{r['total_sales']:>12,.2f}  {r['service']:>10,.2f}  "
                f"{r['product']:>9,.2f}  {r['prod_hours']:>8,.2f}  "
                f"{r['wax_count']:>6.0f} {r['wax']:>8,.2f} "
                f"{r['color']:>9,.2f} {r['treat_count']:>6.0f} {r['treat']:>8,.2f}"
            )
    return "\n".join(lines)


def render_historical_review(rows: list[dict]) -> str:
    """Render historical DATA_MONTHLY rows pivoted by year_month."""
    if not rows:
        return "(no DATA_MONTHLY historical rows)"
    by_ym: dict[str, list[dict]] = {}
    for r in rows:
        by_ym.setdefault(r["year_month"], []).append(r)
    lines = []
    for ym in sorted(by_ym.keys()):
        lines.append(f"\n  year_month={ym} ({len(by_ym[ym])} locations)")
        for r in sorted(by_ym[ym], key=lambda x: x["loc_name"]):
            lines.append(f"    {r['loc_name']:<16}  total_sales = {r['total_sales']:>12,.2f}")
    return "\n".join(lines)


# ─── Sheets write ─────────────────────────────────────────────────────────────

def _load_config(config_path: Path) -> dict:
    return json.loads(config_path.read_text(encoding="utf-8"))


def write_to_sheets(
    cumul_rows: list[dict],
    monthly_rows: list[dict],
    config: dict,
    dry_run: bool,
) -> None:
    from core.sheets_writer import (
        _build_service,
        append_to_cumulative_mtd,
        append_to_data_monthly,
    )

    if dry_run:
        print(">>> DRY RUN — no Sheets write. Would have appended:")
        print(f"    {len(cumul_rows)} rows to CUMULATIVE_MTD")
        print(f"    {len(monthly_rows)} rows to DATA_MONTHLY")
        print(f"    (sheet {config['sheet_id']})\n")
        append_to_cumulative_mtd(service=None, config=config, rows=cumul_rows, dry_run=True)
        append_to_data_monthly(service=None, config=config, rows=monthly_rows, dry_run=True)
        return

    print(
        f">>> WRITING {len(cumul_rows)} CUMULATIVE_MTD rows + "
        f"{len(monthly_rows)} DATA_MONTHLY rows (sheet {config['sheet_id']})"
    )
    service = _build_service(config)
    append_to_cumulative_mtd(service=service, config=config, rows=cumul_rows, dry_run=False)
    append_to_data_monthly(service=service, config=config, rows=monthly_rows, dry_run=False)
    print("    Done.\n")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Seed CUMULATIVE_MTD (May Wk1/Wk2/Wk4) + DATA_MONTHLY "
            "(2025-05, 2019-05, 2025-Q2, 2019-Q2) from Karissa's May 2026 tracker."
        )
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true",
                       help="Parse + render review tables but never write to Sheets")
    group.add_argument("--write", action="store_true",
                       help="Actually write to Sheets after dry-run review")
    parser.add_argument("--tracker", default=None,
                        help=f"Override default tracker path (default: {DEFAULT_TRACKER})")
    parser.add_argument("--config", default=None,
                        help=f"Override default customer config (default: {DEFAULT_CONFIG})")
    parser.add_argument("--verbose", action="store_true", help="Verbose logging")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    tracker_path = Path(args.tracker) if args.tracker else DEFAULT_TRACKER
    config_path = Path(args.config) if args.config else DEFAULT_CONFIG
    config = _load_config(config_path)

    print(f"\n{'=' * 70}")
    print(f"  KARISSA TRACKER BACKFILL — {'DRY RUN' if args.dry_run else 'LIVE WRITE'}")
    print(f"  Tracker: {tracker_path}")
    print(f"  Sheet:   {config['sheet_id']}")
    print(f"{'=' * 70}")

    cumul_rows = load_cumulative_mtd_rows(tracker_path, config)
    monthly_rows = load_historical_data_monthly_rows(tracker_path, config)

    print(f"\nCUMULATIVE_MTD rows: {len(cumul_rows)} "
          f"(target: 12 locations x 3 weeks = 36)")
    print(render_cumulative_review(cumul_rows))

    print(f"\nDATA_MONTHLY historical rows: {len(monthly_rows)} "
          f"(target: 12 locations x 4 periods = 48)")
    print(render_historical_review(monthly_rows))

    print()
    write_to_sheets(cumul_rows, monthly_rows, config, dry_run=args.dry_run)

    if args.dry_run:
        print(">>> Dry run complete. To write to Sheets, re-run with --write.\n")
    else:
        print(">>> Backfill complete.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
