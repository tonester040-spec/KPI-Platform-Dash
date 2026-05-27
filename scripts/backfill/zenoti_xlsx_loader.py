"""
scripts/backfill/zenoti_xlsx_loader.py
──────────────────────────────────────
Cross-validation source for March 2026.

Reads the 9 Zenoti Employee KPI .xlsx files Karissa exported on 2026-04-06
(one per Zenoti location, all dated From: 01 Mar 2026 To: 31 Mar 2026).
Returns ``{loc_name: {invoices, guests, service, product}}`` so the
orchestrator can diff against tracker Week 5 values before any DATA_MONTHLY
write.

These files do NOT have a location in their filename — we identify each
file by scanning for a manager row (e.g. ``"Blaine mgr"``, ``"Andover mgr"``,
``"888-40098-F Sams Roseville, MN_mgr mgr"``).

The Total row layout (validated 2026-05-26 across all 9 files):
  A "Total"  B invoices  C avg_inv_value  D guests  E service_sales
  F avg_svc_value  G product_sales  H avg_prod_value

Dollar amounts (service / product) reliably match Karissa's tracker within
$0.01. Guest counts use a different methodology (Zenoti unique-guests vs
invoice count) and may differ by a few percent — caller should treat
guest-count diffs as informational, not error-level.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)


def _normalize(s: Any) -> str:
    if not s:
        return ""
    return " ".join(str(s).split()).strip().lower()


def _bare_form(loc_norm: str) -> str:
    """Strip the canonical 'Andover FS' → 'andover' bare form used in mgr rows."""
    if loc_norm.endswith(" fs"):
        return loc_norm[:-3]
    if loc_norm.endswith(" full service"):
        return loc_norm[:-13]
    return loc_norm


def _detect_location(ws, zenoti_loc_names: list[str]) -> str | None:
    """Walk every row in column A and return the canonical Zenoti loc name
    whose lowercase form (or bare form, stripping 'FS') appears in a 'mgr' row.

    The mgr rows on Karissa's Zenoti exports always use the bare location
    name (e.g. ``"Andover mgr"`` — not ``"Andover FS mgr"``). So we match
    both the full canonical form and the bare form.

    Handles weird real-world formats like
    ``"888-40098-F Sams Roseville, MN_mgr mgr"`` via substring match.
    """
    candidates: list[tuple[int, str]] = []  # (rank, loc_name); lower rank = stricter match
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell = str(row[0])
        if "mgr" not in cell.lower():
            continue
        cell_norm = _normalize(cell)
        for loc in zenoti_loc_names:
            loc_norm = _normalize(loc)
            bare = _bare_form(loc_norm)
            # Exact "Andover mgr" pattern (rank 0 = strongest)
            if cell_norm == f"{loc_norm} mgr" or cell_norm == f"{bare} mgr":
                candidates.append((0, loc))
                continue
            # Substring match for weird formats (rank 1)
            if loc_norm in cell_norm or bare in cell_norm:
                candidates.append((1, loc))
    if not candidates:
        return None
    # Lowest rank wins; on ties prefer longer (more specific) loc name
    candidates.sort(key=lambda x: (x[0], -len(x[1])))
    return candidates[0][1]


def _read_total_row(ws) -> dict | None:
    """Find the row whose A-cell == 'Total' and return its KPI columns."""
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).strip().lower() == "total":
            return {
                "invoices": row[1] if len(row) > 1 else None,
                "avg_invoice_value": row[2] if len(row) > 2 else None,
                "guests": row[3] if len(row) > 3 else None,
                "service": row[4] if len(row) > 4 else None,
                "avg_service_value": row[5] if len(row) > 5 else None,
                "product": row[6] if len(row) > 6 else None,
                "avg_product_value": row[7] if len(row) > 7 else None,
            }
    return None


def load_zenoti_march_xlsxes(
    xlsx_dir: str | Path,
    customer_config: dict,
) -> dict[str, dict]:
    """Walk the March folder and return ``{loc_name: {invoices, guests, service, product}}``
    for all 9 Zenoti locations.

    Skips non-.xlsx files (the 3 SU Stylist_Tracking_Report .xls files are
    intentionally ignored per Tony's 2026-05-26 decision).

    Raises ValueError if any Zenoti config location can't be matched, or if a
    file has 'mgr' rows but matches no known location.
    """
    xlsx_dir = Path(xlsx_dir)
    if not xlsx_dir.exists():
        raise FileNotFoundError(f"xlsx directory not found: {xlsx_dir}")

    zenoti_locs = [
        loc["name"] for loc in customer_config["locations"]
        if loc["platform"] == "zenoti"
    ]

    out: dict[str, dict] = {}
    unmatched_files: list[str] = []

    for path in sorted(xlsx_dir.glob("*.xlsx")):
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            loc = _detect_location(ws, zenoti_locs)
            if not loc:
                unmatched_files.append(path.name)
                continue
            total = _read_total_row(ws)
            if total is None:
                log.warning("No 'Total' row found in %s — skipping", path.name)
                continue
            out[loc] = {
                "source_file": path.name,
                "invoices": total["invoices"],
                "guests": total["guests"],
                "service": total["service"],
                "product": total["product"],
            }
        finally:
            wb.close()

    missing = [loc for loc in zenoti_locs if loc not in out]
    if missing:
        raise ValueError(
            f"Missing xlsx coverage for Zenoti locations: {missing}. "
            f"Files in dir: {[p.name for p in sorted(xlsx_dir.glob('*.xlsx'))]}. "
            f"Unmatched files: {unmatched_files}"
        )

    log.info("Loaded %d Zenoti monthly xlsx Total rows from %s", len(out), xlsx_dir.name)
    return out


def _read_stylist_rows(ws, loc_name: str, loc_id: str, *,
                       year_month: str, period_start: str, period_end: str) -> list[dict]:
    """Walk every data row in the xlsx, skip Total/mgr/empty/header, emit
    STYLISTS_DATA_MONTHLY row dicts for active stylists only.

    An 'active' stylist is one whose service_sales > 0 OR product_sales > 0.
    Roster rows with all zeros (inactive employees, terminated stylists, etc.)
    are skipped to keep the tab focused.
    """
    out: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        name_lower = name.lower()
        # Skip headers / metadata / total / mgr
        if name_lower in ("employee name", "fantastic sams", "total"):
            continue
        if "employee kpi" in name_lower or "from :" in name_lower:
            continue
        if "mgr" in name_lower:
            continue
        # Need at least 8 cols of data
        if len(row) < 8:
            continue
        try:
            invoices = int(row[1] or 0)
            guests = int(row[3] or 0)
            service = float(row[4] or 0)
            avg_invoice = float(row[2] or 0)
            product = float(row[6] or 0)
        except (TypeError, ValueError):
            continue
        if service <= 0 and product <= 0:
            continue  # inactive stylist roster entry
        out.append({
            "year_month": year_month,
            "name": name,
            "loc_name": loc_name,
            "loc_id": loc_id,
            "platform": "zenoti",
            "invoices": invoices,
            "guests": guests,
            "net_service": service,
            "net_product": product,
            "avg_ticket": avg_invoice,
            "pph": 0,                 # not in xlsx
            "ppg": (product / invoices) if invoices else 0,
            "production_hours": 0,    # not in xlsx
            "source": "zenoti_xlsx",
            "period_start": period_start,
            "period_end": period_end,
            # req_pct + avg_service_time_min aren't in the Employee KPI xlsx
            # export (they're only in the PDF Salon Dashboard's PERFORMANCE
            # DETAILS section). Leave at 0 for March; April onward comes from
            # the PDF loader which captures these.
            "req_pct": 0,
            "avg_service_time_min": 0,
        })
    return out


def load_zenoti_march_stylists(
    xlsx_dir: str | Path,
    customer_config: dict,
    *,
    year_month: str = "2026-03",
    period_start: str = "2026-03-01",
    period_end: str = "2026-03-31",
) -> list[dict]:
    """Walk March xlsxes and emit STYLISTS_DATA_MONTHLY rows for active stylists.

    Inactive stylists (zero service AND zero product) are filtered out.
    Returns rows for all 9 Zenoti locations.
    """
    xlsx_dir = Path(xlsx_dir)
    if not xlsx_dir.exists():
        raise FileNotFoundError(f"xlsx directory not found: {xlsx_dir}")

    zenoti_loc_names = [
        loc["name"] for loc in customer_config["locations"]
        if loc["platform"] == "zenoti"
    ]
    loc_by_name = {
        loc["name"]: loc for loc in customer_config["locations"] if loc["platform"] == "zenoti"
    }

    all_rows: list[dict] = []
    matched_locs: set[str] = set()

    for path in sorted(xlsx_dir.glob("*.xlsx")):
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            loc_name = _detect_location(ws, zenoti_loc_names)
            if not loc_name:
                continue
            cfg = loc_by_name[loc_name]
            stylist_rows = _read_stylist_rows(
                ws, loc_name, cfg["id"],
                year_month=year_month, period_start=period_start, period_end=period_end,
            )
            all_rows.extend(stylist_rows)
            matched_locs.add(loc_name)
        finally:
            wb.close()

    log.info(
        "Loaded %d active stylist rows across %d Zenoti locations for %s",
        len(all_rows), len(matched_locs), year_month,
    )
    return all_rows


def diff_against_tracker(
    tracker_rows: list[dict],
    xlsx_totals: dict[str, dict],
    *,
    dollar_tolerance: float = 0.01,
    guest_tolerance_pct: float = 0.10,
) -> list[dict]:
    """Compare tracker Week 5 rows against Zenoti xlsx Total rows.

    Returns a list of finding dicts:
      {location, severity, code, message, details}

    Severity rules:
      - service/product dollar mismatch > dollar_tolerance → 'error'
      - guest_count differs at all                        → 'info' (methodology diff)
      - guest_count differs by > guest_tolerance_pct      → 'warning'
      - source has only one of the two (xlsx missing or tracker missing) → 'info'
    """
    findings: list[dict] = []
    tracker_by_loc = {r["loc_name"]: r for r in tracker_rows}

    # Only compare locations that appear in both sources (i.e. the 9 Zenoti)
    for loc_name, xlsx_data in xlsx_totals.items():
        if loc_name not in tracker_by_loc:
            findings.append({
                "location": loc_name,
                "severity": "info",
                "code": "MISSING_IN_TRACKER",
                "message": f"{loc_name} present in xlsx but absent from tracker",
                "details": {"xlsx_file": xlsx_data.get("source_file")},
            })
            continue
        t = tracker_by_loc[loc_name]
        x = xlsx_data

        for field in ("service", "product"):
            t_val = float(t.get(field) or 0)
            x_val = float(x.get(field) or 0)
            delta = round(x_val - t_val, 2)
            if abs(delta) > dollar_tolerance:
                findings.append({
                    "location": loc_name,
                    "severity": "error",
                    "code": f"{field.upper()}_MISMATCH",
                    "message": (
                        f"{loc_name} {field}: tracker=${t_val:,.2f} vs xlsx=${x_val:,.2f} "
                        f"(delta={delta:+,.2f}, tol=${dollar_tolerance:.2f})"
                    ),
                    "details": {"tracker": t_val, "xlsx": x_val, "delta": delta},
                })

        # Guest count — informational only (methodology differs)
        t_guests = float(t.get("guests") or 0)
        x_guests = float(x.get("guests") or 0)
        if t_guests != x_guests:
            denom = max(t_guests, 1)
            pct = abs(x_guests - t_guests) / denom
            sev = "warning" if pct > guest_tolerance_pct else "info"
            findings.append({
                "location": loc_name,
                "severity": sev,
                "code": "GUEST_COUNT_DIFF",
                "message": (
                    f"{loc_name} guests: tracker={int(t_guests)} vs xlsx={int(x_guests)} "
                    f"(diff={int(x_guests - t_guests):+}, {pct:.1%}; "
                    f"known methodology diff — Zenoti unique vs invoice count)"
                ),
                "details": {"tracker": int(t_guests), "xlsx": int(x_guests), "pct": pct},
            })

    return findings
