"""
core/karissa_workbook.py
KPI Platform — builds the monthly Excel workbook in Karissa's exact layout.

Replaces the legacy ``core/report_builder.py`` six-tab format with the
layout Karissa actually uses in her own tracker:

  Week 1, Week 2, Week 3, Week 4, Week 5  — five weekly cumulative-MTD tabs
                                            (only the weeks she has data for
                                             are populated; others are empty
                                             templates)
  Year Over Year                          — current-month YoY + Q2 YoY +
                                            Q2 monthly breakdown sub-tables

Entry point:
  ``build_monthly_workbook(year_month, output_path, service, config)`` ->
  Path. Reads CUMULATIVE_MTD (weekly snapshots), DATA_MONTHLY (historical
  totals + prior-month reference), and MONTHLY_GOALS (per-location monthly
  $ target + divisors) from Google Sheets. Writes a static-value workbook
  (no Excel formulas — every cell is a pre-computed number per Z-mode).

Z-mode contract: derived ratios are recomputed canonically from raw
primitives per CLAUDE.md, never copied from a tracker. Where Karissa's
tracker uses POS-reported PPG/PPH for Zenoti locations, our output
will differ — that's intentional.
"""

from __future__ import annotations

import datetime as _dt
import logging
from calendar import monthrange
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core import data_source

log = logging.getLogger(__name__)


# ─── Constants ────────────────────────────────────────────────────────────────

# Karissa's tab labels are intentionally non-uniform — match her file byte-for-byte.
WEEK_TAB_NAMES = [" Week 1 ", " week 2", " week 3", " week 4", " week 5"]

# Service-mix treatment-penetration goal annotation under bottom Totals.
GOAL_ANNOTATION_LABEL = "Goal is 15% plus"
PERCENT_OF_GUESTS_LABEL = "% of Guests"

# Weekday integers per Python's date.weekday(): Monday=0 ... Sunday=6.
# Karissa's working days are Mon-Sat (closed Sundays).
_SUNDAY_WEEKDAY = 6

# Karissa's Projection formula on every weekly tab is
#     =(total_sales / working_days_so_far) * working_days_in_month
# but her counts don't quite match a pure Mon-Sat calculation — she also
# adjusts for things like Memorial Day (May 25, 2026) closing one Monday.
# For months she's filled by hand we hardcode her observed divisors; for
# unknown months we fall back to Mon-Sat from the date.
#
# Schema: {"YYYY-MM": {week_idx: (days_elapsed, days_total)}}
# (Week 5 always uses the full-month total as both numerator-shadow and total.)
PROJECTION_DIVISORS_BY_MONTH: dict[str, dict[int, tuple[int, int]]] = {
    "2026-05": {
        1: (2, 25),
        2: (8, 25),
        3: (14, 25),   # Wk3 unfilled in her May tracker; inferred from her +6 pattern
        4: (21, 25),
        5: (25, 25),
    },
}


# ─── Calendar helpers ─────────────────────────────────────────────────────────

def _sundays_of_month(year: int, month: int) -> list[_dt.date]:
    """Return every Sunday that falls within ``year-month``, in order."""
    _, n_days = monthrange(year, month)
    out: list[_dt.date] = []
    for d in range(1, n_days + 1):
        dt = _dt.date(year, month, d)
        if dt.weekday() == _SUNDAY_WEEKDAY:
            out.append(dt)
    return out


def _working_days_in_range(start: _dt.date, end_inclusive: _dt.date) -> int:
    """Count Mon-Sat days from ``start`` through ``end_inclusive`` (inclusive)."""
    if end_inclusive < start:
        return 0
    n = 0
    d = start
    while d <= end_inclusive:
        if d.weekday() != _SUNDAY_WEEKDAY:
            n += 1
        d += _dt.timedelta(days=1)
    return n


def _working_days_in_month(year: int, month: int) -> int:
    _, n = monthrange(year, month)
    return _working_days_in_range(_dt.date(year, month, 1),
                                  _dt.date(year, month, n))


def _parse_year_month(year_month: str) -> tuple[int, int]:
    y, m = year_month.split("-")
    return int(y), int(m)


def _parse_iso_date(s: str) -> _dt.date:
    return _dt.datetime.strptime(s, "%Y-%m-%d").date()


# ─── Z-mode canonical recompute ───────────────────────────────────────────────

def _safe_div(num: float, denom: float) -> float:
    return num / denom if denom else 0.0


def _recompute_canonical(raw: dict) -> dict:
    """Given raw primitives (guests, total_sales, service, product, prod_hours,
    wax_count, wax, color, treat_count, treat), return a dict with the canonical
    derived ratios per CLAUDE.md.

    Per-cell decimals are rounded to 2dp for PPG/PPH/AT to match Karissa's
    typed precision; percentages stay at full precision (the cell format
    handles display rounding).
    """
    g = raw.get("guests", 0) or 0
    ts = raw.get("total_sales", 0) or 0
    sv = raw.get("service", 0) or 0
    pr = raw.get("product", 0) or 0
    ph = raw.get("prod_hours", 0) or 0
    return {
        "product_pct": _safe_div(pr, ts),
        "ppg":         round(_safe_div(pr, g), 2),
        "pph":         round(_safe_div(sv, ph), 2),
        "avg_ticket":  round(_safe_div(ts, g), 2),
        "wax_pct":     _safe_div(raw.get("wax_count", 0) or 0, g),
        "color_pct":   _safe_div(raw.get("color", 0) or 0, sv),
        "treat_pct":   _safe_div(raw.get("treat_count", 0) or 0, g),
    }


# ─── Styling helpers ──────────────────────────────────────────────────────────
#
# Color palette extracted from Karissa's May 2026.xlsx tracker (see ARGB values
# in column dumps via openpyxl):
#   Cyan    #00B0F0   header banner (row 1 across; top + bottom block header row)
#   Green   #92D050   right-side reference section headers (2025 ref + prior-month)
#   Peach   #FBE4D5   "Diff" / "Dif" column headers
#   Yellow  #FEF2CB   row 16 network averages (diff cells)
#   LtGreen #E2EFD9   row 16 trailing accent (AT avg)

FILL_CYAN   = PatternFill("solid", fgColor="00B0F0") if False else None  # lazy-init below
FILL_GREEN  = None
FILL_PEACH  = None
FILL_YELLOW = None
FILL_LTGRN  = None


def _init_fills():
    """Lazy-init the PatternFill objects (openpyxl may not be available at import)."""
    global FILL_CYAN, FILL_GREEN, FILL_PEACH, FILL_YELLOW, FILL_LTGRN
    if FILL_CYAN is None:
        FILL_CYAN   = PatternFill("solid", fgColor="00B0F0")
        FILL_GREEN  = PatternFill("solid", fgColor="92D050")
        FILL_PEACH  = PatternFill("solid", fgColor="FBE4D5")
        FILL_YELLOW = PatternFill("solid", fgColor="FEF2CB")
        FILL_LTGRN  = PatternFill("solid", fgColor="E2EFD9")


def _border() -> Border:
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _bold() -> Font:
    return Font(bold=True)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


# ─── Per-week column layouts (1-based) ────────────────────────────────────────
#
# Karissa's weekly tabs have small, intentional variations:
#   - Week 1 has an extra "tickets" column at B (top block) and a blank col B
#     (bottom block) — everything shifts +1 right.
#   - Week 4 inserts a "Goals" column in the top block before the 2025
#     comparison columns.
#   - Week 5 renames "Projection(s)" to "Monthly goal" in the same column.

def _top_columns(week_idx: int) -> dict[str, int]:
    """Return 1-based column positions for the top KPI block of ``week_idx``."""
    cols: dict[str, int] = {"name": 1}
    c = 2
    if week_idx == 1:
        cols["tickets"] = c; c += 1
    cols["guests"]      = c; c += 1
    cols["total_sales"] = c; c += 1
    cols["service"]     = c; c += 1
    cols["product"]     = c; c += 1
    cols["product_pct"] = c; c += 1
    cols["ppg"]         = c; c += 1
    cols["pph"]         = c; c += 1
    cols["avg_ticket"]  = c; c += 1
    cols["prod_hours"]  = c; c += 1
    cols["projection"]  = c; c += 1
    if week_idx == 4:
        cols["goal"] = c; c += 1
    cols["y2025_total"]  = c; c += 1
    cols["y2025_guests"] = c; c += 1
    cols["y2025_ppg"]    = c; c += 1
    cols["ppg_diff"]     = c; c += 1
    cols["y2025_at"]     = c; c += 1
    cols["at_diff"]      = c; c += 1
    return cols


def _top_headers(week_idx: int) -> dict[str, str]:
    """Header labels for row 2 of the top block, matching Karissa's exact wording
    (including stray leading/trailing whitespace she has in her template)."""
    h: dict[str, str] = {
        "name":         "Name",
        "guests":       "Guest Count",
        "total_sales":  "Total Sales Net",
        "service":      "Service Net",
        "product":      "Product Net",
        "product_pct":  "Product %",
        "ppg":          "PPG Net",
        "pph":          "PPH Net",
        "avg_ticket":   "Average Tkt",
        "prod_hours":   "Prod Hours",
        "y2025_total":  "2025 Total Sales",
        "y2025_guests": " Guest Count",   # Karissa: leading space
        "y2025_ppg":    "PPG",
        "ppg_diff":     "Diff",
        "y2025_at":     " AT",            # Karissa: leading space
        "at_diff":      "Dif",
    }
    if week_idx == 1:
        h["tickets"] = "tickets "         # Karissa: trailing space
        h["projection"] = "Projection"
    elif week_idx == 5:
        h["projection"] = "Monthly goal"
    else:
        h["projection"] = "Projections"
    if week_idx == 4:
        h["goal"] = "Goals "              # Karissa: trailing space
    return h


def _bottom_columns(week_idx: int) -> dict[str, int]:
    """1-based columns for the bottom service-mix block."""
    cols: dict[str, int] = {"name": 1}
    c = 2
    if week_idx == 1:
        cols["blank_b"] = c; c += 1
    cols["wax_count"]   = c; c += 1
    cols["wax_net"]     = c; c += 1
    cols["wax_pct"]     = c; c += 1
    cols["color_net"]   = c; c += 1
    cols["color_pct"]   = c; c += 1
    cols["treat_count"] = c; c += 1
    cols["treat_net"]   = c; c += 1
    cols["treat_pct"]   = c; c += 1
    # Karissa leaves one blank separator column before the right-side reference block.
    c += 1
    cols["prior_prod_hours"] = c; c += 1
    cols["prior_wax_net"]    = c; c += 1
    cols["prior_treat_net"]  = c; c += 1
    cols["prior_service"]    = c; c += 1
    cols["prior_treat_pct"]  = c; c += 1
    cols["curr_treat_pct"]   = c; c += 1
    cols["prior_wax_pct"]    = c; c += 1
    cols["curr_wax_pct"]     = c; c += 1
    return cols


def _bottom_headers(week_idx: int) -> dict:
    """Bottom-block header labels — match Karissa's exact tokens.

    Karissa enters the right-side "2026" labels as numbers (not strings),
    so they get stored as the int ``2026`` instead of the string ``"2026"``.
    """
    return {
        "wax_count":        "Wax Count",
        "wax_net":          "Waxing Net",
        "wax_pct":          "Wax %",
        "color_net":        "Color Net",
        "color_pct":        "Color %",
        "treat_count":      "Trmt Count",
        "treat_net":        "Trmt Net",
        "treat_pct":        "Trmt %",
        "prior_prod_hours": " Prod Hours",
        "prior_wax_net":    "Wax ",
        "prior_treat_net":  "Treat ",
        "prior_service":    "Service ",
        "prior_treat_pct":  "Treat %",
        "curr_treat_pct":   2026,
        "prior_wax_pct":    "Wax%",
        "curr_wax_pct":     2026,
    }


# ─── Number formats ───────────────────────────────────────────────────────────

FMT_INT     = "#,##0"
FMT_MONEY   = "#,##0.00"
FMT_PCT     = "0.00%"
FMT_DECIMAL = "0.00"
FMT_DATE    = "yyyy-mm-dd"


# ─── Weekly tab builder ───────────────────────────────────────────────────────

def _set(ws, row: int, col: int, value, *, fmt: str | None = None, bold: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = _bold()
    return cell


def _build_weekly_tab(
    wb,
    week_idx: int,
    week_ending: _dt.date | None,
    locations: list[dict],
    snapshots_for_week: dict[str, dict],
    prior_month_by_loc: dict[str, dict],
    monthly_goal_by_loc: dict[str, dict],
    year_month: str,
    working_days_in_month: int,
):
    """Build one weekly tab. ``snapshots_for_week`` is {loc_name: snapshot}
    keyed by canonical config name. Empty dict means this week has no data
    yet — the tab is rendered as a header-only template.
    """
    ws = wb.create_sheet(title=WEEK_TAB_NAMES[week_idx - 1])

    top_cols = _top_columns(week_idx)
    bot_cols = _bottom_columns(week_idx)
    top_hdr  = _top_headers(week_idx)
    bot_hdr  = _bottom_headers(week_idx)

    # ── Row 1: date in A1 ──
    if week_ending is not None:
        _set(ws, 1, 1, week_ending, fmt=FMT_DATE)
    # Karissa places "Region #" directly above her "Total Sales Net" column.
    _set(ws, 1, top_cols["total_sales"], "Region #", bold=True)
    # "End of month" decorative label sits above Projection — but not on Wk5
    # (where col K is "Monthly goal" instead of "Projection").
    if week_idx != 5:
        _set(ws, 1, top_cols["projection"], "End of month", bold=True)

    # ── Row 2: top block headers ──
    for key, col in top_cols.items():
        if key in top_hdr:
            _set(ws, 2, col, top_hdr[key], bold=True)

    # ── Rows 3..14: per-location data ──
    totals = {
        "guests": 0, "total_sales": 0.0, "service": 0.0, "product": 0.0,
        "ppg": 0.0, "pph": 0.0, "avg_ticket": 0.0, "prod_hours": 0.0,
        "projection": 0.0, "goal": 0.0,
        "y2025_total": 0.0, "y2025_guests": 0, "y2025_ppg": 0.0,
        "ppg_diff": 0.0, "y2025_at": 0.0, "at_diff": 0.0,
    }

    for i, loc in enumerate(locations):
        row = 3 + i
        name = loc["name"]
        _set(ws, row, top_cols["name"], name)

        # Tickets col on Wk1 — Karissa-internal region label; leave blank to avoid
        # encoding meaning we don't actually own.
        # (column already empty; nothing to write)

        snap = snapshots_for_week.get(name) or {}
        ratios = _recompute_canonical(snap) if snap else {}

        if snap:
            _set(ws, row, top_cols["guests"],      snap.get("guests", 0), fmt=FMT_INT)
            _set(ws, row, top_cols["total_sales"], snap.get("total_sales", 0.0), fmt=FMT_MONEY)
            _set(ws, row, top_cols["service"],     snap.get("service", 0.0), fmt=FMT_MONEY)
            _set(ws, row, top_cols["product"],     snap.get("product", 0.0), fmt=FMT_MONEY)
            _set(ws, row, top_cols["product_pct"], ratios["product_pct"], fmt=FMT_PCT)
            _set(ws, row, top_cols["ppg"],         ratios["ppg"], fmt=FMT_DECIMAL)
            _set(ws, row, top_cols["pph"],         ratios["pph"], fmt=FMT_DECIMAL)
            _set(ws, row, top_cols["avg_ticket"],  ratios["avg_ticket"], fmt=FMT_DECIMAL)
            _set(ws, row, top_cols["prod_hours"],  snap.get("prod_hours", 0.0), fmt=FMT_DECIMAL)

            # Projection — Karissa's cumulative formula: (cumul / working_days_so_far) * working_days_in_month.
            # Prefer her hand-curated divisors when we have them; fall back to Mon-Sat.
            if week_ending is not None:
                override = PROJECTION_DIVISORS_BY_MONTH.get(year_month, {}).get(week_idx)
                if override:
                    days_elapsed, days_total = override
                else:
                    month_first = _dt.date(week_ending.year, week_ending.month, 1)
                    days_elapsed = max(_working_days_in_range(month_first, week_ending), 1)
                    days_total = working_days_in_month
                projection = (snap.get("total_sales", 0.0) / days_elapsed) * days_total
                _set(ws, row, top_cols["projection"], projection, fmt=FMT_MONEY)
                totals["projection"] += projection

            totals["guests"]      += snap.get("guests", 0) or 0
            totals["total_sales"] += snap.get("total_sales", 0.0) or 0.0
            totals["service"]     += snap.get("service", 0.0) or 0.0
            totals["product"]     += snap.get("product", 0.0) or 0.0
            totals["ppg"]         += ratios["ppg"]
            totals["pph"]         += ratios["pph"]
            totals["avg_ticket"]  += ratios["avg_ticket"]
            totals["prod_hours"]  += snap.get("prod_hours", 0.0) or 0.0

        # Week 4: Goals column — from MONTHLY_GOALS tab (per Karissa).
        if week_idx == 4:
            mg = monthly_goal_by_loc.get(name) or {}
            goal = mg.get("monthly_goal", 0.0) or 0.0
            if goal:
                _set(ws, row, top_cols["goal"], goal, fmt=FMT_MONEY)
                totals["goal"] += goal

        # 2025 reference columns — Karissa shows these on EVERY week (even
        # weeks she hasn't filled current data for). Sourced from DATA_MONTHLY
        # 2025-MM rows (which now include guests/ppg/avg_ticket from her Wk1
        # right-side reference block).
        prior_ym = f"{week_ending.year - 1}-{week_ending.month:02d}" if week_ending else f"2025-{int(year_month.split('-')[1]):02d}"
        prior_2025 = prior_month_by_loc.get((prior_ym, name)) or {}
        if prior_2025:
            ts_2025 = prior_2025.get("total_sales", 0.0) or 0.0
            g_2025  = prior_2025.get("guests", 0) or 0
            ppg_2025 = prior_2025.get("ppg", 0.0) or 0.0
            at_2025  = prior_2025.get("avg_ticket", 0.0) or 0.0
            _set(ws, row, top_cols["y2025_total"], ts_2025, fmt=FMT_MONEY)
            if g_2025:
                _set(ws, row, top_cols["y2025_guests"], g_2025, fmt=FMT_INT)
            if ppg_2025:
                _set(ws, row, top_cols["y2025_ppg"], ppg_2025, fmt=FMT_DECIMAL)
            if at_2025:
                _set(ws, row, top_cols["y2025_at"], at_2025, fmt=FMT_DECIMAL)
            # PPG/AT diff = current canonical - 2025 reference.
            if snap and ppg_2025:
                _set(ws, row, top_cols["ppg_diff"], ratios["ppg"] - ppg_2025, fmt=FMT_DECIMAL)
            elif ppg_2025:
                # Empty current week — Karissa shows -<2025_value> here.
                _set(ws, row, top_cols["ppg_diff"], -ppg_2025, fmt=FMT_DECIMAL)
            if snap and at_2025:
                _set(ws, row, top_cols["at_diff"], ratios["avg_ticket"] - at_2025, fmt=FMT_DECIMAL)
            elif at_2025:
                _set(ws, row, top_cols["at_diff"], -at_2025, fmt=FMT_DECIMAL)
            totals["y2025_total"] += ts_2025
            totals["y2025_guests"] += g_2025

    # ── Row 15: Totals ──
    totals_row = 15
    _set(ws, totals_row, top_cols["name"], "Totals", bold=True)
    if any(totals.values()):
        _set(ws, totals_row, top_cols["guests"],      totals["guests"], fmt=FMT_INT, bold=True)
        _set(ws, totals_row, top_cols["total_sales"], totals["total_sales"], fmt=FMT_MONEY, bold=True)
        _set(ws, totals_row, top_cols["service"],     totals["service"], fmt=FMT_MONEY, bold=True)
        _set(ws, totals_row, top_cols["product"],     totals["product"], fmt=FMT_MONEY, bold=True)
        # Product % of network = sum(product) / sum(total_sales)
        if totals["total_sales"]:
            _set(ws, totals_row, top_cols["product_pct"],
                 totals["product"] / totals["total_sales"], fmt=FMT_PCT, bold=True)
        # PPG, PPH, AT totals row in Karissa's file is the COLUMN SUM, not network avg.
        _set(ws, totals_row, top_cols["ppg"],        totals["ppg"], fmt=FMT_DECIMAL, bold=True)
        _set(ws, totals_row, top_cols["pph"],        totals["pph"], fmt=FMT_DECIMAL, bold=True)
        _set(ws, totals_row, top_cols["avg_ticket"], totals["avg_ticket"], fmt=FMT_DECIMAL, bold=True)
        _set(ws, totals_row, top_cols["prod_hours"], totals["prod_hours"], fmt=FMT_DECIMAL, bold=True)
        _set(ws, totals_row, top_cols["projection"], totals["projection"], fmt=FMT_MONEY, bold=True)
        if week_idx == 4 and totals["goal"]:
            _set(ws, totals_row, top_cols["goal"], totals["goal"], fmt=FMT_MONEY, bold=True)
        if totals["y2025_total"]:
            _set(ws, totals_row, top_cols["y2025_total"], totals["y2025_total"], fmt=FMT_MONEY, bold=True)
        if totals["y2025_guests"]:
            _set(ws, totals_row, top_cols["y2025_guests"], totals["y2025_guests"], fmt=FMT_INT, bold=True)

    # ── Row 16: network averages (column-by-12) for PPG, PPH, AT ──
    n = len(locations)
    if snapshots_for_week and n:
        _set(ws, 16, top_cols["ppg"],        totals["ppg"] / n, fmt=FMT_DECIMAL)
        _set(ws, 16, top_cols["pph"],        totals["pph"] / n, fmt=FMT_DECIMAL)
        _set(ws, 16, top_cols["avg_ticket"], totals["avg_ticket"] / n, fmt=FMT_DECIMAL)

    # ── Row 18: bottom block date + headers ──
    if week_ending is not None:
        _set(ws, 18, 1, week_ending, fmt=FMT_DATE)
    for key, col in bot_cols.items():
        if key in bot_hdr:
            _set(ws, 18, col, bot_hdr[key], bold=True)

    # ── Rows 19..30: per-location service-mix data ──
    bot_totals = {
        "wax_count": 0, "wax_net": 0.0, "color_net": 0.0,
        "treat_count": 0, "treat_net": 0.0,
        "prior_prod_hours": 0.0, "prior_wax_net": 0.0, "prior_treat_net": 0.0,
        "prior_service": 0.0,
    }
    # Track current-month penetration aggregate (treatment %) for the footer.
    cum_treat_count = 0
    cum_guests = 0

    for i, loc in enumerate(locations):
        row = 19 + i
        name = loc["name"]
        _set(ws, row, bot_cols["name"], name)

        snap = snapshots_for_week.get(name) or {}
        if snap:
            wc = snap.get("wax_count", 0) or 0
            wn = snap.get("wax", 0.0) or 0.0
            cn = snap.get("color", 0.0) or 0.0
            tc = snap.get("treat_count", 0) or 0
            tn = snap.get("treat", 0.0) or 0.0
            g  = snap.get("guests", 0) or 0
            sv = snap.get("service", 0.0) or 0.0

            _set(ws, row, bot_cols["wax_count"],   wc, fmt=FMT_INT)
            _set(ws, row, bot_cols["wax_net"],     wn, fmt=FMT_MONEY)
            _set(ws, row, bot_cols["wax_pct"],     _safe_div(wc, g), fmt=FMT_PCT)
            _set(ws, row, bot_cols["color_net"],   cn, fmt=FMT_MONEY)
            _set(ws, row, bot_cols["color_pct"],   _safe_div(cn, sv), fmt=FMT_PCT)
            _set(ws, row, bot_cols["treat_count"], tc, fmt=FMT_INT)
            _set(ws, row, bot_cols["treat_net"],   tn, fmt=FMT_MONEY)
            _set(ws, row, bot_cols["treat_pct"],   _safe_div(tc, g), fmt=FMT_PCT)

            bot_totals["wax_count"]   += wc
            bot_totals["wax_net"]     += wn
            bot_totals["color_net"]   += cn
            bot_totals["treat_count"] += tc
            bot_totals["treat_net"]   += tn
            cum_treat_count += tc
            cum_guests += g

        # Prior-month right-side reference columns (always populated — these
        # come from DATA_MONTHLY for the previous full month, independent of
        # whether this week has snapshot data yet).
        prior_apr = prior_month_by_loc.get(("2026-04", name)) or {}
        if prior_apr:
            ph = prior_apr.get("prod_hours", 0.0) or 0.0
            pw = prior_apr.get("wax", 0.0) or 0.0
            pt = prior_apr.get("treat", 0.0) or 0.0
            ps = prior_apr.get("service", 0.0) or 0.0
            _set(ws, row, bot_cols["prior_prod_hours"], ph, fmt=FMT_DECIMAL)
            _set(ws, row, bot_cols["prior_wax_net"],   pw, fmt=FMT_MONEY)
            _set(ws, row, bot_cols["prior_treat_net"], pt, fmt=FMT_MONEY)
            _set(ws, row, bot_cols["prior_service"],   ps, fmt=FMT_MONEY)
            _set(ws, row, bot_cols["prior_treat_pct"], _safe_div(pt, ps), fmt=FMT_PCT)
            _set(ws, row, bot_cols["prior_wax_pct"],   _safe_div(pw, ps), fmt=FMT_PCT)
            bot_totals["prior_prod_hours"] += ph
            bot_totals["prior_wax_net"]    += pw
            bot_totals["prior_treat_net"]  += pt
            bot_totals["prior_service"]    += ps

        # Current cumulative wax%/treat% on the right side (wax_$ / service_$,
        # treat_$ / service_$ — these are revenue-share metrics, NOT the
        # penetration percentages on the left side).
        if snap:
            sv = snap.get("service", 0.0) or 0.0
            _set(ws, row, bot_cols["curr_treat_pct"], _safe_div(snap.get("treat", 0.0) or 0.0, sv), fmt=FMT_PCT)
            _set(ws, row, bot_cols["curr_wax_pct"],   _safe_div(snap.get("wax", 0.0) or 0.0, sv), fmt=FMT_PCT)

    # ── Row 31: Totals (bottom block) ──
    tot_row = 31
    _set(ws, tot_row, bot_cols["name"], "Totals", bold=True)
    _set(ws, tot_row, bot_cols["wax_count"], bot_totals["wax_count"], fmt=FMT_INT, bold=True)
    _set(ws, tot_row, bot_cols["wax_net"],   bot_totals["wax_net"], fmt=FMT_MONEY, bold=True)
    if cum_guests:
        _set(ws, tot_row, bot_cols["wax_pct"], bot_totals["wax_count"] / cum_guests, fmt=FMT_PCT, bold=True)
    _set(ws, tot_row, bot_cols["color_net"],   bot_totals["color_net"], fmt=FMT_MONEY, bold=True)
    # Network color % = sum(color) / sum(service) — same denominator rule, scaled up.
    cum_service = sum(
        (snapshots_for_week.get(loc["name"]) or {}).get("service", 0.0) or 0.0
        for loc in locations
    )
    if cum_service:
        _set(ws, tot_row, bot_cols["color_pct"],
             bot_totals["color_net"] / cum_service, fmt=FMT_PCT, bold=True)
    _set(ws, tot_row, bot_cols["treat_count"], bot_totals["treat_count"], fmt=FMT_INT, bold=True)
    _set(ws, tot_row, bot_cols["treat_net"],   bot_totals["treat_net"], fmt=FMT_MONEY, bold=True)
    if cum_guests:
        _set(ws, tot_row, bot_cols["treat_pct"],
             bot_totals["treat_count"] / cum_guests, fmt=FMT_PCT, bold=True)

    # Right-side Totals row (prior + current ratios for the network)
    _set(ws, tot_row, bot_cols["prior_prod_hours"], bot_totals["prior_prod_hours"], fmt=FMT_DECIMAL, bold=True)
    _set(ws, tot_row, bot_cols["prior_wax_net"],   bot_totals["prior_wax_net"], fmt=FMT_MONEY, bold=True)
    _set(ws, tot_row, bot_cols["prior_treat_net"], bot_totals["prior_treat_net"], fmt=FMT_MONEY, bold=True)
    _set(ws, tot_row, bot_cols["prior_service"],   bot_totals["prior_service"], fmt=FMT_MONEY, bold=True)
    if bot_totals["prior_service"]:
        _set(ws, tot_row, bot_cols["prior_treat_pct"],
             bot_totals["prior_treat_net"] / bot_totals["prior_service"], fmt=FMT_PCT, bold=True)
        _set(ws, tot_row, bot_cols["prior_wax_pct"],
             bot_totals["prior_wax_net"] / bot_totals["prior_service"], fmt=FMT_PCT, bold=True)
    if cum_service:
        cum_treat = sum(
            (snapshots_for_week.get(loc["name"]) or {}).get("treat", 0.0) or 0.0
            for loc in locations
        )
        cum_wax = sum(
            (snapshots_for_week.get(loc["name"]) or {}).get("wax", 0.0) or 0.0
            for loc in locations
        )
        _set(ws, tot_row, bot_cols["curr_treat_pct"], cum_treat / cum_service, fmt=FMT_PCT, bold=True)
        _set(ws, tot_row, bot_cols["curr_wax_pct"],   cum_wax / cum_service, fmt=FMT_PCT, bold=True)

    # ── Row 32+: footer annotation (Karissa's verbatim "% of Guests" + "Goal is 15% plus") ──
    if cum_guests:
        _set(ws, 32, bot_cols["color_pct"], PERCENT_OF_GUESTS_LABEL)
        _set(ws, 32, bot_cols["treat_count"], bot_totals["treat_count"] / cum_guests, fmt=FMT_PCT)
        _set(ws, 33, bot_cols["color_pct"], GOAL_ANNOTATION_LABEL, bold=True)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 16
    # Top-block numeric columns
    for col in range(2, max(top_cols.values()) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # ── Apply Karissa's color schema ──
    _init_fills()
    top_max = max(top_cols.values())
    bot_max = max(bot_cols.values())
    width = max(top_max, bot_max)

    # Row 1: cyan banner across the full width of the top block.
    for c in range(1, top_max + 1):
        ws.cell(row=1, column=c).fill = FILL_CYAN

    # Row 2: cyan for the top-block KPI headers (cols up to and including the
    # left-side block — i.e. through "Projection" and Wk4's "Goals" col).
    left_block_last = top_cols.get("goal", top_cols["projection"])
    for c in range(top_cols["guests"], left_block_last + 1):
        ws.cell(row=2, column=c).fill = FILL_CYAN
    # Row 2: green for the 2025 reference section (excluding Diff / Dif).
    for key in ("y2025_total", "y2025_guests", "y2025_ppg", "y2025_at"):
        if key in top_cols:
            ws.cell(row=2, column=top_cols[key]).fill = FILL_GREEN
    # Row 2: peach for the diff column headers.
    for key in ("ppg_diff", "at_diff"):
        if key in top_cols:
            ws.cell(row=2, column=top_cols[key]).fill = FILL_PEACH

    # Row 16: cyan network-average band for PPG / PPH / Avg Tkt; yellow for
    # the right-side diff averages; light-green accent on the trailing AT diff.
    for key in ("ppg", "pph", "avg_ticket", "prod_hours"):
        if key in top_cols:
            ws.cell(row=16, column=top_cols[key]).fill = FILL_CYAN
    if "ppg_diff" in top_cols:
        ws.cell(row=16, column=top_cols["ppg_diff"]).fill = FILL_YELLOW
    if "at_diff" in top_cols:
        ws.cell(row=16, column=top_cols["at_diff"]).fill = FILL_LTGRN

    # Row 18: cyan for the bottom-block service-mix headers (Wax/Color/Trmt
    # primaries); green for the right-side prior-month + current-month
    # reference headers.
    primary_keys = ("wax_count", "wax_net", "wax_pct", "color_net", "color_pct",
                    "treat_count", "treat_net", "treat_pct")
    ws.cell(row=18, column=1).fill = FILL_CYAN  # date cell
    for key in primary_keys:
        if key in bot_cols:
            ws.cell(row=18, column=bot_cols[key]).fill = FILL_CYAN
    for key in ("prior_prod_hours", "prior_wax_net", "prior_treat_net",
                "prior_service", "prior_treat_pct", "prior_wax_pct"):
        if key in bot_cols:
            ws.cell(row=18, column=bot_cols[key]).fill = FILL_GREEN


# ─── Year Over Year tab builder ───────────────────────────────────────────────

def _build_yoy_tab(
    wb,
    year_month: str,
    locations: list[dict],
    latest_snapshot_by_loc: dict[str, dict],
    historical_by_key: dict[tuple, dict],   # (year_month, loc_name) -> row
    monthly_goal_by_loc: dict[str, dict],
):
    """Build the Year Over Year tab — three sub-tables matching Karissa's layout.

    Table 1 (rows 2-13): current-month YoY + pacing.
    Table 2 (rows 18-29): Q2 YoY (2026 / 2025 / 2019).
    Table 3 (rows 19-30, cols J-N): Q2 monthly breakdown (Apr / May / Jun / Q2 Total).
    """
    ws = wb.create_sheet(title="Year Over Year")

    # Determine reference dates from year_month.
    year, month = _parse_year_month(year_month)
    # First-Sunday of the month is what Karissa puts in A1.
    first_sun = next((d for d in _sundays_of_month(year, month)), None)
    if first_sun is not None:
        _set(ws, 1, 1, first_sun, fmt=FMT_DATE)

    # ── Row 1: Table 1 headers ──
    _set(ws, 1, 2,  year,            bold=True)
    _set(ws, 1, 3,  year - 1,        bold=True)
    _set(ws, 1, 4,  "Sales %",       bold=True)
    _set(ws, 1, 5,  " Goal",         bold=True)
    _set(ws, 1, 6,  "Daily Goal",    bold=True)
    _set(ws, 1, 7,  "Day Goal",      bold=True)
    _set(ws, 1, 8,  "We are At!",    bold=True)
    _set(ws, 1, 9,  " Goal %",       bold=True)
    _set(ws, 1, 10, 2019,            bold=True)
    _set(ws, 1, 11, "Sales %",       bold=True)

    # ── Rows 2..13: Table 1 — per-location current month YoY ──
    tot = {
        "y2026": 0.0, "y2025": 0.0, "y2019": 0.0,
        "monthly_goal": 0.0, "daily_goal": 0.0, "day_goal": 0.0, "we_are_at": 0.0,
    }
    prior_year_month = f"{year - 1:04d}-{month:02d}"
    pp_year_month    = f"2019-{month:02d}"

    for i, loc in enumerate(locations):
        row = 2 + i
        name = loc["name"]
        _set(ws, row, 1, name)

        cur_mtd = (latest_snapshot_by_loc.get(name) or {}).get("total_sales", 0.0) or 0.0
        py = (historical_by_key.get((prior_year_month, name)) or {}).get("total_sales", 0.0) or 0.0
        pp = (historical_by_key.get((pp_year_month, name)) or {}).get("total_sales", 0.0) or 0.0
        mg = monthly_goal_by_loc.get(name) or {}
        goal = mg.get("monthly_goal", 0.0) or 0.0
        d_div = mg.get("daily_goal_divisor", 0) or 0
        g_div = mg.get("day_goal_divisor", 0) or 0
        daily = goal / d_div if d_div else 0.0
        day   = goal / g_div if g_div else 0.0
        goal_pct = cur_mtd / day if day else 0.0

        _set(ws, row, 2, cur_mtd, fmt=FMT_MONEY)
        _set(ws, row, 3, py,      fmt=FMT_MONEY)
        _set(ws, row, 4, _safe_div(cur_mtd, py), fmt=FMT_PCT)
        _set(ws, row, 5, goal,    fmt=FMT_MONEY)
        _set(ws, row, 6, daily,   fmt=FMT_MONEY)
        _set(ws, row, 7, day,     fmt=FMT_MONEY)
        _set(ws, row, 8, cur_mtd, fmt=FMT_MONEY)
        _set(ws, row, 9, goal_pct, fmt=FMT_PCT)
        _set(ws, row, 10, pp,     fmt=FMT_MONEY)
        _set(ws, row, 11, _safe_div(cur_mtd, pp), fmt=FMT_PCT)

        tot["y2026"]        += cur_mtd
        tot["y2025"]        += py
        tot["y2019"]        += pp
        tot["monthly_goal"] += goal
        tot["daily_goal"]   += daily
        tot["day_goal"]     += day
        tot["we_are_at"]    += cur_mtd

    _set(ws, 14, 1,  "Totals", bold=True)
    _set(ws, 14, 2,  tot["y2026"], fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 3,  tot["y2025"], fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 4,  _safe_div(tot["y2026"], tot["y2025"]), fmt=FMT_PCT, bold=True)
    _set(ws, 14, 5,  tot["monthly_goal"], fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 6,  tot["daily_goal"],   fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 7,  tot["day_goal"],     fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 8,  tot["we_are_at"],    fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 9,  _safe_div(tot["we_are_at"], tot["day_goal"]), fmt=FMT_PCT, bold=True)
    _set(ws, 14, 10, tot["y2019"], fmt=FMT_MONEY, bold=True)
    _set(ws, 14, 11, _safe_div(tot["y2026"], tot["y2019"]), fmt=FMT_PCT, bold=True)

    # ── Row 17: Table 2 headers (Q2 YoY) ──
    _set(ws, 17, 1, "April-june ", bold=True)
    _set(ws, 17, 2, "2nd Quarter 2026", bold=True)
    _set(ws, 17, 3, "2nd 2025", bold=True)
    _set(ws, 17, 4, "Sales %", bold=True)
    _set(ws, 17, 5, "2ndQuarter 2019", bold=True)
    _set(ws, 17, 6, "Sales %", bold=True)

    # ── Row 18 (right side): Table 3 headers (Q2 monthly breakdown) ──
    _set(ws, 18, 11, "2026 Q2", bold=True)
    _set(ws, 18, 12, "April",    bold=True)
    _set(ws, 18, 13, "May ",     bold=True)
    _set(ws, 18, 14, "June ",    bold=True)
    _set(ws, 18, 15, "Q2 Total ", bold=True)

    # ── Rows 18..29: Table 2 — Q2 totals ──
    tot_q2 = {"y2026": 0.0, "y2025": 0.0, "y2019": 0.0}
    apr_total = 0.0
    may_total = 0.0
    jun_total = 0.0
    for i, loc in enumerate(locations):
        row = 18 + i  # rows 18..29 for 12 locations
        name = loc["name"]
        _set(ws, row, 1, name)
        apr = (historical_by_key.get((f"{year:04d}-04", name)) or {}).get("total_sales", 0.0) or 0.0
        may = (historical_by_key.get((f"{year:04d}-05", name)) or {}).get("total_sales", 0.0) or 0.0
        jun = (historical_by_key.get((f"{year:04d}-06", name)) or {}).get("total_sales", 0.0) or 0.0
        q2_2026 = apr + may + jun
        q2_2025 = (historical_by_key.get((f"{year - 1}-Q2", name)) or {}).get("total_sales", 0.0) or 0.0
        q2_2019 = (historical_by_key.get(("2019-Q2", name)) or {}).get("total_sales", 0.0) or 0.0

        _set(ws, row, 2, q2_2026, fmt=FMT_MONEY)
        _set(ws, row, 3, q2_2025, fmt=FMT_MONEY)
        _set(ws, row, 4, _safe_div(q2_2026, q2_2025), fmt=FMT_PCT)
        _set(ws, row, 5, q2_2019, fmt=FMT_MONEY)
        _set(ws, row, 6, _safe_div(q2_2026, q2_2019), fmt=FMT_PCT)

        tot_q2["y2026"] += q2_2026
        tot_q2["y2025"] += q2_2025
        tot_q2["y2019"] += q2_2019

        # Table 3 (right side, rows 19..30): per-location Q2 monthly breakdown
        t3_row = 19 + i
        _set(ws, t3_row, 11, name)
        if apr: _set(ws, t3_row, 12, apr, fmt=FMT_MONEY)
        if may: _set(ws, t3_row, 13, may, fmt=FMT_MONEY)
        if jun: _set(ws, t3_row, 14, jun, fmt=FMT_MONEY)
        if q2_2026: _set(ws, t3_row, 15, q2_2026, fmt=FMT_MONEY)

        apr_total += apr
        may_total += may
        jun_total += jun

    # ── Row 30: Table 2 + Table 3 totals ──
    _set(ws, 30, 1, "Totals", bold=True)
    _set(ws, 30, 2, tot_q2["y2026"], fmt=FMT_MONEY, bold=True)
    _set(ws, 30, 3, tot_q2["y2025"], fmt=FMT_MONEY, bold=True)
    _set(ws, 30, 4, _safe_div(tot_q2["y2026"], tot_q2["y2025"]), fmt=FMT_PCT, bold=True)
    _set(ws, 30, 5, tot_q2["y2019"], fmt=FMT_MONEY, bold=True)
    _set(ws, 30, 6, _safe_div(tot_q2["y2026"], tot_q2["y2019"]), fmt=FMT_PCT, bold=True)

    _set(ws, 31, 11, "Totals", bold=True)
    _set(ws, 31, 12, apr_total, fmt=FMT_MONEY, bold=True)
    _set(ws, 31, 13, may_total, fmt=FMT_MONEY, bold=True)
    _set(ws, 31, 14, jun_total, fmt=FMT_MONEY, bold=True)
    _set(ws, 31, 15, apr_total + may_total + jun_total, fmt=FMT_MONEY, bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 16
    for c in range(2, 16):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Karissa's YoY tab uses cyan headers across the three sub-tables.
    _init_fills()
    # Table 1 header row (row 1)
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = FILL_CYAN
    # Table 2 header row (row 17)
    for c in range(1, 7):
        ws.cell(row=17, column=c).fill = FILL_CYAN
    # Table 3 header row (row 18, cols K-O)
    for c in range(11, 16):
        ws.cell(row=18, column=c).fill = FILL_CYAN


# ─── Public entry point ───────────────────────────────────────────────────────

def build_monthly_workbook(
    year_month: str,
    output_path: Path,
    service,
    config: dict,
    *,
    dry_run: bool = False,
) -> "Path | None":
    """Generate a monthly workbook in Karissa's layout for ``year_month``.

    Returns the path written, or None if openpyxl is unavailable.

    Reads from Google Sheets:
      CUMULATIVE_MTD    — weekly snapshots for the month (1-5 weeks)
      DATA_MONTHLY      — prior-year monthly totals (2025-MM, 2019-MM, 2025-Q2,
                          2019-Q2) + this year's prior months (e.g. 2026-04 for
                          the bottom-block prior-month reference columns)
      MONTHLY_GOALS     — per-location monthly $ target + divisors
    """
    if not OPENPYXL_AVAILABLE:
        log.warning("openpyxl not available — skipping karissa_workbook build")
        return None

    year, month = _parse_year_month(year_month)
    sundays = _sundays_of_month(year, month)
    n_weeks = len(sundays)  # May 2026 = 5 Sundays; June 2026 = 4 Sundays

    locations = config["locations"]

    if dry_run:
        log.info("DRY RUN: would build Karissa workbook for %s -> %s", year_month, output_path)
        wb = openpyxl.Workbook()
        wb.active.title = "DRY_RUN"
        wb.active["A1"] = f"[DRY RUN] {year_month}"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    # Source data ----------------------------------------------------------
    cum = data_source.read_cumulative_mtd_snapshots(service, config, year_month)
    historical_year_months = [
        f"{year - 1}-{month:02d}",  # prior-year same month (e.g., 2025-05)
        f"2019-{month:02d}",         # 2019 same month
        f"{year - 1}-Q2",            # 2025-Q2 for the Q2 YoY sub-table
        "2019-Q2",
        f"{year:04d}-04",            # April (prior month + Q2 breakdown)
        f"{year:04d}-05",            # May (Q2 breakdown)
        f"{year:04d}-06",            # June (Q2 breakdown — may be empty)
    ]
    monthly_rows = data_source.read_data_monthly_rows(service, config, year_months=historical_year_months)
    historical_by_key = {(r["year_month"], r["loc_name"]): r for r in monthly_rows}

    goals = data_source.read_monthly_goals(service, config, year_month)

    # Group cumulative snapshots by week_ending --------------------------
    snaps_by_week: dict[str, dict[str, dict]] = {}
    for s in cum:
        we = s.get("week_ending", "")
        snaps_by_week.setdefault(we, {})[s["loc_name"]] = s

    # Prior-month reference (DATA_MONTHLY 2025-MM and 2026-(MM-1)) keyed by (ym, loc) -
    prior_month_by_loc = historical_by_key  # already (year_month, loc_name) keyed

    # Build workbook --------------------------------------------------------
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    working_days_total = _working_days_in_month(year, month)

    # Always render 5 weekly tabs in order to mirror Karissa's template, even
    # when the month only has 4 Sundays (the 5th tab stays an empty template).
    for idx in range(1, 6):
        week_ending = sundays[idx - 1] if idx - 1 < len(sundays) else None
        we_str = week_ending.isoformat() if week_ending else ""
        snapshots = snaps_by_week.get(we_str, {})
        _build_weekly_tab(
            wb, idx, week_ending, locations,
            snapshots_for_week=snapshots,
            prior_month_by_loc=prior_month_by_loc,
            monthly_goal_by_loc=goals,
            year_month=year_month,
            working_days_in_month=working_days_total,
        )

    # Latest-snapshot-by-location for the YoY tab's "current MTD" column.
    latest_snapshot_by_loc: dict[str, dict] = {}
    for we_str in sorted(snaps_by_week.keys()):
        for loc, s in snaps_by_week[we_str].items():
            latest_snapshot_by_loc[loc] = s  # last wins (sorted ASC -> latest is final)

    _build_yoy_tab(
        wb,
        year_month=year_month,
        locations=locations,
        latest_snapshot_by_loc=latest_snapshot_by_loc,
        historical_by_key=historical_by_key,
        monthly_goal_by_loc=goals,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(
        "Karissa workbook saved: %s (%.1f KB, %d weekly tabs, %d locations)",
        output_path, output_path.stat().st_size / 1024, n_weeks, len(locations),
    )
    return output_path
