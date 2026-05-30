"""
core/report_generator.py
KPI Platform — THE deliverable: Karissa's weekly Excel, generated from the Sheet.

This is the single, canonical report builder. It reads the live Google Sheet
(the source of truth) via ``core/data_source.py`` — the cumulative-MTD weekly
snapshots in CUMULATIVE_MTD, the per-location targets in MONTHLY_GOALS, and the
prior-year reference in DATA_MONTHLY — and writes Karissa's weekly workbook in her
exact layout. It supersedes the earlier ``core/karissa_workbook.py`` (consolidated
here 2026-05-29) and the parallel JSONL ``history_store`` experiment: there is ONE
report path, fed by the Sheet.

WHY SHEET-FED (not a private store)
-----------------------------------
The live Sheet is where the pipeline already lands its data and what the dashboards
already read. A second store (local JSONL) would be a divergent source of truth. So
the report reads the same tabs the rest of the platform writes. The builder is still
storage-agnostic at its core — ``build_workbook_from_data`` takes plain dicts, so
tests (and any future backend) drive it with zero network.

FAITHFUL REPRODUCTION, NOT A REDESIGN
-------------------------------------
Karissa's staff treat her report as canonical; we output a file that looks like the
one she builds by hand so they read it with no retraining: a workbook with one tab
per week (' Week 1 ' .. ' week 5') + a 'Year Over Year' tab; each week tab has two
stacked blocks (top = per-location KPIs, bottom = service-mix penetration); a Totals
row per block; the YoY tab is one row per location.

TWO DELIBERATE DEPARTURES from her hand-built file (both are improvements her own
schema doc asks for — KPI_LOCATIONS_DATA_Schema BUILD NOTE #2):

  1. CANONICAL CONSISTENT HEADERS every week. Her real files drift — May inserts a
     'tickets' column at B (a Region# tag that feeds nothing), her week-4 tab drops
     the Name column, her col-L header rotates 'Projection'/'Projections'/'Monthly
     goal'/'Salon Goals', and the 2025-reference block repeats 'Guest Count'/'PPG'/
     'AT'. We emit ONE fixed header set (the newest = May superset) on every tab so
     positions never drift, and we map by HEADER NAME, never by column index.

  2. READ VALUES AS STORED. The ratios (product_pct, ppg, pph, avg_ticket, wax_pct,
     color_pct, treat_pct) are written verbatim from the snapshot — they were
     already computed canonically + penny-verified by the parsers. We do NOT
     recompute them here (recomputing risks drifting from the stored truth). Only
     Totals-row aggregates are computed (sums / ratio-of-sums / column-sums), exactly
     as Karissa's own Totals row does.

DRIFT-PROOFING (proven necessary)
---------------------------------
``map_top_header_row`` maps an arbitrary top-block header row -> canonical fields,
position-independently and SECTION-AWARE (it splits at '2025 Total Sales' so the
repeated current vs reference labels resolve correctly). It guards the header
contract and lets any future "parse Karissa's own .xlsx" path read March- and
May-style sheets identically.

CUMULATIVE-MTD TIME MODEL (Karissa-confirmed)
---------------------------------------------
Each weekly snapshot is ALREADY cumulative month-to-date and resets each month (her
Week-3 figure already contains Wk1+2+3 — March Forest Lake Total Sales Wk1 8,147 ->
Wk3 22,358 -> Wk5 33,191). We READ the snapshot for each week and PLACE it in that
week's tab — we NEVER sum weekly rows (that would double-count). A closed/holiday
week is a zero row that is still PRESENT (never a skipped tab/row).

YoY GRACEFUL-BLANK
------------------
The 2025 / 2019 / Sales% columns come from prior-year DATA_MONTHLY rows that don't
exist until the separate backload lands. We render them BLANK — never crash, never
zero-fill. Goals (from MONTHLY_GOALS) DO populate today.

NOT WIRED beyond the report step. ``build_monthly_workbook`` matches the old
karissa_workbook signature so main.py / sandbox swap in with a one-line import
change; the broader storage reconciliation is a separate track.
"""
from __future__ import annotations

import datetime as _dt
import json
import logging
from calendar import monthrange
from dataclasses import dataclass
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    OPENPYXL_AVAILABLE = False

from core import data_source

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Karissa's tab labels — intentionally non-uniform whitespace; match byte-for-byte.
# ─────────────────────────────────────────────────────────────────────────────
WEEK_TAB_NAMES = [" Week 1 ", " week 2", " week 3", " week 4", " week 5"]
YOY_TAB_NAME = "Year Over Year"
N_WEEK_TABS = 5  # her template always has 5 week tabs, even if some are empty

PERCENT_OF_GUESTS_LABEL = "% of Guests "
GOAL_ANNOTATION_LABEL = "Goal is 15% plus"
PROVISIONAL_NOTE = "* provisional — data flagged incomplete"

FMT_INT = "#,##0"
FMT_MONEY = "#,##0.00"
FMT_PCT = "0.00%"
FMT_DECIMAL = "0.00"
FMT_DATE = "yyyy-mm-dd"

_SUNDAY_WEEKDAY = 6  # date.weekday(): Mon=0 .. Sun=6. Karissa works Mon-Sat.

# Karissa hand-types =(cumul / days_elapsed) * days_total per weekly tab. Her counts
# don't match a pure Mon-Sat calc (e.g. Memorial Day), so we hardcode her observed
# divisors per filled month and fall back to Mon-Sat otherwise.
# {"YYYY-MM": {week_idx: (days_elapsed, days_total)}}
PROJECTION_DIVISORS_BY_MONTH: dict[str, dict[int, tuple[int, int]]] = {
    "2026-05": {1: (2, 25), 2: (8, 25), 3: (14, 25), 4: (21, 25), 5: (25, 25)},
}


# ─────────────────────────────────────────────────────────────────────────────
# Canonical column spec — SINGLE source of truth shared by the writer (emits
# ``header`` in order) and the reader (``aliases`` map a drifted header back to
# ``field``). ``field`` is the snapshot dict key (data_source CUMULATIVE_MTD shape)
# for current-week metrics, a ``py_*`` prior-year key, a ``diff_*`` computed delta,
# or a ``_``-prefixed shape-only / blank column. ``section`` disambiguates the
# repeated current vs 2025-reference labels.
# ─────────────────────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class ColSpec:
    field: str
    header: object
    section: str
    fmt: str | None = None
    aliases: tuple[str, ...] = ()

    @property
    def is_blank(self) -> bool:
        return isinstance(self.field, str) and self.field.startswith("_")


# Top block — per-location KPIs. Header row 2; data 3..14; Totals 15; net-avg 16.
# Canonical = May Week-1 superset (incl. the noise 'tickets' col at B), emitted
# identically on EVERY week tab so positions never drift.
TOP_BLOCK: list[ColSpec] = [
    ColSpec("loc_name",       "Name",             "current", None,        ("name",)),
    ColSpec("_tickets",       "tickets ",         "current", None,        ("tickets",)),
    ColSpec("guests",         "Guest Count",      "current", FMT_INT,     ("guest count",)),
    ColSpec("total_sales",    "Total Sales Net",  "current", FMT_MONEY,   ("total sales net", "total sales $")),
    ColSpec("service",        "Service Net",      "current", FMT_MONEY,   ("service net", "service $")),
    ColSpec("product",        "Product Net",      "current", FMT_MONEY,   ("product net", "product $")),
    ColSpec("product_pct",    "Product %",        "current", FMT_PCT,     ("product %",)),
    ColSpec("ppg",            "PPG Net",          "current", FMT_DECIMAL, ("ppg net", "ppg $")),
    ColSpec("pph",            "PPH Net",          "current", FMT_DECIMAL, ("pph net", "pph $")),
    ColSpec("avg_ticket",     "Average Tkt",      "current", FMT_DECIMAL, ("average tkt", "avg ticket $", "avg ticket")),
    ColSpec("prod_hours",     "Prod Hours",       "current", FMT_DECIMAL, ("prod hours",)),
    ColSpec("projection",     "Projection",       "current", FMT_MONEY,
            ("projection", "projections", "monthly goal", "salon goals", "salon goal")),
    # ── 2025 reference block (prior-year; blank until backload). Section flips here. ──
    ColSpec("py_total_sales", "2025 Total Sales", "ref2025", FMT_MONEY,   ("2025 total sales",)),
    ColSpec("py_guests",      " Guest Count",     "ref2025", FMT_INT,     ("guest count",)),
    ColSpec("py_ppg",         "PPG",              "ref2025", FMT_DECIMAL, ("ppg", "2025ppg")),
    ColSpec("diff_ppg",       "Diff",             "ref2025", FMT_DECIMAL, ("diff",)),
    ColSpec("py_avg_ticket",  " AT",              "ref2025", FMT_DECIMAL, ("at",)),
    ColSpec("diff_at",        "Dif",              "ref2025", FMT_DECIMAL, ("dif",)),
]

# Bottom block — service-mix penetration. Header row 18; data 19..30; Totals 31.
# Left primaries from the snapshot; the right reference band (prior-month / YoY
# helpers) renders blank-with-headers (graceful — awaits backload).
BOTTOM_BLOCK: list[ColSpec] = [
    ColSpec("loc_name",     "Name",       "primary", None,      ("name",)),
    ColSpec("_blank_b",     "",           "primary", None,      ()),
    ColSpec("wax_count",    "Wax Count",  "primary", FMT_INT,   ("wax count",)),
    ColSpec("wax",          "Waxing Net", "primary", FMT_MONEY, ("waxing net", "wax net", "wax $")),
    ColSpec("wax_pct",      "Wax %",      "primary", FMT_PCT,   ("wax %",)),
    ColSpec("color",        "Color Net",  "primary", FMT_MONEY, ("color net", "color $")),
    ColSpec("color_pct",    "Color %",    "primary", FMT_PCT,   ("color %",)),
    ColSpec("treat_count",  "Trmt Count", "primary", FMT_INT,   ("trmt count", "treatment count")),
    ColSpec("treat",        "Trmt Net",   "primary", FMT_MONEY, ("trmt net", "treatment net", "treatment $")),
    ColSpec("treat_pct",    "Trmt %",     "primary", FMT_PCT,   ("trmt %", "treatment %")),
    ColSpec("_sep",         "",           "ref",     None,      ()),
    ColSpec("_ref_prod_hours", " Prod Hours", "ref", None,      ()),
    ColSpec("_ref_wax",        "Wax ",        "ref", None,      ()),
    ColSpec("_ref_treat",      "Treat ",      "ref", None,      ()),
    ColSpec("_ref_service",    "Service ",    "ref", None,      ()),
    ColSpec("_ref_treat_pct",  "Treat %",     "ref", None,      ()),
    ColSpec("_ref_2026_treat", "2026 treat",  "ref", None,      ()),
    ColSpec("_ref_wax_pct",    "Wax%",        "ref", None,      ()),
    ColSpec("_ref_2026_wax",   "2026 wax",    "ref", None,      ()),
]

# YoY tab — one row per location (spec §3). 2025/2019/Sales% are prior-year (blank
# until backload); Goal/Daily Goal/Day Goal come from MONTHLY_GOALS.
YOY_HEADERS = ["Name", 2026, 2025, "Sales %", " Goal", "Daily Goal",
               "Day Goal", "We are At!", " Goal %", 2019, "Sales %"]


# ─────────────────────────────────────────────────────────────────────────────
# Header-name mapping — drift-proofing. Trim + case-fold, then resolve by SECTION.
# ─────────────────────────────────────────────────────────────────────────────
def normalize_header(value) -> str:
    """Canonicalize a header cell for matching: trim + collapse whitespace + lower.
    Her labels carry inconsistent leading/trailing space ('tickets ', ' Guest Count',
    ' AT'); the metric is the same regardless."""
    if value is None:
        return ""
    return " ".join(str(value).strip().split()).lower()


def _alias_lookup(specs, section) -> dict[str, str]:
    """Build {normalized header/alias -> field} for one section. Shape-only / blank
    columns (``_``-prefixed, e.g. the noise 'tickets' tag) are NOT mapped — the
    reader resolves real metrics and skips noise, so March (no tickets) and May
    (tickets at B) yield identical field sets."""
    out: dict[str, str] = {}
    for s in specs:
        if s.section != section or s.is_blank:
            continue
        for lab in {normalize_header(s.header), *s.aliases}:
            if lab:
                out.setdefault(lab, s.field)
    return out


_TOP_CURRENT_LOOKUP = _alias_lookup(TOP_BLOCK, "current")
_TOP_REF_LOOKUP = _alias_lookup(TOP_BLOCK, "ref2025")
_REF_FLIP_KEY = normalize_header("2025 Total Sales")


def map_top_header_row(header_cells) -> dict[int, str]:
    """Map an arbitrary top-block header row -> {1-based column index: field}.

    Position-INDEPENDENT and SECTION-AWARE: walk left to right; once we cross the
    '2025 Total Sales' header, the repeated ambiguous labels ('Guest Count', 'PPG',
    'AT') resolve to the prior-year ``py_*`` fields instead of the current-week ones.
    Unknown/noise headers ('tickets', 'Region #', blanks) are skipped. This is what
    makes March-style and May-style sheets read identically — the key drift defense.
    """
    mapping: dict[int, str] = {}
    section = "current"
    for idx, raw in enumerate(header_cells, start=1):
        key = normalize_header(raw)
        if not key:
            continue
        if key == _REF_FLIP_KEY:
            mapping[idx] = "py_total_sales"
            section = "ref2025"
            continue
        lookup = _TOP_CURRENT_LOOKUP if section == "current" else _TOP_REF_LOOKUP
        field = lookup.get(key)
        if field:
            mapping[idx] = field
    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# Calendar / projection helpers.
# ─────────────────────────────────────────────────────────────────────────────
def _parse_year_month(year_month: str) -> tuple[int, int]:
    y, m = str(year_month).split("-")[:2]
    return int(y), int(m)


def _sundays_of_month(year: int, month: int) -> list[_dt.date]:
    _, n = monthrange(year, month)
    return [_dt.date(year, month, d) for d in range(1, n + 1)
            if _dt.date(year, month, d).weekday() == _SUNDAY_WEEKDAY]


def _working_days_in_range(start: _dt.date, end_inclusive: _dt.date) -> int:
    if end_inclusive < start:
        return 0
    n, d = 0, start
    while d <= end_inclusive:
        if d.weekday() != _SUNDAY_WEEKDAY:
            n += 1
        d += _dt.timedelta(days=1)
    return n


def _working_days_in_month(year: int, month: int) -> int:
    _, n = monthrange(year, month)
    return _working_days_in_range(_dt.date(year, month, 1), _dt.date(year, month, n))


def _projection(total_sales, year_month, week_idx, week_ending):
    """Karissa's cumulative projection: (cumul / working_days_elapsed) * working_days
    _in_month. Prefer her hand-curated divisors; fall back to Mon-Sat. None-safe."""
    if total_sales is None:
        return None
    override = PROJECTION_DIVISORS_BY_MONTH.get(year_month, {}).get(week_idx)
    if override:
        days_elapsed, days_total = override
    elif week_ending is not None:
        month_first = _dt.date(week_ending.year, week_ending.month, 1)
        days_elapsed = max(_working_days_in_range(month_first, week_ending), 1)
        days_total = _working_days_in_month(week_ending.year, week_ending.month)
    else:
        return None
    return (total_sales / days_elapsed) * days_total if days_elapsed else None


def _as_date(s):
    """ISO 'YYYY-MM-DD' -> date; pass through date/datetime; None/garbage -> None."""
    if s is None or isinstance(s, _dt.date):
        return s if not isinstance(s, _dt.datetime) else s.date()
    try:
        return _dt.datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Grouping — cumulative snapshots into week tabs (data-driven, never summed).
# ─────────────────────────────────────────────────────────────────────────────
def group_snapshots_into_weeks(snapshots) -> tuple[list[str], dict[int, dict[str, dict]]]:
    """Assign each cumulative-MTD snapshot to a week tab. The week is driven BY THE
    DATA: distinct week_ending dates sorted ascending -> Week 1, 2, ... in order.

    NO SUMMING — each week shows its own already-cumulative snapshot verbatim.

    Returns (week_endings_iso_sorted, {week_idx (1-based): {loc_name: snapshot}}).
    """
    week_endings = sorted({str(s.get("week_ending")) for s in snapshots if s.get("week_ending")})
    week_index = {we: i + 1 for i, we in enumerate(week_endings)}
    by_week: dict[int, dict[str, dict]] = {}
    for s in snapshots:
        we = str(s.get("week_ending"))
        if we not in week_index:
            continue
        by_week.setdefault(week_index[we], {})[s.get("loc_name")] = s
    return week_endings, by_week


# ─────────────────────────────────────────────────────────────────────────────
# Roster — all 12 locations in Karissa's report order, joined to snapshots by name.
# ─────────────────────────────────────────────────────────────────────────────
_DEFAULT_CUSTOMER = Path(__file__).resolve().parent.parent / "config" / "customers" / "karissa_001.json"


def default_roster() -> list[str]:
    try:
        cfg = json.loads(_DEFAULT_CUSTOMER.read_text(encoding="utf-8"))
        return [loc["name"] for loc in cfg.get("locations", [])]
    except Exception as e:  # pragma: no cover
        log.warning("default_roster: could not read %s (%s)", _DEFAULT_CUSTOMER, e)
        return []


# ─────────────────────────────────────────────────────────────────────────────
# Styling helpers (palette from her May 2026.xlsx ARGB values).
# ─────────────────────────────────────────────────────────────────────────────
_FILLS: dict[str, object] = {}


def _fills() -> dict[str, object]:
    if not _FILLS and OPENPYXL_AVAILABLE:
        _FILLS["cyan"] = PatternFill("solid", fgColor="00B0F0")
        _FILLS["green"] = PatternFill("solid", fgColor="92D050")
        _FILLS["peach"] = PatternFill("solid", fgColor="FBE4D5")
    return _FILLS


def _bold():
    return Font(bold=True)


def _set(ws, row, col, value, *, fmt=None, bold=False, fill=None):
    """Write a cell. Skips None values (blank — graceful) but still styles, so empty
    reference columns keep their header banner."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
        if fmt:
            cell.number_format = fmt
    if bold:
        cell.font = _bold()
    if fill is not None:
        cell.fill = fill
    return cell


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_div(num, den):
    return num / den if den else None


def _autosize(ws, n_cols):
    ws.column_dimensions["A"].width = 16
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13


# ─────────────────────────────────────────────────────────────────────────────
# Week-tab builder.
# ─────────────────────────────────────────────────────────────────────────────
def _build_week_tab(wb, week_idx, week_ending, roster, snaps_for_week,
                    year_month, prior_year_by_loc):
    """One week tab: top KPI block + bottom service-mix block.

    ``snaps_for_week`` is {loc_name: snapshot}. A location absent from it is listed
    with blank metrics (week not yet reported); a location present with zero values
    shows those zeros (closed-week zero row, never skipped); a snapshot carrying
    data_complete_flag=False gets a '*' provisional marker on its name.
    """
    ws = wb.create_sheet(title=WEEK_TAB_NAMES[week_idx - 1])
    top_col = {s.field: i + 1 for i, s in enumerate(TOP_BLOCK)}
    bot_col = {s.field: i + 1 for i, s in enumerate(BOTTOM_BLOCK)}
    n_top, n_bot = len(TOP_BLOCK), len(BOTTOM_BLOCK)
    fills = _fills()
    we_date = _as_date(week_ending)

    # ── Row 1: banners. ──
    if we_date:
        _set(ws, 1, 1, _dt.datetime(we_date.year, we_date.month, we_date.day), fmt=FMT_DATE)
    _set(ws, 1, top_col["total_sales"], "Region #", bold=True)
    _set(ws, 1, top_col["projection"], "End of month", bold=True)
    if fills:
        for c in range(1, n_top + 1):
            ws.cell(row=1, column=c).fill = fills["cyan"]

    # ── Row 2: top headers (canonical, identical every week). ──
    for i, s in enumerate(TOP_BLOCK):
        fill = None
        if fills:
            if s.field == "loc_name" or s.section == "current":
                fill = fills["cyan"]
            elif s.field in ("diff_ppg", "diff_at"):
                fill = fills["peach"]
            else:
                fill = fills["green"]
        _set(ws, 2, i + 1, s.header, bold=True, fill=fill)

    # ── Rows 3..14: per-location KPIs (verbatim from snapshot). ──
    agg = {k: 0.0 for k in ("guests", "total_sales", "service", "product",
                            "prod_hours", "projection", "ppg", "pph", "avg_ticket")}
    n_data = 0
    provisional = False
    year, month = _parse_year_month(year_month)
    prior_ym = f"{year - 1}-{month:02d}"

    for i, name in enumerate(roster):
        row = 3 + i
        snap = snaps_for_week.get(name)
        display = name
        if snap and snap.get("data_complete_flag") is False:
            display, provisional = f"{name} *", True
        _set(ws, row, top_col["loc_name"], display)
        if not snap:
            continue
        n_data += 1
        # Direct, read-as-stored metrics.
        _set(ws, row, top_col["guests"],      snap.get("guests"),      fmt=FMT_INT)
        _set(ws, row, top_col["total_sales"], snap.get("total_sales"), fmt=FMT_MONEY)
        _set(ws, row, top_col["service"],     snap.get("service"),     fmt=FMT_MONEY)
        _set(ws, row, top_col["product"],     snap.get("product"),     fmt=FMT_MONEY)
        _set(ws, row, top_col["product_pct"], snap.get("product_pct"), fmt=FMT_PCT)
        _set(ws, row, top_col["ppg"],         snap.get("ppg"),         fmt=FMT_DECIMAL)
        _set(ws, row, top_col["pph"],         snap.get("pph"),         fmt=FMT_DECIMAL)
        _set(ws, row, top_col["avg_ticket"],  snap.get("avg_ticket"),  fmt=FMT_DECIMAL)
        _set(ws, row, top_col["prod_hours"],  snap.get("prod_hours"),  fmt=FMT_DECIMAL)
        # Projection (computed — cumulative run-rate).
        proj = _projection(_num(snap.get("total_sales")), year_month, week_idx, we_date)
        _set(ws, row, top_col["projection"], proj, fmt=FMT_MONEY)

        for k in agg:
            if k == "projection":
                if proj is not None:
                    agg[k] += proj
            else:
                nv = _num(snap.get(k))
                if nv is not None:
                    agg[k] += nv

        # 2025 reference + diffs (graceful blank when prior-year row absent).
        py = prior_year_by_loc.get((prior_ym, name))
        if py:
            _set(ws, row, top_col["py_total_sales"], py.get("total_sales"), fmt=FMT_MONEY)
            _set(ws, row, top_col["py_guests"],      py.get("guests"),      fmt=FMT_INT)
            _set(ws, row, top_col["py_ppg"],         py.get("ppg"),         fmt=FMT_DECIMAL)
            _set(ws, row, top_col["py_avg_ticket"],  py.get("avg_ticket"),  fmt=FMT_DECIMAL)
            cur_ppg, py_ppg = _num(snap.get("ppg")), _num(py.get("ppg"))
            cur_at, py_at = _num(snap.get("avg_ticket")), _num(py.get("avg_ticket"))
            if cur_ppg is not None and py_ppg is not None:
                _set(ws, row, top_col["diff_ppg"], round(cur_ppg - py_ppg, 2), fmt=FMT_DECIMAL)
            if cur_at is not None and py_at is not None:
                _set(ws, row, top_col["diff_at"], round(cur_at - py_at, 2), fmt=FMT_DECIMAL)

    # ── Row 15: Totals. ──
    tr = 15
    _set(ws, tr, top_col["loc_name"], "Totals", bold=True)
    if n_data:
        _set(ws, tr, top_col["guests"],      round(agg["guests"]),         fmt=FMT_INT,     bold=True)
        _set(ws, tr, top_col["total_sales"], round(agg["total_sales"], 2), fmt=FMT_MONEY,   bold=True)
        _set(ws, tr, top_col["service"],     round(agg["service"], 2),     fmt=FMT_MONEY,   bold=True)
        _set(ws, tr, top_col["product"],     round(agg["product"], 2),     fmt=FMT_MONEY,   bold=True)
        if agg["total_sales"]:
            _set(ws, tr, top_col["product_pct"], agg["product"] / agg["total_sales"], fmt=FMT_PCT, bold=True)
        _set(ws, tr, top_col["ppg"],         round(agg["ppg"], 2),         fmt=FMT_DECIMAL, bold=True)
        _set(ws, tr, top_col["pph"],         round(agg["pph"], 2),         fmt=FMT_DECIMAL, bold=True)
        _set(ws, tr, top_col["avg_ticket"],  round(agg["avg_ticket"], 2),  fmt=FMT_DECIMAL, bold=True)
        _set(ws, tr, top_col["prod_hours"],  round(agg["prod_hours"], 2),  fmt=FMT_DECIMAL, bold=True)
        _set(ws, tr, top_col["projection"],  round(agg["projection"], 2),  fmt=FMT_MONEY,   bold=True)
        # ── Row 16: network averages (correct columns). ──
        _set(ws, 16, top_col["ppg"],        round(agg["ppg"] / n_data, 2),        fmt=FMT_DECIMAL)
        _set(ws, 16, top_col["pph"],        round(agg["pph"] / n_data, 2),        fmt=FMT_DECIMAL)
        _set(ws, 16, top_col["avg_ticket"], round(agg["avg_ticket"] / n_data, 2), fmt=FMT_DECIMAL)

    # ── Row 18: bottom-block date + headers. ──
    if we_date:
        _set(ws, 18, 1, _dt.datetime(we_date.year, we_date.month, we_date.day), fmt=FMT_DATE)
    for i, s in enumerate(BOTTOM_BLOCK):
        fill = None
        if fills and s.header != "":
            fill = fills["green"] if s.section == "ref" else fills["cyan"]
        _set(ws, 18, i + 1, s.header, bold=True, fill=fill)

    # ── Rows 19..30: per-location service mix (verbatim). ──
    bagg = {k: 0.0 for k in ("wax_count", "wax", "color", "treat_count", "treat")}
    for i, name in enumerate(roster):
        row = 19 + i
        snap = snaps_for_week.get(name)
        _set(ws, row, bot_col["loc_name"], name)
        if not snap:
            continue
        _set(ws, row, bot_col["wax_count"],   snap.get("wax_count"),   fmt=FMT_INT)
        _set(ws, row, bot_col["wax"],         snap.get("wax"),         fmt=FMT_MONEY)
        _set(ws, row, bot_col["wax_pct"],     snap.get("wax_pct"),     fmt=FMT_PCT)
        _set(ws, row, bot_col["color"],       snap.get("color"),       fmt=FMT_MONEY)
        _set(ws, row, bot_col["color_pct"],   snap.get("color_pct"),   fmt=FMT_PCT)
        _set(ws, row, bot_col["treat_count"], snap.get("treat_count"), fmt=FMT_INT)
        _set(ws, row, bot_col["treat"],       snap.get("treat"),       fmt=FMT_MONEY)
        _set(ws, row, bot_col["treat_pct"],   snap.get("treat_pct"),   fmt=FMT_PCT)
        for k in bagg:
            nv = _num(snap.get(k))
            if nv is not None:
                bagg[k] += nv

    # ── Row 31: bottom Totals (penetration uses the network guest/service base). ──
    btr = 31
    _set(ws, btr, bot_col["loc_name"], "Totals", bold=True)
    if n_data:
        _set(ws, btr, bot_col["wax_count"], round(bagg["wax_count"]), fmt=FMT_INT, bold=True)
        _set(ws, btr, bot_col["wax"],       round(bagg["wax"], 2),    fmt=FMT_MONEY, bold=True)
        if agg["guests"]:
            _set(ws, btr, bot_col["wax_pct"], bagg["wax_count"] / agg["guests"], fmt=FMT_PCT, bold=True)
        _set(ws, btr, bot_col["color"], round(bagg["color"], 2), fmt=FMT_MONEY, bold=True)
        if agg["service"]:
            _set(ws, btr, bot_col["color_pct"], bagg["color"] / agg["service"], fmt=FMT_PCT, bold=True)
        _set(ws, btr, bot_col["treat_count"], round(bagg["treat_count"]), fmt=FMT_INT, bold=True)
        _set(ws, btr, bot_col["treat"],       round(bagg["treat"], 2),    fmt=FMT_MONEY, bold=True)
        if agg["guests"]:
            _set(ws, btr, bot_col["treat_pct"], bagg["treat_count"] / agg["guests"], fmt=FMT_PCT, bold=True)
        # ── Footer: her verbatim "% of Guests" treatment-penetration + annotation. ──
        if agg["guests"]:
            _set(ws, 32, bot_col["color_pct"], PERCENT_OF_GUESTS_LABEL)
            _set(ws, 32, bot_col["treat_count"], bagg["treat_count"] / agg["guests"], fmt=FMT_PCT)
        _set(ws, 33, bot_col["color_pct"], GOAL_ANNOTATION_LABEL, bold=True)

    if provisional:
        _set(ws, 35, 1, PROVISIONAL_NOTE)
    _autosize(ws, max(n_top, n_bot))


# ─────────────────────────────────────────────────────────────────────────────
# Year Over Year tab.
# ─────────────────────────────────────────────────────────────────────────────
def _build_yoy_tab(wb, year_month, roster, latest_by_loc, goals, prior_year_by_loc):
    """One row per location: current-month MTD (2026) + Goal/Daily/Day from
    MONTHLY_GOALS; 2025 / 2019 / Sales% render blank until the prior-year backload."""
    ws = wb.create_sheet(title=YOY_TAB_NAME)
    fills = _fills()
    year, month = _parse_year_month(year_month)
    prior_ym = f"{year - 1}-{month:02d}"

    first_sun = next(iter(_sundays_of_month(year, month)), None)
    if first_sun:
        _set(ws, 1, 1, _dt.datetime(first_sun.year, first_sun.month, first_sun.day), fmt=FMT_DATE)
    for i, h in enumerate(YOY_HEADERS):
        if i == 0:
            continue  # A1 holds the date
        _set(ws, 1, i + 1, h, bold=True, fill=(fills["cyan"] if fills else None))

    tot = {"y2026": 0.0, "goal": 0.0, "daily": 0.0, "day": 0.0}
    for i, name in enumerate(roster):
        row = 2 + i
        _set(ws, row, 1, name)
        snap = latest_by_loc.get(name) or {}
        mtd = _num(snap.get("total_sales"))
        mg = goals.get(name) or {}
        goal = _num(mg.get("monthly_goal"))
        d_div = _num(mg.get("daily_goal_divisor"))
        g_div = _num(mg.get("day_goal_divisor"))
        daily = _safe_div(goal, d_div) if goal is not None else None
        day = _safe_div(goal, g_div) if goal is not None else None
        py = prior_year_by_loc.get((prior_ym, name)) or {}
        py_ts = _num(py.get("total_sales"))

        if mtd is not None:
            _set(ws, row, 2, round(mtd, 2), fmt=FMT_MONEY)
            _set(ws, row, 8, round(mtd, 2), fmt=FMT_MONEY)  # We are At!
            tot["y2026"] += mtd
        if py_ts is not None:
            _set(ws, row, 3, round(py_ts, 2), fmt=FMT_MONEY)
            if mtd is not None and py_ts:
                _set(ws, row, 4, mtd / py_ts, fmt=FMT_PCT)
        if goal is not None:
            _set(ws, row, 5, round(goal, 2), fmt=FMT_MONEY)
            tot["goal"] += goal
        if daily is not None:
            _set(ws, row, 6, round(daily, 2), fmt=FMT_MONEY)
            tot["daily"] += daily
        if day is not None:
            _set(ws, row, 7, round(day, 2), fmt=FMT_MONEY)
            tot["day"] += day
        if mtd is not None and day:
            _set(ws, row, 9, mtd / day, fmt=FMT_PCT)  # Goal %
        # cols 10/11 (2019 / Sales%): prior-year — blank until backload.

    tr = 2 + len(roster)
    _set(ws, tr, 1, "Totals", bold=True)
    _set(ws, tr, 2, round(tot["y2026"], 2), fmt=FMT_MONEY, bold=True)
    _set(ws, tr, 8, round(tot["y2026"], 2), fmt=FMT_MONEY, bold=True)
    if tot["goal"]:
        _set(ws, tr, 5, round(tot["goal"], 2), fmt=FMT_MONEY, bold=True)
    if tot["daily"]:
        _set(ws, tr, 6, round(tot["daily"], 2), fmt=FMT_MONEY, bold=True)
    if tot["day"]:
        _set(ws, tr, 7, round(tot["day"], 2), fmt=FMT_MONEY, bold=True)
        _set(ws, tr, 9, _safe_div(tot["y2026"], tot["day"]), fmt=FMT_PCT, bold=True)
    _autosize(ws, len(YOY_HEADERS))


# ─────────────────────────────────────────────────────────────────────────────
# Builders — pure (testable, no network) + the production Sheet-fed entry point.
# ─────────────────────────────────────────────────────────────────────────────
def build_workbook_from_data(year_month, output_path, *, snapshots,
                             monthly_rows=(), goals=None, roster=None) -> "Path | None":
    """Build the workbook from already-fetched data (no network) — the testable core.

    snapshots    : CUMULATIVE_MTD dicts (data_source.read_cumulative_mtd_snapshots shape).
    monthly_rows : DATA_MONTHLY dicts (prior-year reference); blank YoY if empty.
    goals        : {loc_name: MONTHLY_GOALS dict} for the YoY Goal columns.
    roster       : ordered canonical location names; defaults to config order.
    """
    if not OPENPYXL_AVAILABLE:  # pragma: no cover
        log.warning("openpyxl unavailable — cannot generate report")
        return None
    if roster is None:
        roster = default_roster()
    goals = goals or {}
    prior_year_by_loc = {(r.get("year_month"), r.get("loc_name")): r for r in monthly_rows}

    week_endings, by_week = group_snapshots_into_weeks(snapshots)
    latest_by_loc: dict[str, dict] = {}
    for wk in sorted(by_week):  # ascending -> last wins = latest cumulative MTD
        for name, snap in by_week[wk].items():
            latest_by_loc[name] = snap

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for idx in range(1, N_WEEK_TABS + 1):
        we = week_endings[idx - 1] if idx - 1 < len(week_endings) else None
        _build_week_tab(wb, idx, we, roster, by_week.get(idx, {}), year_month, prior_year_by_loc)
    _build_yoy_tab(wb, year_month, roster, latest_by_loc, goals, prior_year_by_loc)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info("report_generator: wrote %s (%d week-endings, %d locations)",
             out, len(week_endings), len(roster))
    return out


def build_monthly_workbook(year_month, output_path, service, config, *,
                           dry_run: bool = False) -> "Path | None":
    """Production entry point — signature-compatible with the superseded
    karissa_workbook.build_monthly_workbook so main.py / sandbox swap in directly.

    Reads CUMULATIVE_MTD (weekly snapshots) + DATA_MONTHLY (prior-year reference) +
    MONTHLY_GOALS (targets) from the live Sheet, then builds the workbook.
    """
    if not OPENPYXL_AVAILABLE:  # pragma: no cover
        log.warning("openpyxl not available — skipping report build")
        return None

    output_path = Path(output_path)
    if dry_run:
        log.info("DRY RUN: would build report for %s -> %s", year_month, output_path)
        wb = openpyxl.Workbook()
        wb.active.title = "DRY_RUN"
        wb.active["A1"] = f"[DRY RUN] {year_month}"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    year, month = _parse_year_month(year_month)
    snapshots = data_source.read_cumulative_mtd_snapshots(service, config, year_month)
    monthly_rows = data_source.read_data_monthly_rows(
        service, config, year_months=[f"{year - 1}-{month:02d}", f"2019-{month:02d}"])
    goals = data_source.read_monthly_goals(service, config, year_month)
    roster = [loc["name"] for loc in config.get("locations", [])] or None

    return build_workbook_from_data(
        year_month, output_path,
        snapshots=snapshots, monthly_rows=monthly_rows, goals=goals, roster=roster,
    )


if __name__ == "__main__":  # pragma: no cover
    import argparse
    ap = argparse.ArgumentParser(description="Generate Karissa's weekly Excel from the Sheet.")
    ap.add_argument("year_month", help="e.g. 2026-05")
    ap.add_argument("--out", default=None)
    ap.add_argument("--customer", default="karissa_001")
    args = ap.parse_args()

    cfg = json.loads((Path(__file__).resolve().parent.parent / "config" / "customers"
                      / f"{args.customer}.json").read_text(encoding="utf-8"))
    from core import sheets_writer as _sw
    svc = _sw._build_service(cfg)
    out = args.out or f"KPI_{args.year_month}.xlsx"
    path = build_monthly_workbook(args.year_month, out, svc, cfg)
    print(f"wrote {path}" if path else "openpyxl unavailable — nothing written")
