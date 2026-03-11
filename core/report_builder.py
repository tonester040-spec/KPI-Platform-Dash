#!/usr/bin/env python3
"""
core/report_builder.py
KPI Platform — builds weekly Excel report with color-coded performance data.

Sheets:
  Summary      → Network headline, top/bottom performers, AI coach briefing
  Locations    → Core KPIs per location (PPH, revenue, volume, product)
  Service Mix  → Wax / Color / Treatment breakdown per location
  Goals & YOY  → 2026 goals and year-over-year comparisons (placeholders until API feeds)
  Stylists     → One row per stylist, key metrics, star/watch indicators
"""

import logging
import datetime
from pathlib import Path

log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    log.warning("openpyxl not installed — report_builder will be a no-op")


# ─── Colors ───────────────────────────────────────────────────────────────────

NAVY           = "0F1117"
WHITE          = "FFFFFF"
GREEN          = "D4EDDA"
YELLOW         = "FFF3CD"
RED            = "F8D7DA"
LGRAY          = "F8F9FA"
DGRAY          = "6C757D"
GOLD           = "FFC107"
BLUE           = "4A90D9"
PLACEHOLDER_BG = "EEEEEE"   # light gray bg for API-pending columns
PLACEHOLDER_FG = "999999"   # gray text for placeholder values
SECTION_BG     = "E8EDF4"   # soft blue-gray for footer/section rows


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=WHITE, size=10, italic=False) -> "Font":
    return Font(bold=bold, color=color, size=size, italic=italic)


def _align(horizontal="left", wrap=False) -> "Alignment":
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _border() -> "Border":
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── Shared helpers ───────────────────────────────────────────────────────────

def _pph_color(pph: float, avg_pph: float) -> str:
    if pph >= avg_pph:
        return GREEN
    elif pph >= avg_pph * 0.90:
        return YELLOW
    return RED


def _pct_color(val: float, avg: float) -> str:
    if val >= avg:
        return GREEN
    elif val >= avg * 0.90:
        return YELLOW
    return RED


def _header_row(ws, row: int, values: list, widths: list = None):
    """Write a navy header row and optionally set column widths."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = _align("center")
        cell.border = _border()
    if widths:
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _placeholder_header(ws, row: int, col: int, label: str):
    """Gray placeholder header cell (API feed pending)."""
    cell = ws.cell(row=row, column=col, value=f"⏳ {label}")
    cell.fill = _fill(PLACEHOLDER_BG)
    cell.font = _font(bold=True, color=PLACEHOLDER_FG, size=9)
    cell.alignment = _align("center")
    cell.border = _border()


def _placeholder_val(ws, row: int, col: int):
    """Gray placeholder data cell."""
    cell = ws.cell(row=row, column=col, value="—")
    cell.fill = _fill(PLACEHOLDER_BG)
    cell.font = _font(color=PLACEHOLDER_FG, size=10)
    cell.alignment = _align("center")
    cell.border = _border()


# ─── Summary sheet ────────────────────────────────────────────────────────────

def _build_summary(wb, data: dict, ai_cards: dict):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    net      = data["network"]
    week     = net.get("week_ending", "")
    briefing = ai_cards.get("coach_briefing", "")

    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"KPI Weekly Report — Week Ending {week}"
    title_cell.font = Font(bold=True, size=14, color=NAVY)
    title_cell.alignment = _align("center")

    rows = [
        ("", ""),
        ("NETWORK SUMMARY", ""),
        ("Total Locations",  net.get("total_locations", 0)),
        ("Total Guests",     f"{net.get('total_guests', 0):,}"),
        ("Total Sales",      f"${net.get('total_sales', 0):,.0f}"),
        ("Avg PPH",          f"${net.get('avg_pph', 0):.2f}"),
        ("Avg Product %",    f"{net.get('avg_product_pct', 0):.1f}%"),
        ("Avg Ticket",       f"${net.get('avg_avg_ticket', 0):.2f}"),
        ("", ""),
        ("TOP PERFORMER (PPH)", net.get("top_pph_loc", "")),
        ("NEEDS ATTENTION",     net.get("low_pph_loc", "")),
        ("", ""),
        ("COACH BRIEFING", ""),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=(label.isupper()), color=NAVY)
        ws.cell(row=i, column=2, value=value)

    briefing_row = len(rows) + 2
    ws.merge_cells(f"A{briefing_row}:B{briefing_row + 10}")
    cell = ws.cell(row=briefing_row, column=1)
    cell.value = briefing
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[briefing_row].height = 200

    log.debug("Summary sheet built")


# ─── Locations sheet (core KPIs) ──────────────────────────────────────────────

def _build_locations(wb, data: dict, ai_cards: dict):
    ws = wb.create_sheet("Locations")
    ws.freeze_panes = "A2"

    headers = [
        "Location",      "Rank",  "Flag",
        "PPH $",         "vs Avg PPH",
        "Total Sales $", "Service Net $", "Product Net $",
        "Product %",     "vs Avg Prod",   "PPG",
        "Avg Ticket $",  "Guests",        "Prod Hours",
        "AI Summary",
    ]
    widths = [22, 6, 8, 10, 10, 14, 14, 13, 10, 10, 8, 12, 8, 11, 50]

    _header_row(ws, 1, headers, widths)

    net      = data["network"]
    avg_pph  = net.get("avg_pph", 0)
    avg_prod = net.get("avg_product_pct", 0)

    for row_idx, loc in enumerate(data["locations"], start=2):
        name       = loc["loc_name"]
        pph        = loc.get("pph", 0)
        prod       = loc.get("product_pct", 0)
        pph_delta  = loc.get("pph_delta", 0) or 0
        prod_delta = loc.get("product_pct_delta", 0) or 0
        flag       = loc.get("flag", "solid")
        card       = ai_cards.get("location_cards", {}).get(name, "")

        flag_str = "⭐ STAR" if flag == "star" else ("⚠️ WATCH" if flag == "watch" else "✓ SOLID")
        pph_c    = _pph_color(pph, avg_pph)
        prod_c   = _pct_color(prod, avg_prod)
        bg       = LGRAY if row_idx % 2 == 0 else WHITE

        values = [
            name,
            loc.get("rank_pph", ""),
            flag_str,
            f"${pph:.2f}",
            f"{pph_delta:+.2f}" if pph_delta else "—",
            f"${loc.get('total_sales', 0):,.0f}",
            f"${loc.get('service', 0):,.0f}",
            f"${loc.get('product', 0):,.0f}",
            f"{prod:.1f}%",
            f"{prod_delta:+.1f}" if prod_delta else "—",
            f"{loc.get('ppg', 0):.2f}",
            f"${loc.get('avg_ticket', 0):.2f}",
            loc.get("guests", 0),
            f"{loc.get('prod_hours', 0):.1f}",
            card,
        ]

        fills = [
            None, None, None,
            pph_c, None,
            None, None, None,
            prod_c, None, None,
            None, None, None,
            None,
        ]

        for col, (val, fill) in enumerate(zip(values, fills), start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _fill(fill if fill else bg)
            cell.alignment = _align("left", wrap=(col == len(values)))
            cell.border = _border()
            cell.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 50

    log.debug("Locations sheet built")


# ─── Service Mix sheet ────────────────────────────────────────────────────────

def _build_service_mix(wb, data: dict):
    ws = wb.create_sheet("Service Mix")
    ws.freeze_panes = "A2"

    headers = [
        "Location",
        "Wax Count", "Wax Net $",   "Wax %",
        "Color Net $",               "Color %",
        "Treat Count", "Treat Net $", "Treat %",
        "Prod Hours",
    ]
    widths = [22, 10, 12, 8, 13, 8, 12, 13, 8, 12]

    _header_row(ws, 1, headers, widths)

    locs = data["locations"]
    avg_wax_pct   = sum(l.get("wax_pct",   0) for l in locs) / max(len(locs), 1)
    avg_color_pct = sum(l.get("color_pct", 0) for l in locs) / max(len(locs), 1)
    avg_treat_pct = sum(l.get("treat_pct", 0) for l in locs) / max(len(locs), 1)

    for row_idx, loc in enumerate(locs, start=2):
        bg        = LGRAY if row_idx % 2 == 0 else WHITE
        wax_pct   = loc.get("wax_pct",   0)
        color_pct = loc.get("color_pct", 0)
        treat_pct = loc.get("treat_pct", 0)

        values = [
            loc["loc_name"],
            loc.get("wax_count",   0),
            f"${loc.get('wax',   0):,.0f}",
            f"{wax_pct:.1f}%",
            f"${loc.get('color', 0):,.0f}",
            f"{color_pct:.1f}%",
            loc.get("treat_count", 0),
            f"${loc.get('treat',  0):,.0f}",
            f"{treat_pct:.1f}%",
            f"{loc.get('prod_hours', 0):.1f}",
        ]

        fills = [
            None,
            None, None, _pct_color(wax_pct,   avg_wax_pct),
            None,       _pct_color(color_pct,  avg_color_pct),
            None, None, _pct_color(treat_pct,  avg_treat_pct),
            None,
        ]

        for col, (val, fill) in enumerate(zip(values, fills), start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _fill(fill if fill else bg)
            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()
            cell.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 20

    # Network averages footer
    footer_row = len(locs) + 2
    avg_footer = [
        "NETWORK AVG", "", "",
        f"{avg_wax_pct:.1f}%",
        "",
        f"{avg_color_pct:.1f}%",
        "", "",
        f"{avg_treat_pct:.1f}%",
        "",
    ]
    for col, val in enumerate(avg_footer, start=1):
        cell = ws.cell(row=footer_row, column=col, value=val)
        cell.fill = _fill(SECTION_BG)
        cell.font = Font(bold=True, color=NAVY, size=10)
        cell.alignment = _align("center" if col > 1 else "left")
        cell.border = _border()

    log.debug("Service Mix sheet built")


# ─── Goals & Year-Over-Year sheet (placeholders until API feeds) ──────────────

def _build_goals_yoy(wb, data: dict):
    ws = wb.create_sheet("Goals & YOY")
    ws.freeze_panes = "A3"

    live_headers = ["Location", "PPH $", "Total Sales $", "Guests", "PPG", "Avg Ticket $"]
    live_widths  = [22, 10, 14, 8, 8, 13]

    ph_headers = [
        "2026 Goal $", "vs Goal",
        "2025 Sales $", "2025 Guests", "2025 PPG", "PPG Diff",
        "2025 Avg Tkt", "Avg Tkt Diff",
    ]

    # Row 1: column headers
    for col, (h, w) in enumerate(zip(live_headers, live_widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = _align("center")
        cell.border = _border()

    ph_start = len(live_headers) + 1
    for i, ph in enumerate(ph_headers):
        col = ph_start + i
        ws.column_dimensions[get_column_letter(col)].width = 13
        _placeholder_header(ws, 1, col, ph)

    # Row 2: note spanning placeholder columns
    ws.merge_cells(
        f"{get_column_letter(ph_start)}2"
        f":{get_column_letter(ph_start + len(ph_headers) - 1)}2"
    )
    note = ws.cell(row=2, column=ph_start,
                   value="⏳ Pending: Zenoti YOY API feed + goal config in karissa_001.json")
    note.fill   = _fill(PLACEHOLDER_BG)
    note.font   = Font(italic=True, color=PLACEHOLDER_FG, size=9)
    note.alignment = _align("center")
    note.border = _border()
    # Fill live-column cells in row 2
    for col in range(1, ph_start):
        cell = ws.cell(row=2, column=col)
        cell.fill   = _fill(LGRAY)
        cell.border = _border()

    # Data rows start at row 3
    for row_idx, loc in enumerate(data["locations"], start=3):
        bg = LGRAY if row_idx % 2 == 0 else WHITE

        live_values = [
            loc["loc_name"],
            f"${loc.get('pph', 0):.2f}",
            f"${loc.get('total_sales', 0):,.0f}",
            loc.get("guests", 0),
            f"{loc.get('ppg', 0):.2f}",
            f"${loc.get('avg_ticket', 0):.2f}",
        ]
        for col, val in enumerate(live_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _fill(bg)
            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()
            cell.font = Font(size=10)

        for i in range(len(ph_headers)):
            _placeholder_val(ws, row_idx, ph_start + i)

        ws.row_dimensions[row_idx].height = 20

    log.debug("Goals & YOY sheet built (placeholders for API-pending columns)")


# ─── Stylists sheet ───────────────────────────────────────────────────────────

def _build_stylists(wb, data: dict, ai_cards: dict):
    ws = wb.create_sheet("Stylists")
    ws.freeze_panes = "A2"

    headers = [
        "Stylist", "Location", "Status", "Tenure (yrs)", "Star?",
        "PPH $", "vs Net Avg", "Rebook %", "Product %", "Avg Ticket $", "Coaching Note"
    ]
    col_widths = [22, 18, 10, 12, 8, 10, 10, 10, 10, 12, 50]

    _header_row(ws, 1, headers, col_widths)

    net     = data["network"]
    avg_pph = net.get("avg_pph", 0)

    for row_idx, s in enumerate(data["stylists"], start=2):
        pph     = s.get("cur_pph", 0)
        vs_net  = s.get("vs_net_pph", 0)
        is_star = s.get("is_star", False)
        note    = ai_cards.get("stylist_cards", {}).get(s["name"], "")

        pph_c = _pph_color(pph, avg_pph)
        bg    = LGRAY if row_idx % 2 == 0 else WHITE

        values = [
            s.get("name", ""),
            s.get("loc_name", ""),
            s.get("status", "").upper(),
            s.get("tenure", 0),
            "⭐" if is_star else "",
            f"${pph:.2f}",
            f"{vs_net:+.2f}" if vs_net else "—",
            f"{s.get('cur_rebook', 0):.1f}%",
            f"{s.get('cur_product', 0):.1f}%",
            f"${s.get('cur_ticket', 0):.2f}",
            note,
        ]

        fills = [None, None, None, None, None, pph_c, None, None, None, None, None]

        for col, (val, fill) in enumerate(zip(values, fills), start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _fill(fill if fill else bg)
            cell.alignment = _align("left", wrap=(col == len(values)))
            cell.border = _border()
            cell.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 50

    log.debug("Stylists sheet built")


# ─── Main entry ───────────────────────────────────────────────────────────────

def build_report(data: dict, ai_cards: dict, output_dir: Path, dry_run: bool = False) -> "Path | None":
    """
    Build the Excel report. Returns the Path to the created file, or None on skip.
    """
    if not OPENPYXL_AVAILABLE:
        log.warning("openpyxl not available — skipping report build")
        return None

    week     = data["network"].get("week_ending", datetime.date.today().strftime("%Y-%m-%d"))
    filename = f"KPI_Weekly_Report_{week}.xlsx"
    output_path = output_dir / filename

    if dry_run:
        log.info("DRY RUN: Would create report at %s", output_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = f"[DRY RUN] KPI Report — {week}"
        ws["A2"] = "Placeholder generated in DRY_RUN mode."
        wb.save(output_path)
        log.info("DRY RUN: placeholder report saved to %s", output_path)
        return output_path

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_summary(wb, data, ai_cards)
    _build_locations(wb, data, ai_cards)
    _build_service_mix(wb, data)
    _build_goals_yoy(wb, data)
    _build_stylists(wb, data, ai_cards)

    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Report saved: %s (%.1f KB)", output_path, output_path.stat().st_size / 1024)
    return output_path
